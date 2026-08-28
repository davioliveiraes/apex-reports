# -*- coding: utf-8 -*-
"""
Leitor do export .xlsx do Meta Ads Manager.

Identifica as colunas por palavras-chave (funciona com export em PT ou EN),
consolida os KPIs do período, monta o funil de vendas e os dados dos gráficos
(funil visual e share de resultados por campanha/unidade).
"""

import re
import unicodedata
from dataclasses import asdict
from datetime import date, datetime

from openpyxl import load_workbook

from . import analysis, benchmarks, indicadores


# ----------------------------------------------------------------------
# Helpers
# ----------------------------------------------------------------------
def _norm(texto):
    """minúsculas, sem acento — para casar nomes de coluna."""
    if texto is None:
        return ""
    texto = str(texto).lower().strip()
    return "".join(
        c for c in unicodedata.normalize("NFD", texto)
        if unicodedata.category(c) != "Mn"
    )


def _to_float(v):
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace("R$", "").replace("\xa0", "").strip()
    # formato brasileiro: 1.234,56
    if "," in s:
        s = s.replace(".", "").replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None


def _fmt_moeda(v):
    if v is None:
        return "—"
    return f"R$ {v:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


def _fmt_int(v):
    if v is None:
        return "—"
    return f"{int(round(v)):,}".replace(",", ".")


def _fmt_dec(v, casas=2):
    if v is None:
        return "—"
    return f"{v:.{casas}f}".replace(".", ",")


def _fmt_data(v):
    if isinstance(v, (datetime, date)):
        return v.strftime("%d/%m/%Y")
    if v:
        s = str(v).strip()
        try:
            return datetime.strptime(s[:10], "%Y-%m-%d").strftime("%d/%m/%Y")
        except ValueError:
            return s
    return None


# Mapeamento coluna → (palavras-chave alternativas, termos proibidos no nome)
_COLUNAS = {
    "campanha":    ([["nome da campanha"], ["campaign name"]], []),
    "conjunto":    ([["nome do conjunto"], ["ad set name"]], []),
    "anuncio":     ([["nome do anuncio"], ["ad name"]], []),
    # O Meta alterna o rótulo da coluna de verba entre exports ("Valor usado
    # (BRL)" e "Valor gasto (BRL)"). Sem todas as variantes o relatório sai com
    # investimento e custo por resultado zerados, e o zero passa por número
    # legítimo em vez de coluna não encontrada.
    "investimento": ([["valor usado"], ["valor gasto"], ["valor investido"],
                      ["amount spent"]], []),
    "resultados":  ([["resultados"], ["results"]], ["custo", "cost", "indicador", "indicator", "tipo", "type", "taxa", "rate"]),
    "custo_resultado": ([["custo por resultado"], ["cost per result"]], []),
    "indicador":   ([["indicador de resultado"], ["result indicator"], ["tipo de resultado"], ["result type"]], []),
    "impressoes":  ([["impressoes"], ["impressions"]], ["cpm", "custo", "cost"]),
    "cliques":     ([["cliques no link"], ["link clicks"], ["cliques", "link"]],
                    ["unicos", "unique", "ctr", "custo", "cost", "taxa", "rate"]),
    "alcance":     ([["alcance"], ["reach"]], ["custo", "cost"]),
    "frequencia":  ([["frequencia"], ["frequency"]], []),
    "cpm":         ([["cpm"]], []),
    "inicio":      ([["inicio dos relatorios"], ["reporting starts"]], []),
    "termino":     ([["termino dos relatorios"], ["encerramento dos relatorios"], ["reporting ends"]], []),
    "orcamento":   ([["orcamento"], ["budget"]], []),
}


# O recorte por status de veiculação (todas / ativas / inativas) saiu daqui em
# 24/08/2026, junto com a coluna que ele lia. Ele recortava pelo que a agência
# fez com a campanha; a seleção por grupo, logo abaixo, recorta pelo produto
# anunciado — que é o que o cliente reconhece e o que se quer comparar. Manter
# os dois obrigava a combinar dois filtros para chegar onde um só chega.


# ----------------------------------------------------------------------
# Grupos de campanha (o produto anunciado, lido do nome da campanha)
# ----------------------------------------------------------------------
# Rótulo do grupo das linhas cujo nome de campanha veio em branco. Elas não
# somem no filtro: aparecem como um grupo à parte, para o operador decidir —
# descartar em silêncio verba que não dá para atribuir seria pior.
GRUPO_SEM_NOME = "(campanha sem nome)"

_TOKENS_DO_NOME = re.compile(r"\[([^\]]*)\]")


def chave_grupo_campanha(nome):
    """Grupo a que a campanha pertence — na prática, o produto anunciado.

    O padrão de nome em uso é `[OBJETIVO][PRODUTO][REGIÃO][ESTRUTURA][DATA]`, e
    são os dois primeiros colchetes que se repetem entre as unidades:
    `[LEADS][CELULAR-BOLETO][SALTO][ABO][13JUL26]` e
    `[LEADS][CELULAR-BOLETO][ITU][ABO][01SET25]` caem os dois em
    "LEADS · CELULAR-BOLETO", enquanto `[LEADS][ULTRA][ABO][24JUL26]` fica à
    parte. Região e data ficam de fora de propósito: são justamente o que varia
    entre anexos do mesmo produto, e é por produto que se quer recortar.

    Nome fora do padrão vira grupo dele mesmo — aí o operador escolhe campanha
    por campanha, que é o pior caso aceitável, não um erro.
    """
    tokens = [t.strip() for t in _TOKENS_DO_NOME.findall(str(nome or "")) if t.strip()]
    if len(tokens) >= 2:
        return " · ".join(tokens[:2])
    return str(nome or "").strip() or GRUPO_SEM_NOME


def grupos_de_campanha(registros):
    """Grupos presentes nestes registros, na ordem em que aparecem:
    `[{"chave": ..., "campanhas": [nomes]}]`."""
    grupos = {}
    for r in registros:
        nome = str(r.get("campanha") or "").strip()
        campanhas = grupos.setdefault(chave_grupo_campanha(nome), [])
        if nome and nome not in campanhas:
            campanhas.append(nome)
    return [{"chave": k, "campanhas": v} for k, v in grupos.items()]


def filtrar_campanhas(registros, chaves):
    """Linhas dos grupos de campanha escolhidos.

    `chaves` vazio ou None não filtra nada — é o que faz o fluxo sem seleção
    (um grupo só nos anexos, ou sessão antiga) seguir como sempre foi.
    """
    if not chaves:
        return list(registros)
    escolhidos = set(chaves)
    return [r for r in registros
            if chave_grupo_campanha(r.get("campanha")) in escolhidos]


def _mapear_colunas(header, colunas=None):
    """Retorna {chave: índice}. Prioriza match exato; depois 'contém', respeitando exclusões.

    `colunas` troca o mapa de palavras-chave — é o que deixa o leitor do preset
    VERBA (`parser_verba.py`) reusar esta busca sem duplicá-la. Sem o argumento,
    o mapa é o do export de desempenho, como sempre foi.
    """
    normalizados = [_norm(h) for h in header]
    mapa = {}
    for chave, (alternativas, proibidos) in (colunas or _COLUNAS).items():
        # 1º passe: nome exatamente igual a uma keyword
        for i, nome in enumerate(normalizados):
            if any(len(alt) == 1 and nome == alt[0] for alt in alternativas):
                mapa[chave] = i
                break
        if chave in mapa:
            continue
        # 2º passe: contém todas as keywords e nenhum termo proibido
        for i, nome in enumerate(normalizados):
            if not nome:
                continue
            if any(p in nome for p in proibidos):
                continue
            if any(all(kw in nome for kw in alt) for alt in alternativas):
                mapa[chave] = i
                break
    return mapa


# ----------------------------------------------------------------------
# Leitura principal
# ----------------------------------------------------------------------
def ler_export_meta(arquivo, conta=None, perfil=None, meta_cpa=None):
    """
    Lê o .xlsx exportado do Meta Ads Manager e devolve um dicionário com:
    kpis, metricas_extra, funil, gráficos (funil visual e share por campanha),
    detalhes_campanha, período detectado e a análise sugerida.
    Levanta ValueError com mensagem amigável se o arquivo não for reconhecido.

    `perfil` e `meta_cpa` alimentam a análise do período (ver consolidar).
    """
    registros, mapa = ler_registros(arquivo)
    return consolidar(registros, mapa, conta, perfil, meta_cpa)


def ler_registros(arquivo):
    """
    Linhas de dados do export + mapa de colunas reconhecidas, sem consolidar.

    Separado de `ler_export_meta` para que um mesmo arquivo possa ser
    consolidado mais de uma vez — é o que permite refazer a leitura com outra
    seleção de campanhas na revisão, sem pedir o anexo de novo.
    """
    wb = load_workbook(arquivo, data_only=True, read_only=True)
    ws = wb.active

    linhas = list(ws.iter_rows(values_only=True))
    wb.close()
    if not linhas:
        raise ValueError("A planilha está vazia.")

    # Localiza a linha de cabeçalho (nem sempre é a primeira)
    header_idx, mapa = None, {}
    for i, linha in enumerate(linhas[:10]):
        m = _mapear_colunas(linha)
        if "investimento" in m or "resultados" in m or "impressoes" in m:
            header_idx, mapa = i, m
            break
    if header_idx is None:
        raise ValueError(
            "Não foi possível reconhecer as colunas do export do Meta Ads Manager. "
            "Confira se o arquivo é o .xlsx exportado direto do Gerenciador de Anúncios."
        )

    registros = []
    for linha in linhas[header_idx + 1:]:
        if linha is None or all(c is None or str(c).strip() == "" for c in linha):
            continue
        reg = {}
        for chave, idx in mapa.items():
            valor = linha[idx] if idx < len(linha) else None
            # As colunas de período vêm ora como texto ISO, ora como data
            # tipada, conforme o arquivo que o Meta gerou. Normaliza aqui
            # porque os registros vão para a sessão, que serializa em JSON e
            # não sabe gravar `date` — `_fmt_data` e `_dias_periodo` já leem o
            # texto ISO, então nada além disto muda.
            if isinstance(valor, date):
                valor = valor.strftime("%Y-%m-%d")
            reg[chave] = valor
        # Ignora linhas de total do próprio export
        nome_ref = _norm(reg.get("anuncio") or reg.get("conjunto") or reg.get("campanha"))
        if nome_ref.startswith(("total", "resultados de")):
            continue
        registros.append(reg)

    if not registros:
        raise ValueError("Nenhuma linha de dados encontrada na planilha.")

    return registros, mapa


def consolidar(registros, mapa, conta=None, perfil=None, meta_cpa=None):
    """
    `conta` só identifica a origem no log de indicador não mapeado.

    `perfil` (ver analysis.benchmarks.Perfil) escolhe a faixa de CPA de
    referência e `meta_cpa` a substitui quando a conta já tem meta combinada.
    Ambos são opcionais e ficam nos padrões enquanto não houver onde
    configurá-los por conta.
    """
    # ---- Totais do período ----
    investimento = sum(v for v in (_to_float(r.get("investimento")) for r in registros) if v) or 0.0
    resultados = sum(v for v in (_to_float(r.get("resultados")) for r in registros) if v) or 0.0
    impressoes = sum(v for v in (_to_float(r.get("impressoes")) for r in registros) if v) or 0.0
    alcance = sum(v for v in (_to_float(r.get("alcance")) for r in registros) if v) or 0.0
    cliques = sum(v for v in (_to_float(r.get("cliques")) for r in registros) if v) or 0.0

    custo_resultado = investimento / resultados if resultados else None
    frequencia = impressoes / alcance if alcance else None
    cpm = investimento / impressoes * 1000 if impressoes else None
    # Métricas de meio/fundo de funil — calculadas quando o export não as traz
    ctr = cliques / impressoes * 100 if cliques and impressoes else None
    cpc = investimento / cliques if cliques else None
    taxa_conversao = resultados / cliques * 100 if cliques and resultados else None

    # O indicador da conta é o de maior soma de resultados, não o da primeira
    # linha: uma planilha com campanhas de objetivos diferentes rotularia o
    # relatório pela campanha que por acaso abre o arquivo.
    indicador = indicadores.dominante(registros, para_numero=_to_float)
    indicador_curto = indicadores.rotulo(indicador, conta)

    inicio = next((r.get("inicio") for r in registros if r.get("inicio")), None)
    termino = next((r.get("termino") for r in registros if r.get("termino")), None)
    periodo = ""
    if inicio and termino:
        periodo = f"{_fmt_data(inicio)} a {_fmt_data(termino)}"

    dados = {
        "periodo": periodo,
        # Colunas efetivamente reconhecidas no export — permite distinguir
        # "métrica ausente na planilha" de "métrica igual a zero".
        "_colunas": sorted(mapa),
        "kpis": [
            {"label": "Investimento", "valor": _fmt_moeda(investimento)},
            {"label": indicador_curto, "valor": _fmt_int(resultados)},
            {"label": "Custo / Resultado", "valor": _fmt_moeda(custo_resultado)},
        ],
        "metricas_extra": {
            "Impressões": _fmt_int(impressoes),
            "Alcance": _fmt_int(alcance),
            "Frequência": _fmt_dec(frequencia),
            "CPM": _fmt_moeda(cpm),
        },
        "_num": {  # valores numéricos para a análise automática e consolidação
            "investimento": investimento, "resultados": resultados,
            "custo_resultado": custo_resultado, "cpm": cpm, "frequencia": frequencia,
            "impressoes": impressoes, "alcance": alcance,
            "cliques": cliques, "ctr": ctr, "cpc": cpc, "taxa_conversao": taxa_conversao,
        },
    }

    # ---- Funil de vendas (topo / meio / fundo) ----
    dados["funil"] = {"etapas": _montar_funil(dados["_num"], indicador)}
    dados["indicador"] = indicador

    # ---- Funil visual (gráfico de barras decrescentes) ----
    dados["grafico_funil"] = _dados_grafico_funil(dados["_num"], indicador)

    # ---- Desempenho por campanha ----
    campanhas = {}
    for r in registros:
        nome = str(r.get("campanha") or "").strip()
        if not nome:
            continue
        c = campanhas.setdefault(nome, {"res": 0.0, "inv": 0.0, "imp": 0.0, "alc": 0.0})
        c["res"] += _to_float(r.get("resultados")) or 0
        c["inv"] += _to_float(r.get("investimento")) or 0
        c["imp"] += _to_float(r.get("impressoes")) or 0
        c["alc"] += _to_float(r.get("alcance")) or 0
    if campanhas:
        # Sem coluna de Status: se a campanha estava ativa ou parada é decisão
        # interna da agência, e não aparece no relatório do cliente.
        dados["detalhes_campanha"] = {
            "titulo": "Desempenho por Campanha",
            "header": ["Campanha", "Resultados", "Investimento", "Custo/Resultado"],
            "linhas": [
                [nome, _fmt_int(c["res"]), _fmt_moeda(c["inv"]),
                 _fmt_moeda(c["inv"] / c["res"] if c["res"] else None)]
                for nome, c in campanhas.items()
            ],
            "legenda": f"Indicador de resultado: {indicador_curto}." if indicador else None,
        }
        # Rosca com o share de resultados por campanha (melhor = menor custo, em verde)
        dados["grafico_campanhas"] = _dados_grafico_campanhas(campanhas, resultados)

    # ---- Análise do Período sugerida (editável na revisão) ----
    # O motor classifica o período e escolhe o próximo passo; o texto é
    # derivado dessa decisão, não o contrário. A avaliação fica guardada em
    # `dados` porque é ela — não o texto — que vira payload nas etapas
    # seguintes (prompt de IA, mensagem de WhatsApp), sem reprocessar números.
    # `_metricas` e `_dias` ficam junto para a análise poder ser recalculada
    # sem pedir o anexo de novo.
    dados["_metricas"] = _metricas_analise(dados["_num"], campanhas)
    dados["_dias"] = _dias_periodo(inicio, termino)
    dados["_perfil"] = perfil
    # Os brutos por campanha, para o payload do redator de IA montar o recorte
    # sem reler a planilha nem desformatar o "R$ 225,75" da tabela.
    dados["_campanhas"] = campanhas
    regerar_analise(dados, meta_cpa=meta_cpa)

    return dados


def regerar_analise(dados, *, meta_cpa=None, contexto=None):
    """
    Recalcula avaliação e texto a partir do que já está em `dados`.

    Chamada na leitura do anexo, e disponível para recalcular depois com meta
    ou contexto informados. Não relê a planilha: tudo que o motor precisa foi
    guardado na primeira passagem.
    """
    if dados.get("modo") == "grupo":
        avaliacao = analysis.rules.avaliar_grupo(
            [{"nome": u["nome"], "metricas": _metricas_unidade(u["num"])}
             for u in dados["unidades"]],
            _metricas_unidade(dados["_num"]),
            perfil=dados.get("_perfil"), meta_cpa=meta_cpa, contexto=contexto,
            dias_periodo=dados.get("_dias"))
        texto = analysis.templates.redigir_grupo(
            avaliacao, dados["_num"], destino="pdf")
    else:
        metricas = dados["_metricas"]
        avaliacao = analysis.rules.avaliar(
            metricas, perfil=dados.get("_perfil"), meta_cpa=meta_cpa,
            contexto=contexto, dias_periodo=dados.get("_dias"))
        texto = analysis.templates.redigir(avaliacao, metricas, destino="pdf")

    dados["avaliacao"] = asdict(avaliacao)
    dados["analise_sugerida"] = texto
    return dados


def _metricas_unidade(n):
    """Totais no formato do motor. Sem campanhas: no consolidado a unidade é a
    unidade de análise, e a estrutura interna dela não está na sessão."""
    return dict(n, cpa=n.get("custo_resultado"))


def _metricas_analise(n, campanhas):
    """Entrada do motor de análise: os totais da conta mais as campanhas na
    forma de lista, que é o que `rules.avaliar` espera para medir estrutura e
    concentração."""
    return dict(n, cpa=n.get("custo_resultado"),
                campanhas=[{"nome": nome, "resultados": c["res"],
                            "investimento": c["inv"]}
                           for nome, c in campanhas.items()])


def _dias_periodo(inicio, termino):
    """Dias corridos do período, inclusive as duas pontas. None quando o
    export não trouxe as datas — aí a frequência usa a janela mensal cheia."""
    datas = []
    for v in (inicio, termino):
        if isinstance(v, datetime):
            v = v.date()
        if not isinstance(v, date):
            try:
                v = datetime.strptime(str(v).strip()[:10], "%Y-%m-%d").date()
            except (ValueError, TypeError):
                return None
        datas.append(v)
    return (datas[1] - datas[0]).days + 1 if datas[1] >= datas[0] else None


def _dados_grafico_funil(n, indicador):
    """Estágios do funil visual: Alcance → Cliques → Conversas (só os presentes)."""
    indicador_curto = indicadores.rotulo(indicador, avisar=False)
    estagios = []
    if n.get("alcance"):
        estagios.append({"rotulo": "Alcance", "valor": n["alcance"],
                         "texto": f"{_fmt_int(n['alcance'])} pessoas"})
    elif n.get("impressoes"):
        estagios.append({"rotulo": "Impressões", "valor": n["impressoes"],
                         "texto": f"{_fmt_int(n['impressoes'])} visualizações"})
    if n.get("cliques"):
        estagios.append({"rotulo": "Cliques no Link", "valor": n["cliques"],
                         "texto": f"{_fmt_int(n['cliques'])} cliques"})
    if n.get("resultados"):
        estagios.append({"rotulo": indicador_curto, "valor": n["resultados"],
                         "texto": _fmt_int(n["resultados"])})
    return estagios if len(estagios) >= 2 else []


def _dados_grafico_campanhas(campanhas, resultados_total):
    """Itens da rosca de share por campanha; a de menor custo é o destaque."""
    com_res = [(nome, c) for nome, c in campanhas.items() if c.get("res")]
    if len(com_res) < 2 or not resultados_total:
        return []
    nome_melhor = min(com_res, key=lambda kv: kv[1]["inv"] / kv[1]["res"])[0]
    itens = sorted(com_res, key=lambda kv: -kv[1]["res"])
    return [{
        "nome": nome,
        "valor": c["res"],
        "texto": _fmt_int(c["res"]),
        "share": c["res"] / resultados_total * 100,
        "melhor": nome == nome_melhor,
    } for nome, c in itens]


def _montar_funil(n, indicador):
    """Etapas do funil (topo/meio/fundo) a partir dos totais numéricos `n`."""
    eh_conversa = indicadores.eh_conversa(indicador)
    indicador_curto = indicadores.rotulo(indicador, avisar=False)
    rotulo_custo = "Custo por Conversa (CPA)" if eh_conversa else "Custo por Resultado (CPA)"
    av = benchmarks.avaliar_metricas(n)

    topo = [["Investimento Total", _fmt_moeda(n["investimento"]),
             "Verba total aplicada em mídia no período."]]
    if n["alcance"]:
        topo.append(["Alcance", f"{_fmt_int(n['alcance'])} pessoas",
                     "Pessoas únicas alcançadas pelos anúncios."])
    if n["impressoes"]:
        topo.append(["Impressões", f"{_fmt_int(n['impressoes'])} visualizações",
                     "Total de vezes que os anúncios foram exibidos."])
    if n["frequencia"]:
        topo.append(["Frequência", _fmt_dec(n["frequencia"]),
                     _leitura_metrica("frequencia", av)])
    if n["cpm"]:
        topo.append(["CPM (custo por mil)", _fmt_moeda(n["cpm"]),
                     _leitura_metrica("cpm", av)])

    meio = []
    if n["cliques"]:
        meio.append(["Cliques no Link", f"{_fmt_int(n['cliques'])} cliques",
                     "Cliques direcionados ao destino (WhatsApp / página)."])
        if n["ctr"] is not None:
            meio.append(["CTR (taxa de cliques)", f"{_fmt_dec(n['ctr'])}%",
                         _leitura_metrica("ctr", av)])
        if n["cpc"] is not None:
            meio.append(["CPC (custo por clique)", _fmt_moeda(n["cpc"]),
                         _leitura_metrica("cpc", av)])

    fundo = []
    if n["resultados"]:
        fundo.append([indicador_curto, _fmt_int(n["resultados"]),
                      "Total de conversões registradas no período."])
    if n["custo_resultado"] is not None:
        fundo.append([rotulo_custo, _fmt_moeda(n["custo_resultado"]),
                      "Quanto custou, em média, cada conversão."])
    if n["taxa_conversao"] is not None:
        fundo.append(["Taxa de Conversão (clique → conversa)",
                      f"{_fmt_dec(n['taxa_conversao'])}%",
                      _leitura_metrica("taxa_conversao", av)])

    return [
        {"titulo": t, "linhas": l} for t, l in [
            ("Topo de Funil — Atração", topo),
            ("Meio de Funil — Interesse e Clique", meio),
            ("Fundo de Funil — Conversão", fundo),
        ] if l
    ]


# ----------------------------------------------------------------------
# Consolidação multi-contas (grupo de unidades)
# ----------------------------------------------------------------------
def consolidar_grupo(unidades):
    """
    Consolida os dados de 2+ contas/unidades num relatório único de grupo.

    `unidades`: lista de {"nome": str, "dados": dict retornado por ler_export_meta}
    (o dict de cada unidade precisa ainda conter a chave "_num").

    Regra de agregação: soma investimento/impressões/alcance/cliques/resultados
    e recalcula as taxas sobre os totais — nunca média simples de percentuais.
    """
    us = []
    for u in unidades:
        d = u["dados"]
        us.append({
            "nome": u["nome"],
            "num": d["_num"],
            "funil": d.get("funil"),
            "kpis": d.get("kpis", []),
            "indicador": d.get("indicador", ""),
            "periodo": d.get("periodo", ""),
        })

    n = _totais_grupo(us)
    # Mesma regra da conta individual, um nível acima: o indicador do grupo é
    # o das unidades que respondem pela maior parte dos resultados.
    indicador = indicadores.dominante(
        [{"indicador": u["indicador"], "resultados": u["num"].get("resultados")}
         for u in us])
    eh_conversa = indicadores.eh_conversa(indicador)
    indicador_curto = indicadores.rotulo(indicador, avisar=False)

    funil = {"etapas": _montar_funil(n, indicador)}
    if n["alcance"]:
        funil["legenda"] = ("Alcance geral: soma das unidades — pode haver "
                            "sobreposição de audiência entre as contas.")

    dados = {
        "modo": "grupo",
        "periodo": _periodo_grupo(us),
        "kpis": [
            {"label": "Investimento", "valor": _fmt_moeda(n["investimento"])},
            {"label": indicador_curto, "valor": _fmt_int(n["resultados"])},
            {"label": "Custo / Resultado", "valor": _fmt_moeda(n["custo_resultado"])},
        ],
        "funil": funil,
        "grafico_funil": _dados_grafico_funil(n, indicador),
        "composicao": montar_composicao(us),
        "unidades": us,
    }
    # Aviso de indicador divergente: só na tela de revisão — não vai ao PDF.
    aviso = _aviso_indicador(us)
    if aviso:
        dados["aviso_indicador"] = aviso
    # ---- Análise do Período do grupo (editável na revisão) ----
    # Aqui o motor faz o que só o consolidado permite: mede cada unidade
    # contra o CPA do próprio grupo, no mesmo período. É a única referência do
    # sistema que não é estimativa nossa — e é ela que aponta qual praça está
    # cara e qual tem o método que vale copiar.
    dados["_num"] = n
    dados["_dias"] = _dias_periodo_texto(dados["periodo"])
    regerar_analise(dados)
    return dados


def _dias_periodo_texto(periodo):
    """Dias corridos a partir do "dd/mm/aaaa a dd/mm/aaaa" já montado para o
    grupo. None quando as unidades não trouxeram datas."""
    try:
        ini, fim = (datetime.strptime(x.strip(), "%d/%m/%Y")
                    for x in (periodo or "").split(" a "))
    except ValueError:
        return None
    return (fim - ini).days + 1 if fim >= ini else None


def _totais_grupo(us):
    """Totais do grupo somando as unidades; taxas recalculadas sobre os totais."""
    def soma(chave):
        return sum(v for v in (u["num"].get(chave) for u in us) if v) or 0.0

    investimento = soma("investimento")
    resultados = soma("resultados")
    impressoes = soma("impressoes")
    alcance = soma("alcance")
    cliques = soma("cliques")
    return {
        "investimento": investimento, "resultados": resultados,
        "impressoes": impressoes, "alcance": alcance, "cliques": cliques,
        "custo_resultado": investimento / resultados if resultados else None,
        "frequencia": impressoes / alcance if alcance else None,
        "cpm": investimento / impressoes * 1000 if impressoes else None,
        "ctr": cliques / impressoes * 100 if cliques and impressoes else None,
        "cpc": investimento / cliques if cliques else None,
        "taxa_conversao": resultados / cliques * 100 if cliques and resultados else None,
    }


def montar_composicao(unidades):
    """
    Participação de cada unidade no total de resultados do grupo — dados do
    gráfico de composição (barras horizontais). Público para permitir remontar
    com os nomes de unidade editados na revisão.
    """
    total = sum(u["num"].get("resultados") or 0 for u in unidades)
    ordenadas = sorted(unidades, key=lambda u: -(u["num"].get("resultados") or 0))
    itens = [{
        "nome": u["nome"],
        "valor": u["num"].get("resultados") or 0,
        "texto": _fmt_int(u["num"].get("resultados") or 0),
        "share": (u["num"].get("resultados") or 0) / total * 100 if total else 0,
    } for u in ordenadas]
    return {"titulo": "Participação de Cada Unidade nos Resultados", "itens": itens}


def _aviso_indicador(us):
    """Texto de aviso quando os anexos não usam o mesmo indicador de resultado."""
    grupos = {}
    for u in us:
        if u.get("indicador"):
            # Agrupado pelo rótulo legível: o aviso é para o operador decidir,
            # não para ele decifrar "actions:post_engagement".
            grupos.setdefault(indicadores.rotulo(u["indicador"], avisar=False),
                              []).append(u["nome"])
    if len(grupos) <= 1:
        return None
    partes = "; ".join(f'"{ind}" ({", ".join(nomes)})' for ind, nomes in grupos.items())
    return (
        "Os anexos não usam o mesmo indicador de resultado — "
        f"{partes}. Os totais do grupo somam métricas diferentes; "
        "confira se a comparação faz sentido antes de gerar o PDF."
    )


def _periodo_grupo(us):
    """Período do grupo: do menor início ao maior fim entre as unidades."""
    datas = []
    for u in us:
        try:
            ini, fim = (datetime.strptime(x.strip(), "%d/%m/%Y")
                        for x in (u.get("periodo") or "").split(" a "))
            datas.append((ini, fim))
        except ValueError:
            continue
    if datas:
        ini = min(d[0] for d in datas)
        fim = max(d[1] for d in datas)
        return f"{ini:%d/%m/%Y} a {fim:%d/%m/%Y}"
    return next((u["periodo"] for u in us if u.get("periodo")), "")


# ----------------------------------------------------------------------
# Leituras automáticas das métricas do funil (1 linha por card no PDF)
# ----------------------------------------------------------------------
# Enquadramento sempre positivo ou neutro: a classificação de benchmark é
# interna; para o cliente, a agência aparece em ação contínua.
_LEITURAS_CARD = {
    "frequencia": {
        benchmarks.ABAIXO: "Frequência saudável — público longe da saturação.",
        benchmarks.DENTRO: "Frequência saudável — público longe da saturação.",
        benchmarks.ACIMA: ("Boa presença junto ao público — estamos ampliando "
                           "as audiências para manter a entrega eficiente."),
    },
    "cpm": {
        benchmarks.ABAIXO: ("Custo de entrega competitivo — bom momento para "
                            "ganhar volume."),
        benchmarks.DENTRO: "Custo de entrega competitivo.",
        benchmarks.ACIMA: ("Público concorrido — otimizamos a entrega para "
                           "manter o custo sob controle."),
    },
    "ctr": {
        benchmarks.ABAIXO: ("Estamos renovando os criativos para elevar a "
                            "taxa de cliques."),
        benchmarks.DENTRO: "Taxa de cliques dentro da faixa esperada.",
        benchmarks.ACIMA: "Anúncios com ótima atratividade para o público.",
    },
    "cpc": {
        benchmarks.ABAIXO: "Custo por clique competitivo.",
        benchmarks.DENTRO: "Custo por clique dentro da faixa esperada.",
        benchmarks.ACIMA: ("Estamos refinando públicos e entrega para reduzir "
                           "o custo por clique."),
    },
    "taxa_conversao": {
        benchmarks.ABAIXO: ("Estamos alinhando o fluxo de atendimento para "
                            "aproveitar melhor cada clique."),
        benchmarks.DENTRO: "Boa eficiência do fluxo de atendimento.",
        benchmarks.ACIMA: "Boa eficiência do fluxo de atendimento.",
        benchmarks.EXCELENTE: ("Excelente eficiência — parte das conversas "
                               "chega direto do anúncio, sem depender do clique."),
    },
}


def _leitura_metrica(metrica, avaliacao):
    """Leitura curta da métrica conforme a classificação de benchmark."""
    classe = avaliacao.get(metrica)
    return _LEITURAS_CARD[metrica].get(classe, "") if classe else ""


# Rótulo da linha do funil (como `_montar_funil` escreve) → chave que
# `redator_ia.gerar_leituras_funil` usa no JSON. Só as 4 métricas que têm
# leitura no PDF (`gerador_pdf._METRICAS_LEITURA`) — CPC fica de fora dos
# dois lados.
_ROTULO_METRICA_FUNIL = {
    "Frequência": "frequencia",
    "CPM (custo por mil)": "cpm",
    "CTR (taxa de cliques)": "ctr",
    "Taxa de Conversão (clique → conversa)": "taxa_conversao",
}


def substituir_leituras(funil, leituras):
    """Troca, nas linhas do funil, a leitura estática pela da IA.

    `leituras`: dict como o de `redator_ia.gerar_leituras_funil` — chave
    ausente ou não reconhecida simplesmente não mexe naquela linha, que fica
    com o texto do catálogo que `_montar_funil` já tinha posto. Sem retorno:
    `funil` é mutado no lugar, mesmo dict que já está em `dados["funil"]`.
    """
    if not funil or not leituras:
        return
    for etapa in funil.get("etapas", []):
        linhas = etapa.get("linhas", [])
        for i, linha in enumerate(linhas):
            chave = _ROTULO_METRICA_FUNIL.get(linha[0])
            nova = leituras.get(chave) if chave else None
            if nova:
                linhas[i] = [linha[0], linha[1], nova]


# A Análise do Período — de conta e de grupo — saiu daqui para `analysis/`:
# motor de regras em vez de paráfrase dos números. Saíram junto as funções que
# montavam aquele texto por composição (_frases_mencoes, _frase_resumo,
# _frases_continuidade e o catálogo _CONTINUIDADE), sem uso desde então.
