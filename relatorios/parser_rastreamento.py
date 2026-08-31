# -*- coding: utf-8 -*-
"""
Leitor do export do preset `RASTREAMENTO` do Gerenciador de Anúncios.

Terceiro parser do projeto, pelo mesmo motivo dos outros dois: outro recorte
de colunas, com outras colisões. Aqui as armadilhas são os **pares
único/não-único** — `Cliques no link` × `Cliques no link únicos`, `CTR` × `CTR
único` — e o `(BRL)` que o Meta cola no fim de tudo que é dinheiro.

Camada de aliases centralizada
------------------------------
`_COLUNAS_RASTREAMENTO` é o **único** lugar do projeto que conhece a grafia do
Meta para este preset. Da saída deste módulo para a frente, a regra de negócio
lê `link_ctr` e `video_50`, nunca "CTR (taxa de cliques no link)". Trocar o
nome de uma coluna no Gerenciador é editar uma linha aqui.

O que este preset NÃO traz
--------------------------
Resultados, conversas, contatos, vendas, ROAS, impressões e alcance. As duas
últimas são **recuperáveis** (ver `metricas.py`); as outras não existem e
nenhuma frase pode depender delas. É a diferença entre esta frente e a Análise
de Desempenho: lá se mede o que a campanha produziu, aqui o caminho que o
usuário percorreu até sumir.

Tolerância a coluna ausente
---------------------------
Diferente dos outros dois presets, aqui **não há lista de colunas
obrigatórias**. Um anúncio de imagem não tem métrica de vídeo; uma campanha de
WhatsApp não tem página de destino; um anúncio novo não tem classificação de
relevância. Recusar o arquivo por isso seria recusar o caso normal. A regra é
outra: o arquivo passa se der para montar **pelo menos um** dos quatro blocos
(ver `blocos_possiveis`).
"""

import re
import unicodedata
from datetime import date

from openpyxl import load_workbook

from .parser_xlsx import _mapear_colunas, _norm

# Mesmo formato de `parser_xlsx._COLUNAS`: chave → (alternativas, proibidos).
# `_mapear_colunas` casa nome exato antes de "contém", que é o que separa
# `Cliques no link` de `Cliques no link únicos` sem depender da ordem.
_COLUNAS_RASTREAMENTO = {
    # -------------------------------------------------------- contexto
    "report_start": ([["inicio dos relatorios"], ["reporting starts"]], []),
    "report_end":   ([["encerramento dos relatorios"],
                      ["termino dos relatorios"], ["reporting ends"]], []),
    "campaign_name": ([["nome da campanha"], ["campaign name"]], []),
    "adset_name":    ([["nome do conjunto"], ["ad set name"]], []),
    "ad_name":       ([["nome do anuncio"], ["ad name"]], []),
    "delivery":      ([["veiculacao"], ["delivery"]], []),

    # -------------------------------------------------------- clique
    # "unico" afasta o par único de cada métrica. Sem isso o passe por
    # "contém" pegaria a primeira que casasse, e qual é a primeira depende da
    # ordem em que o operador arrastou as colunas no personalizador.
    "link_clicks": ([["cliques no link"], ["link clicks"]],
                    ["unico", "unique", "ctr", "cpc", "custo", "cost", "taxa",
                     "rate"]),
    "unique_link_clicks": ([["cliques no link unicos"], ["cliques unicos"],
                            ["unique link clicks"]],
                           ["ctr", "cpc", "custo", "cost", "taxa", "rate"]),
    "link_ctr": ([["ctr", "cliques no link"], ["ctr", "link click"]],
                 ["unico", "unique"]),
    "unique_link_ctr": ([["ctr", "unico"], ["ctr", "unique"]], []),
    "link_cpc": ([["cpc"], ["custo por clique no link"],
                  ["cost per link click"]], ["unico", "unique"]),

    # -------------------------------------------------------- destino
    "landing_page_views": ([["visualizacoes da pagina de destino"],
                            ["landing page views"]],
                           ["custo", "cost", "taxa", "rate"]),
    "cost_per_landing_page_view": ([["custo", "pagina de destino"],
                                    ["cost per landing page"]], []),
    "attribution_setting": ([["configuracao de atribuicao"],
                             ["attribution setting"]], []),

    # -------------------------------------------------------- relevância
    "quality_ranking": ([["classificacao de qualidade"],
                         ["quality ranking"]], []),
    "engagement_rate_ranking": ([["classificacao", "engajamento"],
                                 ["engagement rate ranking"]], []),
    "conversion_rate_ranking": ([["classificacao", "conversao"],
                                 ["conversion rate ranking"]], []),

    # -------------------------------------------------------- vídeo
    "video_3s_views": ([["3 segundos"], ["3-second"], ["3 second"]], []),
    "thruplays": ([["thruplays"], ["thruplay"]], ["custo", "cost"]),
    "cost_per_thruplay": ([["custo", "thruplay"], ["cost per thruplay"]], []),
    "video_25":  ([["25%", "video"], ["25%", "reproduc"]], ["custo", "cost"]),
    "video_50":  ([["50%", "video"], ["50%", "reproduc"]], ["custo", "cost"]),
    "video_75":  ([["75%", "video"], ["75%", "reproduc"]], ["custo", "cost"]),
    "video_100": ([["100%", "video"], ["100%", "reproduc"]], ["custo", "cost"]),
}

# Campos que a regra de negócio lê como número.
NUMERICOS = ("link_clicks", "unique_link_clicks", "link_ctr",
             "unique_link_ctr", "link_cpc", "landing_page_views",
             "cost_per_landing_page_view", "video_3s_views", "thruplays",
             "cost_per_thruplay", "video_25", "video_50", "video_75",
             "video_100")

# Taxas — chegam em UNIDADE DE PERCENTUAL (0,58 significa 0,58%), conferido
# contra o export real: 582 cliques com CTR 0,58193 reconstroem exatamente as
# 100.012 impressões que o export de desempenho da mesma conta declara.
PERCENTUAIS = ("link_ctr", "unique_link_ctr")

# Classificações do Meta. Categóricas de propósito: viram texto, nunca número
# (ver §12 — converter "Acima da média" em 3 seria inventar uma escala).
RANKINGS = ("quality_ranking", "engagement_rate_ranking",
            "conversion_rate_ranking")

# Como o Meta escreve "não se aplica" numa célula. Tudo isto vira None, e não
# zero: um anúncio de imagem com 0 ThruPlays é diferente de um anúncio de
# imagem que não tem a métrica.
VAZIOS = frozenset(("", "-", "--", "---", "—", "–", "n/a", "na", "nan",
                    "nao disponivel", "not available", "sem dados"))

# Os quatro blocos do diagnóstico.
BLOCO_CLIQUE = "clique"
BLOCO_DESTINO = "destino"
BLOCO_RELEVANCIA = "relevancia"
BLOCO_RETENCAO = "retencao"

# O que cada bloco precisa para existir. Basta UMA das métricas listadas ter
# valor em alguma linha — a lista é de suficiência, não de obrigatoriedade.
EXIGENCIAS = {
    BLOCO_CLIQUE: ("link_clicks", "unique_link_clicks", "link_ctr",
                   "unique_link_ctr", "link_cpc"),
    BLOCO_DESTINO: ("landing_page_views", "cost_per_landing_page_view"),
    BLOCO_RELEVANCIA: RANKINGS,
    BLOCO_RETENCAO: ("video_3s_views", "thruplays", "cost_per_thruplay",
                     "video_25", "video_50", "video_75", "video_100"),
}

# Nome de cada métrica para a tela de erro (§30: "Métricas esperadas").
ROTULOS = {
    "link_clicks": "Cliques no link",
    "unique_link_clicks": "Cliques no link únicos",
    "link_ctr": "CTR (taxa de cliques no link)",
    "unique_link_ctr": "CTR único (taxa de cliques no link)",
    "link_cpc": "CPC (custo por clique no link)",
    "landing_page_views": "Visualizações da página de destino",
    "cost_per_landing_page_view": "Custo por visualização da página de destino",
    "attribution_setting": "Configuração de atribuição",
    "quality_ranking": "Classificação de qualidade",
    "engagement_rate_ranking": "Classificação da taxa de engajamento",
    "conversion_rate_ranking": "Classificação da taxa de conversão",
    "video_3s_views": "Reproduções de vídeo por no mínimo 3 segundos",
    "thruplays": "ThruPlays",
    "cost_per_thruplay": "Custo por ThruPlay",
    "video_25": "Reproduções de 25% do vídeo",
    "video_50": "Reproduções de 50% do vídeo",
    "video_75": "Reproduções de 75% do vídeo",
    "video_100": "Reproduções de 100% do vídeo",
}

_SO_NUMERO = re.compile(r"[^\d,.\-]")


class ErroDePreset(ValueError):
    """Arquivo lido, métricas insuficientes — com os dois lados da conta.

    Carrega o que foi encontrado e o que era esperado, para a tela poder
    mostrar as duas listas lado a lado (§30) em vez de uma frase genérica.
    """

    def __init__(self, mensagem, encontradas=(), esperadas=()):
        super().__init__(mensagem)
        self.encontradas = list(encontradas)
        self.esperadas = list(esperadas)


def numero(valor):
    """Célula → float, ou `None` quando não há número ali.

    Aceita o que o Meta realmente exporta: número tipado, "1.234,56",
    "R$ 4,52", "0,58%", "--", "—", "N/A" e célula vazia. O `%` é removido sem
    dividir por 100 — as taxas deste preset já vêm em unidade de percentual.
    """
    if valor is None:
        return None
    if isinstance(valor, bool):
        return None
    if isinstance(valor, (int, float)):
        # NaN não é igual a si mesmo. Chega aqui quando a planilha foi salva
        # por uma ferramenta que escreveu #DIV/0! como float.
        return None if valor != valor else float(valor)

    texto = str(valor).strip()
    if _norm(texto) in VAZIOS:
        return None

    texto = _SO_NUMERO.sub("", texto)
    if not texto or texto in ("-", ",", "."):
        return None
    # Formato brasileiro: a vírgula é o decimal e o ponto é o milhar.
    if "," in texto:
        texto = texto.replace(".", "").replace(",", ".")
    try:
        return float(texto)
    except ValueError:
        return None


def texto(valor):
    """Célula → string limpa, ou `None`. Para nome, veiculação e ranking."""
    if valor is None:
        return None
    limpo = str(valor).strip()
    return None if _norm(limpo) in VAZIOS else limpo


def _sem_acento(valor):
    return "".join(c for c in unicodedata.normalize("NFD", str(valor or "").lower())
                   if unicodedata.category(c) != "Mn")


def abaixo_da_media(ranking):
    """A classificação do Meta sinaliza abaixo da média?

    Casa por trecho porque a grafia varia entre exports ("Abaixo da média",
    "Abaixo da média (20% inferiores)", "Below average"). O valor original é
    preservado em toda a saída — isto aqui só responde a pergunta.
    """
    if not ranking:
        return False
    n = _sem_acento(ranking)
    return "abaixo da media" in n or "below average" in n


def ler_planilha_rastreamento(arquivo):
    """`(linhas, disponiveis)` — as linhas normalizadas e o que o arquivo tem.

    `disponiveis` é o conjunto de chaves internas que apareceram **com valor**
    em alguma linha; ter a coluna e não ter dado nenhum não conta, porque é
    exatamente o caso do export de WhatsApp, que traz as colunas de página de
    destino sempre vazias.
    """
    wb = load_workbook(arquivo, data_only=True, read_only=True)
    ws = wb.active
    brutas = list(ws.iter_rows(values_only=True))
    wb.close()
    if not brutas:
        raise ValueError("A planilha está vazia.")

    cabecalho, mapa = None, {}
    for i, linha in enumerate(brutas[:10]):
        m = _mapear_colunas(linha, _COLUNAS_RASTREAMENTO)
        # Qualquer métrica de rastreamento serve para achar o cabeçalho — não
        # dá para exigir uma específica num preset em que quase tudo é
        # opcional.
        if any(c in m for c in
               ("link_clicks", "link_ctr", "landing_page_views", "thruplays",
                "quality_ranking", "video_25")):
            cabecalho, mapa = i, m
            break
    if cabecalho is None:
        raise ErroDePreset(
            "Não foi possível reconhecer nenhuma métrica de rastreamento "
            "nesta planilha. Aplique a predefinição RASTREAMENTO em "
            "Colunas → Personalizar colunas antes de exportar.",
            esperadas=[ROTULOS[c] for c in EXIGENCIAS[BLOCO_CLIQUE]])

    linhas, disponiveis = [], set()
    for bruta in brutas[cabecalho + 1:]:
        if bruta is None or all(c is None or str(c).strip() == ""
                                for c in bruta):
            continue
        reg = {}
        for chave, idx in mapa.items():
            valor = bruta[idx] if idx < len(bruta) else None
            if isinstance(valor, date):
                # A sessão serializa em JSON e não sabe gravar `date`.
                valor = valor.strftime("%Y-%m-%d")
            reg[chave] = (numero(valor) if chave in NUMERICOS
                          else texto(valor))
            if reg[chave] is not None:
                disponiveis.add(chave)
        # Linha de total do próprio export.
        if _norm(reg.get("ad_name") or reg.get("adset_name")
                 or reg.get("campaign_name")).startswith(
                     ("total", "resultados de")):
            continue
        linhas.append(reg)

    if not linhas:
        raise ValueError("Nenhuma linha de dados encontrada na planilha.")

    possiveis = blocos_possiveis(disponiveis)
    if not possiveis:
        raise ErroDePreset(
            "Este arquivo não contém métricas suficientes para uma Análise "
            "de Rastreamento.",
            encontradas=sorted(ROTULOS[c] for c in disponiveis
                               if c in ROTULOS),
            esperadas=[ROTULOS[c] for c in
                       EXIGENCIAS[BLOCO_CLIQUE] + EXIGENCIAS[BLOCO_DESTINO]])
    return linhas, disponiveis


def blocos_possiveis(disponiveis):
    """Quais dos quatro blocos o arquivo sustenta.

    É esta função que substitui a lista de colunas obrigatórias dos outros
    dois presets: o arquivo é válido se sustentar **um** bloco.
    """
    return [bloco for bloco, campos in EXIGENCIAS.items()
            if any(c in disponiveis for c in campos)]


def ler_arquivo_rastreamento(arquivo):
    """`(linhas, disponiveis, erro, encontradas, esperadas)`.

    A view não trata exceção: recebe os dados ou a frase pronta com as duas
    listas que a tela precisa mostrar.
    """
    try:
        linhas, disponiveis = ler_planilha_rastreamento(arquivo)
        return linhas, disponiveis, None, [], []
    except ErroDePreset as e:
        return None, set(), f'Arquivo "{arquivo.name}": {e}', \
            e.encontradas, e.esperadas
    except ValueError as e:
        return None, set(), f'Arquivo "{arquivo.name}": {e}', [], []
    except Exception:
        # O engano provável não é arquivo corrompido — é o export do preset
        # errado. Os quatro saem do mesmo Gerenciador, para o mesmo cliente,
        # e todos abrem sem reclamar.
        return None, set(), (
            f'Não foi possível ler "{arquivo.name}". Confira se é o .xlsx '
            "exportado do Gerenciador de Anúncios com a predefinição "
            "RASTREAMENTO — os exports dos presets DESEMPENHO e VERBA não "
            "servem aqui."), [], []
