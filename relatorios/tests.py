# -*- coding: utf-8 -*-
"""
Testes do fluxo web (upload → revisão → PDF), cobrindo o modo individual
(1 anexo) e o consolidado (2 a 20 anexos) no layout dark de UMA página
(WeasyPrint). A contagem de páginas usa `pdfinfo` quando disponível e
PyMuPDF como fallback.
"""
import io
import json
import os
import re
import shutil
import subprocess
import sys
import tempfile
from datetime import date
from pathlib import Path
from types import SimpleNamespace
from unittest.mock import MagicMock, patch

import fitz  # PyMuPDF — extração de texto e contagem de páginas
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import SimpleTestCase, TestCase, override_settings
from openpyxl import Workbook

from . import analysis, benchmarks, metricas, parser_xlsx, redator_ia
from .analysis import templates
from .parser_xlsx import consolidar_grupo, ler_export_meta
from .views import _MESES_PT, _paragrafos as _paragrafos_analise

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

CABECALHO = [
    "Nome da campanha", "Veiculação da campanha", "Resultados",
    "Indicador de resultado", "Valor usado (BRL)", "Impressões", "Alcance",
    "Cliques no link", "Início dos relatórios", "Término dos relatórios",
]

INDICADOR = "Conversas por mensagem iniciadas"


def _planilha(campanhas, inicio="2026-07-01", fim="2026-07-15", indicador=INDICADOR):
    wb = Workbook()
    ws = wb.active
    ws.append(CABECALHO)
    for c in campanhas:
        ws.append([
            c["nome"], c.get("status", "active"), c.get("res"),
            c.get("indicador", indicador), c.get("inv"), c.get("imp"),
            c.get("alc"), c.get("cliques"), inicio, fim,
        ])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _arquivo(nome, campanhas, **kw):
    return SimpleUploadedFile(nome, _planilha(campanhas, **kw), content_type=XLSX_MIME)


def _planilha_sem_veiculacao(campanhas, inicio="2026-07-01", fim="2026-07-15"):
    """Export legado, sem a coluna de veiculação da campanha."""
    wb = Workbook()
    ws = wb.active
    ws.append([c for c in CABECALHO if c != "Veiculação da campanha"])
    for c in campanhas:
        ws.append([c["nome"], c.get("res"), INDICADOR, c.get("inv"), c.get("imp"),
                   c.get("alc"), c.get("cliques"), inicio, fim])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _bytes_pdf(resposta):
    return b"".join(resposta.streaming_content)


def _texto_pdf(pdf_bytes):
    """Texto de todas as páginas do PDF, com espaços normalizados."""
    pdf = fitz.open(stream=pdf_bytes, filetype="pdf")
    texto = "\n".join(p.get_text() for p in pdf)
    pdf.close()
    return " ".join(texto.split())


def _paginas(pdf_bytes):
    """Nº de páginas via `pdfinfo` (poppler); fallback PyMuPDF se ausente."""
    if shutil.which("pdfinfo"):
        with tempfile.NamedTemporaryFile(suffix=".pdf") as f:
            f.write(pdf_bytes)
            f.flush()
            saida = subprocess.run(["pdfinfo", f.name], capture_output=True,
                                   text=True, check=True).stdout
        for linha in saida.splitlines():
            if linha.startswith("Pages:"):
                return int(linha.split(":")[1])
    pdf = fitz.open(stream=pdf_bytes, filetype="pdf")
    n = len(pdf)
    pdf.close()
    return n


def _tem_imagem(pdf_bytes, minimo=2):
    """True se a página tiver ao menos `minimo` imagens (logo + donut)."""
    pdf = fitz.open(stream=pdf_bytes, filetype="pdf")
    n = len(pdf[0].get_images())
    pdf.close()
    return n >= minimo


class FluxoIndividualTest(TestCase):
    """1 anexo: página única com funil, tabela + donut por campanha e análise."""

    def _upload(self):
        f = _arquivo("iloc.xlsx", [
            {"nome": "Campanha A", "res": 40, "inv": 80.0, "imp": 4000,
             "alc": 2500, "cliques": 300},
            {"nome": "Campanha B", "res": 20, "inv": 120.0, "imp": 2000,
             "alc": 1000, "cliques": 120},
        ])
        return self.client.post("/", {"cliente": "ILOC", "arquivos": [f]})

    def _gerar_pdf(self):
        return self.client.post("/revisao/", {
            "cliente": "ILOC", "periodo": "01/07/2026 a 15/07/2026",
            "analise": "Texto de análise.",
        })

    def test_fluxo_completo(self):
        r = self._upload()
        self.assertEqual(r.status_code, 302)

        r = self.client.get("/revisao/")
        self.assertEqual(r.status_code, 200)
        html = r.content.decode()
        self.assertIn("Topo de Funil — Atração", html)
        self.assertNotIn("checklist_agencia", html)

        dados = self.client.session["relatorio_apex"]
        self.assertNotIn("modo", dados)

        r = self._gerar_pdf()
        self.assertEqual(r.status_code, 200)
        self.assertIn('filename="ILOC-anexounico-1-jul-26-15-jul-26.pdf"',
                      r["Content-Disposition"])
        pdf = _bytes_pdf(r)
        self.assertEqual(_paginas(pdf), 1, "o relatório deve ter UMA página")

        texto = _texto_pdf(pdf)
        self.assertIn("Funil de Vendas", texto)
        self.assertIn("Desempenho por Campanha", texto)
        self.assertIn("Análise do Período", texto)
        # Tabela: share de resultados por campanha (40/60 = 67%)
        self.assertIn("Campanha A", texto)
        self.assertIn("67%", texto)
        # Donut embutido como imagem (além do logo)
        self.assertTrue(_tem_imagem(pdf), "donut ausente do PDF")
        # Sem seções antigas do layout ReportLab
        self.assertNotIn("Checklist", texto)
        self.assertNotIn("Composição por Unidade", texto)
        # Status de veiculação é assunto interno — não aparece no relatório
        self.assertNotIn("Status", texto)

    def test_toda_campanha_do_anexo_sai_no_pdf(self):
        """Nenhuma campanha é resumida em "Outras (N)" (12/08/2026).

        A soma fechava, mas o cliente não via quanto cada praça custou — e num
        relatório de uma campanha por cidade é exatamente isso que ele quer
        saber. Se não couber, o PDF ganha página.
        """
        f = _arquivo("conta.xlsx", [
            {"nome": f"Campanha {i}", "res": 10 * (12 - i), "inv": 50.0 + i,
             "imp": 3000, "alc": 2000, "cliques": 200}
            for i in range(12)
        ])
        r = self.client.post("/", {"cliente": "ILOC", "arquivos": [f]})
        self.assertEqual(r.status_code, 302)

        r = self._gerar_pdf()
        self.assertEqual(r.status_code, 200)
        pdf = _bytes_pdf(r)
        texto = _texto_pdf(pdf)
        for i in range(12):
            self.assertIn(f"Campanha {i}", texto)
        self.assertNotIn("Outras", texto)
        self.assertEqual(_paginas_com_rodape_invadido(pdf), [])

    def test_tabela_longa_empilha_em_vez_de_dividir_a_largura(self):
        """Em flex o WeasyPrint não parte a seção: ela pularia inteira para a
        página seguinte e deixaria meia folha em branco."""
        from .gerador_pdf import MAX_LINHAS_LADO_A_LADO as MAX, _empilhar
        self.assertFalse(_empilhar([{"nome": "C"}] * MAX))
        self.assertTrue(_empilhar([{"nome": "C"}] * (MAX + 1)))
        self.assertFalse(_empilhar(None))


class FluxoConsolidadoTest(TestCase):
    """2+ anexos: página única com funil geral, donut por unidade e análise."""

    def _upload_tres(self):
        # A: custo/resultado 2,00 (melhor) · B: 6,00 · C: sem resultados
        arquivos = [
            _arquivo("TIM_Sao_Jose.xlsx", [
                {"nome": "Camp SJ", "res": 50, "inv": 100.0, "imp": 10000,
                 "alc": 5000, "cliques": 500},
            ]),
            _arquivo("TIM_Braganca.xlsx", [
                {"nome": "Camp BR", "res": 50, "inv": 300.0, "imp": 20000,
                 "alc": 10000, "cliques": 1000},
            ]),
            _arquivo("TIM_Maxshopping.xlsx", [
                {"nome": "Camp MX", "res": 0, "inv": 200.0, "imp": 30000,
                 "alc": 20000, "status": "not_delivering"},
            ]),
        ]
        return self.client.post("/", {"cliente": "TIM Brasil", "arquivos": arquivos})

    def test_agregacao(self):
        r = self._upload_tres()
        self.assertEqual(r.status_code, 302)

        dados = self.client.session["relatorio_apex"]
        self.assertEqual(dados["modo"], "grupo")
        self.assertEqual(len(dados["unidades"]), 3)

        # Funil geral: totais somados e taxas recalculadas sobre os totais
        linhas = {m: v for etapa in dados["funil"]["etapas"]
                  for m, v, _ in etapa["linhas"]}
        self.assertEqual(linhas["Investimento Total"], "R$ 600,00")
        self.assertEqual(linhas["Impressões"], "60.000 visualizações")
        # CPM = 600 / 60000 * 1000 = 10,00 (não média dos CPMs por unidade)
        self.assertEqual(linhas["CPM (custo por mil)"], "R$ 10,00")
        self.assertEqual(linhas["CTR (taxa de cliques)"], "2,50%")
        self.assertEqual(linhas["Conversas Iniciadas"], "100")
        self.assertEqual(linhas["Custo por Conversa (CPA)"], "R$ 6,00")

    def test_pdf_consolidado(self):
        self._upload_tres()

        r = self.client.get("/revisao/")
        self.assertEqual(r.status_code, 200)
        self.assertContains(r, "unidade_0")

        r = self.client.post("/revisao/", {
            "cliente": "TIM Brasil", "periodo": "01/07/2026 a 15/07/2026",
            "analise": "Análise do grupo.",
            "unidade_0": "TIM São José", "unidade_1": "TIM Bragança",
            "unidade_2": "TIM Maxshopping",
        })
        self.assertEqual(r.status_code, 200)
        pdf = _bytes_pdf(r)
        self.assertEqual(_paginas(pdf), 1, "o consolidado deve ter UMA página")

        texto = _texto_pdf(pdf)
        self.assertIn("Consolidado de 3 unidades", texto)
        self.assertIn("R$ 600,00", texto)                 # total somado no funil
        self.assertIn("Composição por Unidade", texto)
        # Legenda do donut: nome + valor + percentual (unidades com resultado)
        self.assertIn("TIM São José — 50 (50%)", texto)
        self.assertIn("TIM Bragança — 50 (50%)", texto)
        # Rodapé lista as unidades somadas + nota de sobreposição de alcance
        self.assertIn("TIM São José, TIM Bragança, TIM Maxshopping", texto)
        self.assertIn("sobreposição de audiência", texto)
        # Donut embutido; sem tabela de campanhas no consolidado
        self.assertTrue(_tem_imagem(pdf), "donut ausente do PDF")
        self.assertNotIn("Desempenho por Campanha", texto)
        self.assertNotIn("Ranking", texto)
        # Status de veiculação (unidade C está not_delivering) não vaza
        self.assertNotIn("not_delivering", texto)

    def test_indicador_divergente_gera_aviso(self):
        arquivos = [
            _arquivo("loja_a.xlsx", [{"nome": "C", "res": 10, "inv": 50.0,
                                      "imp": 1000, "alc": 800}]),
            _arquivo("loja_b.xlsx", [{"nome": "C", "res": 5, "inv": 40.0,
                                      "imp": 900, "alc": 700}],
                     indicador="Compras"),
        ]
        r = self.client.post("/", {"cliente": "Grupo", "arquivos": arquivos})
        self.assertEqual(r.status_code, 302)

        # Aviso na tela de revisão (não bloqueia o fluxo)
        r = self.client.get("/revisao/")
        self.assertContains(r, "não usam o mesmo indicador")

        # A validação é interna: o aviso NÃO vaza para o PDF do cliente
        r = self.client.post("/revisao/", {
            "cliente": "Grupo", "periodo": "01/07/2026 a 15/07/2026",
            "analise": "Análise.", "unidade_0": "Loja A", "unidade_1": "Loja B",
        })
        self.assertEqual(r.status_code, 200)
        pdf = _bytes_pdf(r)
        self.assertEqual(_paginas(pdf), 1)
        self.assertNotIn("não usam o mesmo indicador", _texto_pdf(pdf))

    def test_vinte_arquivos_uma_pagina(self):
        arquivos = [
            _arquivo(f"unidade_{i:02d}.xlsx", [
                {"nome": f"Camp {i}", "res": 10 + i, "inv": 50.0 + i,
                 "imp": 1000, "alc": 800, "cliques": 100},
            ])
            for i in range(20)
        ]
        r = self.client.post("/", {"cliente": "Grupo 20", "arquivos": arquivos})
        self.assertEqual(r.status_code, 302)
        self.assertEqual(len(self.client.session["relatorio_apex"]["unidades"]), 20)

        post = {"cliente": "Grupo 20", "periodo": "01/07/2026 a 15/07/2026",
                "analise": "Análise."}
        post.update({f"unidade_{i}": f"Unidade {i:02d}" for i in range(20)})
        r = self.client.post("/revisao/", post)
        self.assertEqual(r.status_code, 200)
        pdf = _bytes_pdf(r)
        texto = _texto_pdf(pdf)
        self.assertIn("Consolidado de 20 unidades", texto)
        # As 20 na legenda do donut, uma a uma — sem "Outras (N)" desde
        # 12/08/2026 —, e as 20 outra vez no rodapé.
        self.assertNotIn("Outras", texto)
        for i in range(20):
            self.assertIn(f"Unidade {i:02d}", texto)
        self.assertEqual(_paginas_com_rodape_invadido(pdf), [])


def _dados(campanhas, **kw):
    """Atalho: parseia uma planilha sintética direto pelo ler_export_meta."""
    return ler_export_meta(io.BytesIO(_planilha(campanhas, **kw)))


def _linhas_funil(dados):
    """{métrica: (valor, leitura)} de todas as etapas do funil."""
    return {m: (v, l) for etapa in dados["funil"]["etapas"]
            for m, v, l in etapa["linhas"]}


class CasosReaisTest(TestCase):
    """
    Duas contas reais, do anexo à página impressa.

    Elix Finance sai do PDF em `docs/Elix-Finance-anexounico-31-jul-26-6-ago-26.pdf`.
    TecnoCell Chapada não tem PDF versionado: os números vêm do export
    `Tecno-Cell-Chapada-Campanhas-30-de-jul-de-2026-5-de-ago-de-2026.xlsx`,
    lido em 07/08/2026 — foi ele que revelou o rótulo `Valor gasto (BRL)`.
    """

    # 31/07 a 06/08/2026 · R$ 257,86 · 13 conversas · CPA R$ 19,84
    ELIX = [{"nome": "[LEADS][CLINICA-BOLETO][ABO][31JUL26]", "res": 13,
             "inv": 171.24, "imp": 3900, "alc": 1950},
            {"nome": "[LEADS][CLINICA-BOL][ABO][04AGO26]", "res": 0,
             "inv": 70.51, "imp": 1200, "alc": 600},
            {"nome": "[LEADS][CLINICA-BOL][ABO][04AGO26] — Cópia", "res": 0,
             "inv": 16.11, "imp": 460, "alc": 213}]

    # 30/07 a 05/08/2026 · R$ 338,80 · 82 conversas · CPA R$ 4,13 · 1 campanha
    TECNOCELL = [{"nome": "[VENDAS][IPHONE-ASSTECH][ABO][27JUL26]", "res": 82,
                  "inv": 338.80, "imp": 35840, "alc": 10206}]

    def _ler(self, campanhas, inicio, fim):
        return parser_xlsx.ler_export_meta(io.BytesIO(
            _planilha(campanhas, inicio=inicio, fim=fim)))

    def _elix(self):
        return self._ler(self.ELIX, "2026-07-31", "2026-08-06")

    def _tecnocell(self):
        return self._ler(self.TECNOCELL, "2026-07-30", "2026-08-05")

    # ---- Elix: verba parada é o assunto do período ----
    def test_elix_numeros_batem_com_o_pdf(self):
        n = self._elix()["_num"]
        self.assertAlmostEqual(n["investimento"], 257.86, places=2)
        self.assertEqual(n["resultados"], 13)
        self.assertAlmostEqual(n["custo_resultado"], 19.83, places=1)
        self.assertAlmostEqual(n["frequencia"], 2.01, places=2)

    def test_elix_aponta_a_verba_parada_e_a_faixa_esperada(self):
        av = self._elix()["avaliacao"]
        self.assertEqual(av["derivados"]["verba_sem_retorno"], 70.51)
        self.assertEqual(av["derivados"]["verba_em_aprendizado"], 16.11)
        self.assertEqual(av["derivados"]["resultados_esperados"], [3, 4])
        self.assertEqual(av["proximo_passo"], "corrigir_a_captacao")

    def test_elix_no_pdf_diz_que_o_gargalo_e_a_captacao(self):
        dados = self._elix()
        self.client.post("/", {"cliente": "Elix Finance", "arquivos": [
            _arquivo("elix.xlsx", self.ELIX, inicio="2026-07-31",
                     fim="2026-08-06")]})
        r = self.client.post("/revisao/", {
            "cliente": "Elix Finance", "periodo": "31/07/2026 a 06/08/2026",
            "analise": dados["analise_sugerida"].replace("\n", "\r\n")})
        pdf = _bytes_pdf(r)
        self.assertEqual(_paginas(pdf), 1)
        texto = _texto_pdf(pdf)
        self.assertIn("R$ 70,51", texto)
        self.assertIn("de 3 a 4 registros", texto)
        self.assertIn("na etapa de captação", texto)
        # E o número derivado NÃO é a soma das duas zeradas: a de R$ 16,11
        # ainda não gastou o equivalente a um contato.
        self.assertNotIn("R$ 86,62", texto)

    # ---- TecnoCell: entrega sadia, público saturado ----
    def test_tecnocell_numeros_batem_com_o_export(self):
        n = self._tecnocell()["_num"]
        self.assertAlmostEqual(n["investimento"], 338.80, places=2)
        self.assertEqual(n["resultados"], 82)
        self.assertAlmostEqual(n["custo_resultado"], 4.13, places=2)
        self.assertAlmostEqual(n["frequencia"], 3.51, places=2)

    def test_tecnocell_sem_verba_parada_e_publico_saturado(self):
        av = self._tecnocell()["avaliacao"]
        self.assertEqual(av["classificacao"], "BOM")   # 4,13 > teto 4,00
        self.assertNotIn("verba_sem_retorno", av["sinais"])
        self.assertIn("campanha_unica", av["sinais"])
        self.assertIn("frequencia_saturada", av["sinais"])
        self.assertEqual(av["proximo_passo"], "ampliar_publico_e_criativos")

    def test_tecnocell_no_pdf_nao_repete_os_numeros_da_tabela(self):
        dados = self._tecnocell()
        self.assertNotRegex(dados["analise_sugerida"], r"\d")
        self.assertNotIn("R$", dados["analise_sugerida"])

    def test_as_duas_contas_sao_lidas_de_forma_diferente(self):
        # Mesma janela de 7 dias e mesma aplicação: o texto tem que separar
        # uma conta com gargalo de captação de uma com público desgastado.
        elix = self._elix()["analise_sugerida"]
        tecno = self._tecnocell()["analise_sugerida"]
        self.assertNotEqual(elix, tecno)
        self.assertIn("captação", elix)
        self.assertNotIn("captação", tecno)
        self.assertIn("já viu os anúncios muitas vezes", tecno)


MARCA_RODAPE = "APEX — Gestão de Tráfego Pago"


def _paginas_com_rodape_invadido(pdf_bytes):
    """Páginas em que algum texto passa por baixo do rodapé.

    Era o defeito de 12/08/2026: a página tinha altura fixa e `overflow:
    hidden`, com o rodapé preso no fundo por `position: absolute`, então texto
    longo era DESENHADO POR CIMA dele em vez de empurrar página. Contar
    páginas não pega isso — o PDF continua com uma página, só que ilegível.
    Aqui a conferência é geométrica: nenhum bloco de texto pode terminar
    abaixo do topo do rodapé.
    """
    doc = fitz.open(stream=pdf_bytes, filetype="pdf")
    invadidas = []
    for numero, pagina in enumerate(doc, 1):
        blocos = [b for b in pagina.get_text("blocks") if b[4].strip()]
        rodape = [b for b in blocos if MARCA_RODAPE in b[4]]
        if not rodape:
            invadidas.append((numero, "sem rodapé"))
            continue
        topo = min(b[1] for b in rodape)
        for b in blocos:
            if (MARCA_RODAPE not in b[4]
                    and "Relatório gerado a partir" not in b[4]
                    and b[3] > topo + 1):          # 1pt de tolerância
                invadidas.append((numero, b[4].strip()[:40]))
    doc.close()
    return invadidas


class OrcamentoDePaginaTest(TestCase):
    """
    O teto de caracteres é o ORÇAMENTO DO MOTOR — não uma promessa de página.

    Até 12/08/2026 ele prometia uma página: o motor cortava blocos até o texto
    caber. A promessa caiu junto com a página de altura fixa, e por dois
    motivos. O primeiro é que ela nunca foi verdade — a medição antiga contava
    páginas num layout que desenhava o excedente por cima do rodapé. O segundo
    é que sustentá-la agora custaria conteúdo: o consolidado passou a listar
    todas as unidades, e o teto que garantiria uma página cortaria um bloco
    inteiro dos textos reais do motor. Encurtar o que o cliente lê para
    economizar uma folha é a troca errada.

    O que o teto ainda faz, e é testado aqui: impedir que o texto do motor
    cresça sem limite. No teto, ele sai INTEIRO, sem sobrepor o rodapé e sem
    passar de duas páginas — e o número continua ancorado no que de fato cabe
    numa página, medido por bissecção. O texto da IA não passa por aqui: esse
    transborda de propósito.
    """

    ROTULOS = ["Leitura do período.", "Ponto de atenção.", "Leitura atual.",
               "O que vamos fazer.", "Objetivo do próximo ciclo."]
    FRASE = ("O custo por resultado ficou dentro da faixa de trabalho da conta "
             "e a verba investida segue virando contato real com cliente de "
             "forma previsível ao longo de todo o período observado. ")
    # O teto pode passar do que cabe numa página, mas não pode virar outra
    # coisa: acima disso o relatório deixaria de ser de leitura rápida.
    TETO_SOBRE_CAPACIDADE = 2.0
    BLOCOS = 5

    def _texto(self, total, blocos=None):
        """String de exatamente `total` caracteres, em parágrafos rotulados —
        rótulos e separadores contam, porque é a string inteira que o limite
        controla."""
        blocos = blocos or self.BLOCOS
        marcas = ["<b>%s</b> " % self.ROTULOS[i % len(self.ROTULOS)]
                  for i in range(blocos)]
        sobra = total - sum(len(m) for m in marcas) - 2 * (blocos - 1)
        por_bloco = sobra // blocos
        partes = []
        for i, marca in enumerate(marcas):
            n = por_bloco if i < blocos - 1 else sobra - por_bloco * (blocos - 1)
            partes.append(marca + (self.FRASE * (n // len(self.FRASE) + 2))[:n])
        texto = "\n\n".join(partes)
        self.assertEqual(len(texto), total)
        return texto

    def _cabe(self, campos, total):
        r = self.client.post("/revisao/", dict(campos, analise=self._texto(total)))
        return _paginas(_bytes_pdf(r)) == 1

    def _maximo(self, campos, teto_busca=4000):
        baixo, alto = 600, teto_busca
        self.assertTrue(self._cabe(campos, baixo), "nem o texto mínimo cabe")
        self.assertFalse(self._cabe(campos, alto),
                         "o limite de busca não estoura a página — suba o teto")
        while alto - baixo > 25:
            meio = (baixo + alto) // 2
            if self._cabe(campos, meio):
                baixo = meio
            else:
                alto = meio
        return baixo

    def _conferir(self, limite, medido, modo):
        self.assertLessEqual(
            limite, int(medido * self.TETO_SOBRE_CAPACIDADE),
            "o teto de %s está solto demais: cabem %d caracteres numa página e "
            "o teto é %d — mais que o dobro. Ou o layout encolheu muito, ou o "
            "teto virou outra coisa; reveja templates.py"
            % (modo, medido, limite))

    def test_no_teto_o_texto_sai_inteiro_e_sem_sobrepor(self):
        # O que o teto promete hoje: no limite, o texto aparece inteiro, não
        # invade o rodapé e não passa de duas páginas — em qualquer número de
        # blocos e nos dois modos.
        casos = [
            ("conta", templates.LIMITE_PDF, [
                _arquivo("a.xlsx", [
                    {"nome": "Camp A", "res": 40, "inv": 80.0, "imp": 4000,
                     "alc": 2500, "cliques": 300},
                    {"nome": "Camp B", "res": 20, "inv": 120.0, "imp": 2000,
                     "alc": 1000, "cliques": 120},
                    {"nome": "Camp C", "res": 10, "inv": 60.0, "imp": 1500,
                     "alc": 900, "cliques": 80},
                    {"nome": "Camp D", "res": 5, "inv": 40.0, "imp": 900,
                     "alc": 500, "cliques": 40}])],
             {"cliente": "Medição", "periodo": "01/07/2026 a 31/07/2026"}),
            ("consolidado", templates.LIMITE_PDF_GRUPO,
             [_arquivo("u%d.xlsx" % i, [
                 {"nome": "C", "res": 40 + i, "inv": 80.0 + i, "imp": 4000,
                  "alc": 2500, "cliques": 300}]) for i in range(3)],
             {"cliente": "Medição", "periodo": "01/07/2026 a 31/07/2026",
              "unidade_0": "Praça A", "unidade_1": "Praça B",
              "unidade_2": "Praça C"}),
        ]
        for modo, limite, arquivos, campos in casos:
            with self.subTest(modo=modo):
                self.client.post("/", {"cliente": "Medição",
                                       "arquivos": arquivos})
                for blocos in (3, 4, 5):
                    texto = self._texto(limite, blocos)
                    r = self.client.post("/revisao/",
                                         dict(campos, analise=texto))
                    pdf = _bytes_pdf(r)
                    onde = "%s: %d caracteres em %d blocos" % (modo, limite,
                                                               blocos)
                    self.assertEqual(_paginas_com_rodape_invadido(pdf), [],
                                     onde + " passou por cima do rodapé")
                    self.assertLessEqual(_paginas(pdf), 2,
                                         onde + " precisou de mais de 2 páginas")
                    # Última frase do último bloco: prova que nada foi cortado.
                    fim = " ".join(texto.split()[-6:])
                    self.assertIn(fim, _texto_pdf(pdf), onde + " saiu cortado")

    def test_teto_da_conta_individual(self):
        self.client.post("/", {"cliente": "Medição", "arquivos": [
            _arquivo("a.xlsx", [
                {"nome": "Camp A", "res": 40, "inv": 80.0, "imp": 4000,
                 "alc": 2500, "cliques": 300},
                {"nome": "Camp B", "res": 20, "inv": 120.0, "imp": 2000,
                 "alc": 1000, "cliques": 120},
                {"nome": "Camp C", "res": 10, "inv": 60.0, "imp": 1500,
                 "alc": 900, "cliques": 80},
                {"nome": "Camp D", "res": 5, "inv": 40.0, "imp": 900,
                 "alc": 500, "cliques": 40}])]})
        campos = {"cliente": "Medição", "periodo": "01/07/2026 a 31/07/2026"}
        self._conferir(templates.LIMITE_PDF, self._maximo(campos), "conta")

    def test_teto_do_consolidado(self):
        self.client.post("/", {"cliente": "Medição", "arquivos": [
            _arquivo("u%d.xlsx" % i, [
                {"nome": "C", "res": 40 + i, "inv": 80.0 + i, "imp": 4000,
                 "alc": 2500, "cliques": 300}]) for i in range(3)]})
        campos = {"cliente": "Medição", "periodo": "01/07/2026 a 31/07/2026",
                  "unidade_0": "Praça A", "unidade_1": "Praça B",
                  "unidade_2": "Praça C"}
        self._conferir(templates.LIMITE_PDF_GRUPO, self._maximo(campos),
                       "consolidado")

    def test_o_texto_real_do_motor_fica_bem_dentro_do_orcamento(self):
        """O orçamento é rede de segurança, não tesoura de uso diário.

        Se um texto real do motor encostar no teto, o corte por precedência
        começa a agir no dia a dia — e o cliente passa a receber um bloco a
        menos sem que ninguém tenha decidido isso.
        """
        maiores = []
        for cpa in (1.0, 5.0, 12.0, 30.0):
            for resultados in (5, 50, 500):
                for frequencia in (1.2, 3.9):
                    dados = _dados([{"nome": "C", "res": resultados,
                                     "inv": cpa * resultados, "imp": 40000,
                                     "alc": int(40000 / frequencia),
                                     "cliques": 500}])
                    maiores.append(len(dados["analise_sugerida"]))
        self.assertLess(max(maiores), templates.LIMITE_PDF * 0.9,
                        "texto do motor encostando no teto: %d de %d"
                        % (max(maiores), templates.LIMITE_PDF))

    def test_o_texto_real_do_grupo_fica_bem_dentro_do_orcamento(self):
        grupo = consolidar_grupo([
            {"nome": "Praça %d" % i,
             "dados": _dados([{"nome": "C", "res": 10 + i * 30,
                               "inv": 400.0 - i * 90, "imp": 9000,
                               "alc": 4000, "cliques": 400}])}
            for i in range(4)])
        self.assertLess(len(grupo["analise_sugerida"]),
                        templates.LIMITE_PDF_GRUPO * 0.9,
                        "texto do grupo encostando no teto: %d de %d"
                        % (len(grupo["analise_sugerida"]),
                           templates.LIMITE_PDF_GRUPO))


RESPOSTA_IA ="""*Período analisado: 01/07/2026 a 15/07/2026*

*Leitura do período: BOM*

No período, investimos R$ 257,86 nas campanhas e geramos 13 conversas, \
resultando em um custo médio de R$ 19,84 por contato. A operação manteve um \
volume estável de oportunidades.

O principal destaque positivo foi Hortolândia, com custo de R$ 4,92 por \
conversa. Em contrapartida, Jaguariúna ficou em R$ 21,96.

No geral, os números mostram uma operação equilibrada, com diferenças \
relevantes entre as praças."""

RESPOSTA_LEITURAS_FUNIL = json.dumps({
    "frequencia": "Alcance renovado periodicamente para o público não saturar.",
    "cpm": "Leilão competitivo agora, ótimo momento para ganhar escala.",
    "ctr": "Criativos com apelo forte, gerando cliques acima da média do setor.",
    "taxa_conversao": "Atendimento respondendo bem ao volume de cliques recebido.",
}, ensure_ascii=False)


@override_settings(OPENAI_API_KEY="chave-de-teste", OPENAI_MODEL="modelo-de-teste")
class AnalisePorIATest(TestCase):
    """
    O botão "Escrever com IA" na tela 02.

    Nenhum teste desta classe toca a rede: `_chamar` é a única função que fala
    com a OpenAI e é ela que fica trocada por uma resposta fixa. Suíte offline
    continua sendo requisito do projeto — e chamada de verdade custa dinheiro.
    """

    ELIX = [{"nome": "[VAGAS][TIM][ABO][HORTOLANDIA][30JUL26]", "res": 13,
             "inv": 171.24, "imp": 3900, "alc": 1950, "cliques": 40},
            {"nome": "[VAGAS][TIM][ABO][JAGUARIUNA][23JUL26]", "res": 0,
             "inv": 70.51, "imp": 1200, "alc": 600, "cliques": 12},
            {"nome": "[VAGAS][REALME][ABO][ITU][07AGOS26]", "res": 0,
             "inv": 16.11, "imp": 460, "alc": 213}]

    def _importar(self, campanhas=None, cliente="Elix"):
        f = _arquivo("e.xlsx", campanhas or self.ELIX)
        self.client.post("/", {"cliente": cliente, "arquivos": [f]})
        return self.client.session["relatorio_apex"]

    def _pedir(self, resposta=RESPOSTA_IA, analise="texto do motor", **campos):
        """Clica em "Escrever com IA" com a resposta do modelo já decidida."""
        base = {"cliente": "Elix", "periodo": "01/07/2026 a 15/07/2026",
                "analise": analise, "analise_ia": "1"}
        base.update(campos)
        alvo = "relatorios.redator_ia._chamar"
        efeito = ({"side_effect": resposta} if isinstance(resposta, Exception)
                  else {"return_value": resposta})
        with patch(alvo, **efeito) as chamada:
            r = self.client.post("/revisao/", base)
        self.assertEqual(r.status_code, 200)
        return r, self.client.session["relatorio_apex"], chamada

    # ---- a tela ----
    def test_botao_so_aparece_havendo_chave(self):
        self._importar()
        self.assertIn('name="analise_ia"',
                      self.client.get("/revisao/").content.decode())
        with override_settings(OPENAI_API_KEY=""):
            self.assertNotIn('name="analise_ia"',
                             self.client.get("/revisao/").content.decode())

    def test_texto_do_modelo_entra_no_textarea_e_na_sessao(self):
        do_motor = self._importar()["analise_sugerida"]
        r, dados, _ = self._pedir()

        self.assertIn("O principal destaque positivo foi Hortolândia",
                      r.content.decode())
        self.assertIn("Hortolândia", dados["analise_ia"])
        # O texto do motor continua guardado: a IA não o apaga, só deixa de ser
        # o que a tela mostra.
        self.assertEqual(dados["analise_sugerida"], do_motor)
        self.assertEqual(dados["analise_ia_bruta"], RESPOSTA_IA)

    def test_asterisco_vira_negrito_e_o_periodo_repetido_sai(self):
        self._importar()
        _r, dados, _ = self._pedir()
        texto = dados["analise_ia"]
        self.assertIn("<b>Leitura do período: BOM</b>", texto)
        self.assertNotIn("*", texto)
        # O período já está no cabeçalho do PDF — repeti-lo na análise é
        # duplicação que o cliente lê como descuido.
        self.assertNotIn("Período analisado", texto)

    def test_pedir_analise_nao_gera_pdf(self):
        self._importar()
        r, _dados, _ = self._pedir()
        self.assertEqual(r["Content-Type"].split(";")[0], "text/html")

    def test_recarregar_a_tela_mantem_o_texto_da_ia(self):
        self._importar()
        self._pedir()
        self.assertIn("Hortolândia", self.client.get("/revisao/").content.decode())

    def test_pdf_sai_com_o_texto_da_ia(self):
        self._importar()
        _r, dados, _ = self._pedir()
        r = self.client.post("/revisao/", {
            "cliente": "Elix", "periodo": "01/07/2026 a 15/07/2026",
            "analise": dados["analise_ia"].replace("\n", "\r\n")})
        texto = _texto_pdf(_bytes_pdf(r))
        self.assertIn("destaque positivo foi Hortolândia", texto)
        self.assertEqual(_paginas(_bytes_pdf(self.client.post("/revisao/", {
            "cliente": "Elix", "periodo": "01/07/2026 a 15/07/2026",
            "analise": dados["analise_ia"].replace("\n", "\r\n")}))), 1)

    # ---- o que chega ao modelo ----
    def test_payload_leva_os_totais_e_o_recorte_por_campanha(self):
        dados = self._importar()
        payload = redator_ia.montar_payload(dados)

        self.assertEqual(payload["periodo_analisado"], "01/07/2026 a 15/07/2026")
        self.assertEqual(payload["totais"]["resultados"], 13)
        self.assertEqual(payload["totais"]["investimento_reais"], 257.86)
        self.assertEqual(payload["totais"]["custo_por_resultado_reais"], 19.84)
        # Custo por campanha calculado aqui, não pelo modelo: divisão feita por
        # LLM sai errada sem ninguém perceber.
        primeira = payload["campanhas"][0]
        self.assertEqual(primeira["resultados"], 13)
        self.assertEqual(primeira["custo_por_resultado_reais"], 13.17)
        # Campanha sem resultado não vira divisão por zero.
        self.assertIsNone(payload["campanhas"][-1]["custo_por_resultado_reais"])

    def test_payload_declara_o_que_o_relatorio_nao_tem(self):
        """A lista de ausências é o que segura a alucinação: sem ela o modelo
        compara com "o mês passado" que não existe no arquivo."""
        payload = redator_ia.montar_payload(self._importar())
        ausentes = " ".join(payload["dados_ausentes"])
        for esperado in ("período anterior", "recorte por dia",
                         "conjunto de anúncios", "ROAS"):
            self.assertIn(esperado, ausentes)

    def test_coluna_ausente_no_export_e_declarada_ausente(self):
        """Coluna que não veio ≠ métrica igual a zero. A soma de uma coluna
        inexistente dá zero, e mandar "0 cliques" ao modelo é oferecer um
        parágrafo sobre um fracasso que ninguém mediu."""
        wb = Workbook()
        ws = wb.active
        ws.append([c for c in CABECALHO if c != "Cliques no link"])
        ws.append(["[VAGAS][TIM][ABO][ITU][06AGOS26]", "active", 13, INDICADOR,
                   171.24, 3900, 1950, "2026-07-01", "2026-07-15"])
        buf = io.BytesIO()
        wb.save(buf)
        self.client.post("/", {"cliente": "Elix", "arquivos": [
            SimpleUploadedFile("s.xlsx", buf.getvalue(), content_type=XLSX_MIME)]})

        payload = redator_ia.montar_payload(self.client.session["relatorio_apex"])
        self.assertNotIn("cliques_no_link", payload["totais"])
        self.assertIn("cliques, CTR e custo por clique", payload["dados_ausentes"])
        # E o que a planilha traz continua indo.
        self.assertEqual(payload["totais"]["resultados"], 13)

    def test_zero_medido_continua_sendo_zero(self):
        """O contrário do teste acima: a coluna veio, e veio vazia. Isso é
        informação — o modelo pode dizer que não houve clique."""
        payload = redator_ia.montar_payload(
            self._importar([dict(c, cliques=None) for c in self.ELIX]))
        self.assertEqual(payload["totais"]["cliques_no_link"], 0)
        self.assertNotIn("cliques, CTR e custo por clique",
                         payload["dados_ausentes"])

    def test_o_prompt_do_operador_vai_inteiro_no_system(self):
        self._importar()
        _r, _dados, chamada = self._pedir()
        # A primeira chamada é a Análise do Período; a segunda (não coberta
        # por este teste) é a das legendas do funil — call_args_list[0], não
        # call_args, que pegaria a mais recente.
        mensagens = chamada.call_args_list[0].args[0]
        self.assertEqual(mensagens[0]["role"], "system")
        self.assertIn("Atue como gestor de tráfego pago sênior",
                      mensagens[0]["content"])
        self.assertIn("Exatamente 3 parágrafos de análise", mensagens[0]["content"])
        # E o relatório entra como JSON na mensagem do usuário.
        self.assertEqual(json.loads(mensagens[1]["content"])["totais"]["resultados"], 13)

    # ---- o que volta do modelo ----
    def test_html_vindo_do_modelo_e_escapado(self):
        """O template do PDF renderiza a análise com `|safe`. O que volta da
        API é texto de terceiro e não pode chegar lá como marcação."""
        self._importar()
        _r, dados, _ = self._pedir(
            resposta='*Leitura do período: BOM*\n\n<script>alert(1)</script>')
        self.assertNotIn("<script>", dados["analise_ia"])
        self.assertIn("&lt;script&gt;", dados["analise_ia"])

    def test_texto_longo_avisa_da_segunda_pagina_sem_cortar(self):
        self._importar()
        longo = "*Leitura do período: BOM*\n\n" + ("palavra " * 600)
        r, dados, _ = self._pedir(resposta=longo)
        self.assertIn("desce inteira para uma segunda", r.content.decode())
        # E o texto chega inteiro ao textarea: quem encurta é o operador.
        self.assertGreater(len(dados["analise_ia"]), templates.LIMITE_PDF)

    def test_formato_fora_do_esperado_avisa(self):
        self._importar()
        r, _dados, _ = self._pedir(resposta="Um parágrafo só, sem cabeçalho.")
        self.assertIn("vieram 1 blocos", r.content.decode())

    # ---- quando falha ----
    def test_falha_da_api_nao_custa_o_relatorio(self):
        do_motor = self._importar()["analise_sugerida"]
        r, dados, _ = self._pedir(
            resposta=redator_ia.ErroDeIA("sem crédito na conta"),
            analise=do_motor)

        html = r.content.decode()
        self.assertIn("sem crédito na conta", html)
        self.assertIn("não foi gerada", html)
        # O texto que estava na tela continua lá (o textarea escapa as tags do
        # motor, então a comparação é pela frase), e nada foi gravado.
        self.assertIn(_paragrafos_analise(do_motor)[0][-60:].replace("<b>", ""),
                      html)
        self.assertNotIn("analise_ia", dados)

    def test_sem_chave_configurada_o_erro_diz_o_que_fazer(self):
        with override_settings(OPENAI_API_KEY=""):
            with self.assertRaises(redator_ia.ErroDeIA) as e:
                redator_ia.gerar({})
        self.assertIn("OPENAI_API_KEY", str(e.exception))
        self.assertEqual(e.exception.motivo, "chave")

    def test_credito_acabado_vira_aviso_proprio_e_tira_o_botao(self):
        """O caso que motivou a classificação: saldo zerado.

        Um 429 genérico convidaria a clicar de novo; sem crédito, clicar de
        novo é gastar o tempo do operador para receber o mesmo erro.
        """
        self._importar()
        r, _dados, _ = self._pedir(resposta=redator_ia.ErroDeIA(
            "Os créditos da IA acabaram.", "credito"))

        html = r.content.decode()
        self.assertIn("créditos da IA acabaram", html)
        self.assertIn('class="erro"', html)          # vermelho, não amarelo
        self.assertNotIn('name="analise_ia"', html)  # e sem botão para repetir

    def test_falha_passageira_mantem_o_botao_na_tela(self):
        """O oposto do teste acima: rede caiu, clicar de novo resolve."""
        self._importar()
        r, _dados, _ = self._pedir(resposta=redator_ia.ErroDeIA(
            "A OpenAI não respondeu em 90 segundos.", "rede"))

        html = r.content.decode()
        self.assertIn('class="aviso"', html)
        self.assertIn('name="analise_ia"', html)

    # ---- de que a SDK falhou para o que a tela diz ----
    def _erro_da_sdk(self, status=None, code=None, classe="APIStatusError",
                     param=None):
        """Imita o erro da SDK: status HTTP, `code` e `param` no corpo."""
        e = type(classe, (Exception,), {})("Error code: %s" % status)
        e.status_code = status
        if code or param:
            e.body = {"error": {"code": code, "type": "insufficient_quota",
                                "param": param}}
        return e

    def test_credito_e_excesso_de_chamadas_nao_se_confundem(self):
        """Os dois chegam como HTTP 429; só o `code` separa um do outro."""
        motivo, msg = redator_ia._classificar(
            self._erro_da_sdk(429, "insufficient_quota"))
        self.assertEqual(motivo, "credito")
        self.assertIn("créditos", msg)

        motivo, msg = redator_ia._classificar(
            self._erro_da_sdk(429, "rate_limit_exceeded"))
        self.assertEqual(motivo, "limite")
        self.assertNotIn("créditos", msg)

    def test_chave_modelo_rede_e_desconhecido_saem_classificados(self):
        casos = [
            (self._erro_da_sdk(401, "invalid_api_key"), "chave", "OPENAI_API_KEY"),
            (self._erro_da_sdk(404, "model_not_found"), "modelo", "modelo-de-teste"),
            # Modelo que não raciocina recusa o parâmetro em vez de o ignorar.
            (self._erro_da_sdk(400, "unsupported_value",
                               param="reasoning_effort"),
             "modelo", "esforço de raciocínio"),
            (self._erro_da_sdk(500), "servico", "instabilidade"),
            (self._erro_da_sdk(classe="APITimeoutError"), "rede", "90 segundos"),
            (ValueError("coisa nunca vista"), "desconhecido", "coisa nunca vista"),
        ]
        for erro, esperado, trecho in casos:
            with self.subTest(motivo=esperado):
                motivo, msg = redator_ia._classificar(erro)
                self.assertEqual(motivo, esperado)
                self.assertIn(trecho, msg)

    def test_o_code_do_atributo_vale_tanto_quanto_o_do_corpo(self):
        """Versões da SDK diferem em preencher `.code`; o corpo é o fallback."""
        e = self._erro_da_sdk(429)
        e.code = "insufficient_quota"
        self.assertEqual(redator_ia._classificar(e)[0], "credito")

    def test_o_prompt_leva_o_limite_da_pagina_e_ele_muda_com_o_modo(self):
        """No consolidado sobra menos folha: a tabela de unidades come o resto.

        Fosse um número fixo escrito no prompt, um dos dois modos escreveria
        para o limite do outro — e no consolidado errar para cima é o cliente
        recebendo um PDF de duas folhas.
        """
        def sistema_de(modo):
            with patch("relatorios.redator_ia._chamar", return_value="x") as c:
                redator_ia.gerar({"modo": modo})
            return c.call_args[0][0][0]["content"]

        reserva = redator_ia.RESERVA_DO_CABECALHO
        self.assertIn(f"máximo {templates.LIMITE_PDF_GRUPO - reserva} "
                      "caracteres", sistema_de("grupo"))
        self.assertIn(f"máximo {templates.LIMITE_PDF - reserva} caracteres",
                      sistema_de("individual"))

    def test_o_tamanho_pedido_e_faixa_e_nao_so_teto(self):
        """Só o teto fez o modelo tratá-lo como alvo a evitar: ele escrevia
        70% do que cabia e a análise saía curta à toa."""
        regra = redator_ia._regra_de_tamanho(templates.LIMITE_PDF_GRUPO)
        cabe = templates.LIMITE_PDF_GRUPO - redator_ia.RESERVA_DO_CABECALHO
        teto = int(cabe / redator_ia.CHARS_POR_PALAVRA)
        self.assertIn(f"entre {int(teto * redator_ia.PISO)} e", regra)
        self.assertLess(int(teto * redator_ia.PISO), teto)

    def test_a_requisicao_leva_modelo_teto_e_esforco_de_raciocinio(self):
        """O que sai na chamada — a parte que nenhum outro teste enxerga.

        Os demais trocam `_chamar` inteiro por uma resposta fixa, então o
        `reasoning_effort` poderia sumir da requisição sem quebrar nada: sem
        ele a OpenAI aplica o padrão dela, e o relatório muda de tom no dia em
        que esse padrão mudar.
        """
        resposta = SimpleNamespace(choices=[SimpleNamespace(
            finish_reason="stop",
            message=SimpleNamespace(content="texto do modelo", refusal=None))])
        criar = MagicMock(return_value=resposta)
        cliente = SimpleNamespace(chat=SimpleNamespace(
            completions=SimpleNamespace(create=criar)))

        with patch("openai.OpenAI", return_value=cliente):
            texto = redator_ia._chamar([{"role": "user", "content": "x"}])

        self.assertEqual(texto, "texto do modelo")
        kwargs = criar.call_args.kwargs
        self.assertEqual(kwargs["model"], "modelo-de-teste")
        self.assertEqual(kwargs["max_completion_tokens"], redator_ia.MAX_TOKENS)
        self.assertEqual(kwargs["reasoning_effort"], redator_ia.ESFORCO)

    def test_o_esforco_configurado_e_um_dos_que_a_api_aceita(self):
        """`chat.completions` recusa `max` com 400 — só o `/v1/responses` o
        aceita, e não é por ele que este projeto fala."""
        self.assertIn(redator_ia.ESFORCO,
                      ("none", "low", "medium", "high", "xhigh"))

    # ---- HTTP 200 sem texto dentro ----
    def _escolha_vazia(self, finish_reason, refusal=None):
        """Imita `resposta.choices[0]` quando o modelo não escreveu nada."""
        return SimpleNamespace(
            finish_reason=finish_reason,
            message=SimpleNamespace(content="", refusal=refusal))

    def test_teto_estourado_nao_convida_a_clicar_de_novo(self):
        """`length` sem texto é o raciocínio tendo comido o MAX_TOKENS inteiro.

        Repetir gasta de novo para receber o mesmo nada — daí ser definitivo —,
        e a mensagem precisa apontar o teto, porque a conta do operador está
        boa e mandá-lo conferir chave ou saldo é despistá-lo.
        """
        motivo, msg = redator_ia._diagnosticar_vazio(
            self._escolha_vazia("length"))
        self.assertEqual(motivo, "teto")
        self.assertIn(motivo, redator_ia.DEFINITIVOS)
        self.assertIn(str(redator_ia.MAX_TOKENS), msg)

    def test_recusa_do_modelo_chega_inteira_na_tela(self):
        motivo, msg = redator_ia._diagnosticar_vazio(
            self._escolha_vazia("stop", refusal="não escrevo sobre isso"))
        self.assertEqual(motivo, "vazio")
        self.assertIn("não escrevo sobre isso", msg)

    def test_vazio_inexplicado_guarda_o_finish_reason_e_mantem_o_botao(self):
        """Sem o `finish_reason` na mensagem não sobra pista nenhuma."""
        motivo, msg = redator_ia._diagnosticar_vazio(
            self._escolha_vazia("content_filter"))
        self.assertEqual(motivo, "vazio")
        self.assertNotIn(motivo, redator_ia.DEFINITIVOS)
        self.assertIn("content_filter", msg)

    # ---- consolidado ----
    def test_consolidado_manda_as_unidades_e_aceita_o_texto(self):
        arquivos = [
            _arquivo("a.xlsx", [{"nome": "CA", "res": 50, "inv": 100.0,
                                 "imp": 5000, "alc": 4000}]),
            _arquivo("b.xlsx", [{"nome": "CB", "res": 10, "inv": 200.0,
                                 "imp": 5000, "alc": 4000}]),
        ]
        self.client.post("/", {"cliente": "Grupo", "arquivos": arquivos})
        dados = self.client.session["relatorio_apex"]

        payload = redator_ia.montar_payload(dados)
        self.assertEqual([u["nome_da_unidade"] for u in payload["unidades"]],
                         ["a", "b"])
        self.assertNotIn("campanhas", payload)

        with patch("relatorios.redator_ia._chamar", return_value=RESPOSTA_IA):
            r = self.client.post("/revisao/", {
                "cliente": "Grupo", "periodo": "01/07/2026 a 15/07/2026",
                "analise": "x", "analise_ia": "1",
                "unidade_0": "Praça A", "unidade_1": "Praça B"})

        self.assertEqual(r.status_code, 200)
        html = r.content.decode()
        self.assertIn("Hortolândia", html)
        # Os nomes das unidades sobrevivem à volta para a tela.
        self.assertIn("Praça A", html)

    def test_consolidado_usa_o_teto_de_pagina_do_consolidado(self):
        """O consolidado cabe menos texto que o individual (a composição por
        unidade ocupa a página), então o aviso tem que vir antes."""
        arquivos = [_arquivo(f"{n}.xlsx", [{"nome": n, "res": 50, "inv": 100.0,
                                            "imp": 5000, "alc": 4000}])
                    for n in ("a", "b")]
        self.client.post("/", {"cliente": "Grupo", "arquivos": arquivos})
        entre_os_dois = "x" * ((templates.LIMITE_PDF_GRUPO
                                + templates.LIMITE_PDF) // 2)
        with patch("relatorios.redator_ia._chamar", return_value=entre_os_dois):
            r = self.client.post("/revisao/", {
                "cliente": "Grupo", "periodo": "01/07/2026 a 15/07/2026",
                "analise": "x", "analise_ia": "1",
                "unidade_0": "A", "unidade_1": "B"})
        self.assertIn(str(templates.LIMITE_PDF_GRUPO), r.content.decode())


@override_settings(OPENAI_API_KEY="chave-de-teste", OPENAI_MODEL="modelo-de-teste")
class LeiturasDoFunilIATest(TestCase):
    """A segunda chamada do botão "Escrever com IA": as legendas do funil.

    Mesma regra de `AnalisePorIATest` — `_chamar` nunca fala com a rede de
    verdade. Aqui o `side_effect` é uma lista de duas respostas: a primeira
    chamada é sempre a Análise do Período (`RESPOSTA_IA`), a segunda são as
    legendas do funil — é essa que cada teste varia.
    """

    # Mesma campanha de `AnalisePorIATest.ELIX[0]` (Hortolândia): frequência
    # 2,0 (dentro), CPM 43,91 (dentro), CTR 1,03% (abaixo), taxa de conversão
    # 32,5% (acima) — dá pra prever o texto estático de cada uma.
    CAMPANHAS = [{"nome": "[VAGAS][TIM][ABO][HORTOLANDIA][30JUL26]", "res": 13,
                  "inv": 171.24, "imp": 3900, "alc": 1950, "cliques": 40}]

    def _importar(self):
        f = _arquivo("e.xlsx", self.CAMPANHAS)
        self.client.post("/", {"cliente": "Elix", "arquivos": [f]})
        return self.client.session["relatorio_apex"]

    def _pedir(self, resposta_funil):
        base = {"cliente": "Elix", "periodo": "01/07/2026 a 15/07/2026",
                "analise": "texto do motor", "analise_ia": "1"}
        alvo = "relatorios.redator_ia._chamar"
        with patch(alvo, side_effect=[RESPOSTA_IA, resposta_funil]) as chamada:
            r = self.client.post("/revisao/", base)
        self.assertEqual(r.status_code, 200)
        return r, self.client.session["relatorio_apex"], chamada

    @staticmethod
    def _linha(funil, rotulo):
        for etapa in funil["etapas"]:
            for linha in etapa["linhas"]:
                if linha[0] == rotulo:
                    return linha
        raise AssertionError(f"linha {rotulo!r} não está no funil")

    def test_json_valido_substitui_as_4_leituras(self):
        self._importar()
        _r, dados, chamada = self._pedir(RESPOSTA_LEITURAS_FUNIL)

        self.assertEqual(chamada.call_count, 2)
        funil = dados["funil"]
        self.assertEqual(self._linha(funil, "Frequência")[2],
                         "Alcance renovado periodicamente para o público não "
                         "saturar.")
        self.assertEqual(self._linha(funil, "CPM (custo por mil)")[2],
                         "Leilão competitivo agora, ótimo momento para ganhar "
                         "escala.")
        self.assertEqual(self._linha(funil, "CTR (taxa de cliques)")[2],
                         "Criativos com apelo forte, gerando cliques acima da "
                         "média do setor.")
        self.assertEqual(
            self._linha(funil, "Taxa de Conversão (clique → conversa)")[2],
            "Atendimento respondendo bem ao volume de cliques recebido.")

    def test_a_segunda_chamada_usa_esforco_e_teto_proprios(self):
        """A chamada das legendas é bem mais barata que a da análise — se ela
        silenciosamente voltasse a usar o teto/esforço padrão, o custo por
        clique no botão dobraria sem ninguém perceber."""
        self._importar()
        _r, _dados, chamada = self._pedir(RESPOSTA_LEITURAS_FUNIL)
        kwargs = chamada.call_args_list[1].kwargs
        self.assertEqual(kwargs["max_tokens"],
                         redator_ia.MAX_TOKENS_LEITURAS_FUNIL)
        self.assertEqual(kwargs["esforco"], redator_ia.ESFORCO_LEITURAS_FUNIL)

    def test_chave_ausente_no_json_mantem_o_texto_estatico_so_naquela_linha(self):
        parcial = json.dumps({"cpm": "Leilão competitivo agora."})
        self._importar()
        _r, dados, _ = self._pedir(parcial)
        funil = dados["funil"]
        self.assertEqual(self._linha(funil, "CPM (custo por mil)")[2],
                         "Leilão competitivo agora.")
        # Frequência não veio no JSON: continua com o texto do catálogo.
        self.assertEqual(self._linha(funil, "Frequência")[2],
                         "Frequência saudável — público longe da saturação.")

    def test_valor_acima_do_limite_de_sanidade_e_descartado(self):
        longo = json.dumps({"cpm": "x " * redator_ia.LIMITE_LEITURA_FUNIL})
        self._importar()
        _r, dados, _ = self._pedir(longo)
        # Estourou o teto: cai no catálogo, não no texto gigante do modelo.
        self.assertEqual(self._linha(dados["funil"], "CPM (custo por mil)")[2],
                         "Custo de entrega competitivo.")

    def test_chave_desconhecida_no_json_e_ignorada(self):
        com_lixo = json.dumps({"cpm": "Leilão competitivo agora.",
                               "algo_que_nao_pedimos": "x"})
        self._importar()
        _r, dados, _ = self._pedir(com_lixo)
        self.assertEqual(self._linha(dados["funil"], "CPM (custo por mil)")[2],
                         "Leilão competitivo agora.")

    def test_json_malformado_nao_afeta_a_analise_principal_e_vira_aviso(self):
        self._importar()
        r, dados, _ = self._pedir("isso não é json")
        self.assertIn("Hortolândia", dados["analise_ia"])
        self.assertIn("Legendas do funil", r.content.decode())
        # Sem chave nenhuma vinda do modelo, tudo fica no texto estático.
        self.assertEqual(self._linha(dados["funil"], "CPM (custo por mil)")[2],
                         "Custo de entrega competitivo.")

    def test_cerca_de_bloco_de_codigo_e_tolerada(self):
        cercado = "```json\n" + RESPOSTA_LEITURAS_FUNIL + "\n```"
        self._importar()
        _r, dados, _ = self._pedir(cercado)
        self.assertEqual(self._linha(dados["funil"], "CPM (custo por mil)")[2],
                         "Leilão competitivo agora, ótimo momento para ganhar "
                         "escala.")

    def test_falha_de_rede_na_segunda_chamada_nao_bloqueia_a_tela(self):
        self._importar()
        alvo = "relatorios.redator_ia._chamar"
        with patch(alvo, side_effect=[
                RESPOSTA_IA, redator_ia.ErroDeIA("instabilidade", "servico")]):
            r = self.client.post("/revisao/", {
                "cliente": "Elix", "periodo": "01/07/2026 a 15/07/2026",
                "analise": "texto do motor", "analise_ia": "1"})
        self.assertEqual(r.status_code, 200)
        html = r.content.decode()
        self.assertIn("Hortolândia", html)  # a análise principal segue lá
        self.assertIn("instabilidade", html)
        self.assertNotIn("A análise por IA não foi gerada", html)


class TextoLongoNoPdfTest(TestCase):
    """
    Análise longa transborda para outra página — nunca por cima do rodapé.

    A escrita por IA não tem catálogo fixo de blocos: o tamanho do texto é
    decidido pelo modelo, e o relatório precisa aguentar isso. Antes de
    12/08/2026 a página tinha altura travada e o excedente era desenhado sobre
    o rodapé; com o export novo (que preenche o meio do funil) bastavam 1.200
    caracteres para o defeito aparecer.
    """

    CAMPANHAS = [{"nome": f"[VAGAS][TIM][ABO][CIDADE{i}][23JUL26]",
                  "res": 40 - i, "inv": 200.0 - i, "imp": 12000, "alc": 6000,
                  "cliques": 90} for i in range(5)]

    def _pdf(self, analise):
        f = _arquivo("e.xlsx", self.CAMPANHAS)
        self.client.post("/", {"cliente": "Mobile Magazine", "arquivos": [f]})
        r = self.client.post("/revisao/", {
            "cliente": "Mobile Magazine", "periodo": "05/08/2026 a 11/08/2026",
            "analise": analise.replace("\n", "\r\n")})
        return _bytes_pdf(r)

    def _analise(self, blocos):
        frase = ("O custo por resultado ficou dentro da faixa de trabalho da "
                 "conta e a verba investida segue virando contato real. ")
        return "\n\n".join("<b>Bloco %d.</b> %s" % (i, frase * 4)
                           for i in range(1, blocos + 1))

    def test_analise_curta_continua_em_uma_pagina(self):
        """A regressão a evitar do outro lado: relatório normal não pode
        passar a sair com duas páginas."""
        pdf = self._pdf("<b>Leitura do período.</b> Texto curto do motor.")
        self.assertEqual(_paginas(pdf), 1)
        self.assertEqual(_paginas_com_rodape_invadido(pdf), [])

    def test_analise_longa_ganha_pagina_em_vez_de_sobrepor(self):
        pdf = self._pdf(self._analise(6))
        self.assertGreater(_paginas(pdf), 1)
        self.assertEqual(_paginas_com_rodape_invadido(pdf), [])

    def test_o_rodape_sai_em_toda_pagina(self):
        pdf = self._pdf(self._analise(6))
        doc = fitz.open(stream=pdf, filetype="pdf")
        paginas = [MARCA_RODAPE in p.get_text() for p in doc]
        doc.close()
        self.assertTrue(all(paginas), f"rodapé faltando em {paginas}")

    def test_a_analise_nao_se_parte_ao_meio(self):
        """Cabeçalho da seção e texto viajam juntos: partir deixava um
        parágrafo órfão numa página em branco."""
        pdf = self._pdf(self._analise(3))
        doc = fitz.open(stream=pdf, filetype="pdf")
        paginas = [p.get_text() for p in doc]
        doc.close()
        com_titulo = [i for i, t in enumerate(paginas) if "Análise do Período" in t]
        com_texto = [i for i, t in enumerate(paginas) if "Bloco 1." in t]
        self.assertEqual(com_titulo, com_texto)
        # E o último bloco não ficou para trás numa página sozinho.
        self.assertIn("Bloco 3.", paginas[com_titulo[0]])

    def test_consolidado_tambem_transborda_sem_sobrepor(self):
        arquivos = [_arquivo(f"{n}.xlsx", [{"nome": n, "res": 50, "inv": 100.0,
                                            "imp": 5000, "alc": 4000}])
                    for n in ("a", "b", "c")]
        self.client.post("/", {"cliente": "Grupo", "arquivos": arquivos})
        r = self.client.post("/revisao/", {
            "cliente": "Grupo", "periodo": "01/07/2026 a 15/07/2026",
            "unidade_0": "A", "unidade_1": "B", "unidade_2": "C",
            "analise": self._analise(6).replace("\n", "\r\n")})
        pdf = _bytes_pdf(r)
        self.assertGreater(_paginas(pdf), 1)
        self.assertEqual(_paginas_com_rodape_invadido(pdf), [])


class ColunaDeInvestimentoTest(TestCase):
    """O Meta alterna o rótulo da coluna de verba entre exports. Não
    reconhecê-la zera investimento e custo por resultado sem avisar
    ninguém — o zero passa por número legítimo."""

    def _dados(self, rotulo):
        wb = Workbook()
        ws = wb.active
        ws.append([c if c != "Valor usado (BRL)" else rotulo
                   for c in CABECALHO])
        ws.append(["Campanha A", "active", 82, INDICADOR, 338.80, 35840,
                   10206, 500, "2026-07-30", "2026-08-05"])
        buf = io.BytesIO()
        wb.save(buf)
        return ler_export_meta(io.BytesIO(buf.getvalue()))

    def test_variantes_de_rotulo_da_coluna(self):
        for rotulo in ("Valor usado (BRL)", "Valor gasto (BRL)",
                       "Valor investido (BRL)", "Amount spent (BRL)"):
            with self.subTest(rotulo=rotulo):
                n = self._dados(rotulo)["_num"]
                self.assertAlmostEqual(n["investimento"], 338.80, places=2)
                self.assertAlmostEqual(n["custo_resultado"], 4.13, places=2)

    def test_coluna_desconhecida_nao_vira_periodo_otimo(self):
        # Sem a verba o CPA sai zero, e zero é mais barato que a faixa mais
        # barata de qualquer perfil. A análise tem que recusar a leitura.
        dados = self._dados("Verba aplicada")
        self.assertEqual(dados["_num"]["investimento"], 0.0)
        self.assertEqual(dados["avaliacao"]["classificacao"], "ATENCAO")
        self.assertEqual(dados["avaliacao"]["motivo_principal"],
                         "sem_investimento")
        self.assertIn("não trouxe o valor investido", dados["analise_sugerida"])


class BenchmarksTest(TestCase):
    """Classificação das métricas nas faixas de referência configuráveis."""

    def test_classificacao_basica(self):
        av = benchmarks.avaliar_metricas({
            "ctr": 1.0, "cpm": 30.0, "cpc": 4.0,
            "frequencia": 2.0, "taxa_conversao": 5.0,
        })
        self.assertEqual(av, {
            "ctr": benchmarks.ABAIXO, "cpm": benchmarks.DENTRO,
            "cpc": benchmarks.ACIMA, "frequencia": benchmarks.DENTRO,
            "taxa_conversao": benchmarks.DENTRO,
        })

    def test_taxa_conversao_acima_de_100_e_excelente(self):
        # Conversas podem originar de visualizações sem clique — não é anomalia
        av = benchmarks.avaliar_metricas({"taxa_conversao": 250.0})
        self.assertEqual(av["taxa_conversao"], benchmarks.EXCELENTE)

    def test_override_de_faixas_por_chamada(self):
        self.assertEqual(benchmarks.avaliar_metricas({"ctr": 1.5}),
                         {"ctr": benchmarks.ABAIXO})
        av = benchmarks.avaliar_metricas({"ctr": 1.5}, faixas={"ctr": (1.0, 3.0)})
        self.assertEqual(av, {"ctr": benchmarks.DENTRO})

    def test_faixa_de_frequencia_para_retargeting(self):
        self.assertEqual(benchmarks.avaliar_metricas({"frequencia": 4.0}),
                         {"frequencia": benchmarks.ACIMA})
        av = benchmarks.avaliar_metricas({"frequencia": 4.0}, retargeting=True)
        self.assertEqual(av, {"frequencia": benchmarks.DENTRO})

    def test_metricas_ausentes_ficam_fora(self):
        self.assertEqual(benchmarks.avaliar_metricas({"ctr": None}), {})


class AnaliseAutomaticaTest(TestCase):
    """
    Integração do motor de regras (`analysis/`) com o parser: a conta
    individual passa pela avaliação, o consolidado ainda usa o texto por
    composição. As regras em si têm testes próprios em analysis/tests/.
    """

    def test_status_nao_vaza_para_analise_nem_para_o_pdf(self):
        f = _arquivo("conta.xlsx", [
            {"nome": "Campanha A", "res": 40, "inv": 80.0, "imp": 4000,
             "alc": 2500, "cliques": 300, "status": "inactive"},
            {"nome": "Campanha B", "res": 0, "inv": 60.0, "imp": 2000,
             "alc": 1000, "status": "not_delivering"},
        ])
        r = self.client.post("/", {"cliente": "ILOC", "arquivos": [f]})
        self.assertEqual(r.status_code, 302)

        analise = self.client.session["relatorio_apex"]["analise_sugerida"]
        for termo in ("inactive", "not_delivering", "inativ", "fora do ar",
                      "Ads Manager", "bloqueio", "reprova"):
            self.assertNotIn(termo, analise)

        # PDF gerado com a análise sugerida: nada de status ou auditoria
        r = self.client.post("/revisao/", {
            "cliente": "ILOC", "periodo": "01/07/2026 a 15/07/2026",
            "analise": analise,
        })
        texto = _texto_pdf(_bytes_pdf(r))
        for termo in ("inactive", "not_delivering", "Status", "fora do ar",
                      "verificando"):
            self.assertNotIn(termo, texto)

    def test_a_analise_mais_longa_ainda_cabe_em_uma_pagina(self):
        # A análise saiu de dois parágrafos curtos para quatro blocos, e o
        # relatório individual é de UMA página fechada. Este teste pega o mais
        # longo dos textos que o motor sabe produzir e leva até o PDF.
        f = _arquivo("conta.xlsx", [
            {"nome": "Campanha A", "res": 40, "inv": 80.0, "imp": 4000,
             "alc": 2500, "cliques": 300},
            {"nome": "Campanha B", "res": 20, "inv": 120.0, "imp": 2000,
             "alc": 1000, "cliques": 120},
        ])
        self.client.post("/", {"cliente": "ILOC", "arquivos": [f]})

        metricas = {"investimento": 644.98, "alcance": 16279,
                    "impressoes": 57965, "frequencia": 3.56, "cpm": 11.13,
                    "resultados": 344, "cpa": 1.87, "ctr": 0.5,
                    "campanhas": [{"nome": "a", "resultados": 344}]}
        textos = []
        for cpa in (1.87, 6.0, 30.0):
            for meta in (None, 5.00):
                for frequencia in (1.1, 2.0, 3.0, 4.0):
                    for cpm in (11.13, 25.0, 80.0):
                        av = analysis.rules.avaliar(
                            dict(metricas, cpa=cpa, frequencia=frequencia,
                                 cpm=cpm), meta_cpa=meta)
                        textos.append(analysis.templates.redigir(av, metricas))
        maior = max(textos, key=len)

        r = self.client.post("/revisao/", {
            "cliente": "ILOC", "periodo": "01/07/2026 a 15/07/2026",
            "analise": maior,
        })
        pdf = _bytes_pdf(r)
        self.assertEqual(_paginas(pdf), 1,
                         "a análise mais longa estourou a página do relatório")
        # E o texto chegou inteiro: o último bloco é o que cairia fora.
        self.assertIn("Objetivo do próximo ciclo", _texto_pdf(pdf))

    def test_quebra_de_bloco_sobrevive_ao_textarea(self):
        # O navegador manda \r\n (HTML spec). Separar por "\n\n" cru não acha
        # separador nenhum e a análise inteira sai num parágrafo só — era
        # assim que o bug aparecia no PDF, com os rótulos vermelhos correndo
        # no meio do texto.
        f = _arquivo("conta.xlsx", [
            {"nome": "Campanha A", "res": 40, "inv": 80.0, "imp": 4000,
             "alc": 2500, "cliques": 300}])
        self.client.post("/", {"cliente": "ILOC", "arquivos": [f]})
        analise = self.client.session["relatorio_apex"]["analise_sugerida"]
        blocos = analise.split("\n\n")
        self.assertGreaterEqual(len(blocos), 3)

        r = self.client.post("/revisao/", {
            "cliente": "ILOC", "periodo": "01/07/2026 a 15/07/2026",
            "analise": analise.replace("\n", "\r\n"),
        })
        texto = _texto_pdf(_bytes_pdf(r))
        # Cada bloco começa num parágrafo próprio: o rótulo do último não pode
        # aparecer grudado no fim do penúltimo.
        for rotulo in ("Leitura do período.", "O que vamos fazer.",
                       "Objetivo do próximo ciclo."):
            self.assertIn(rotulo, texto)
        self.assertEqual(len(_paragrafos_analise(analise.replace("\n", "\r\n"))),
                         len(blocos))

    def test_linha_em_branco_com_espaco_ainda_separa(self):
        # Operador que edita o texto à mão deixa espaço na linha vazia.
        self.assertEqual(_paragrafos_analise("Um.\r\n   \r\nDois."),
                         ["Um.", "Dois."])
        self.assertEqual(_paragrafos_analise("Um.\n\nDois.\n\n\nTrês."),
                         ["Um.", "Dois.", "Três."])

    def test_analise_da_conta_vem_do_motor_de_regras(self):
        # 80/20 entre duas campanhas, CPA de R$ 2,60: período ótimo, e o
        # próximo passo sai da concentração de resultados.
        dados = _dados([
            {"nome": "Campanha A", "res": 80, "inv": 160.0, "imp": 8000,
             "alc": 5000, "cliques": 400},
            {"nome": "Campanha B", "res": 20, "inv": 100.0, "imp": 2000,
             "alc": 1500, "cliques": 100},
        ])
        av = dados["avaliacao"]
        self.assertEqual(av["classificacao"], "OTIMO")
        self.assertIn("resultados_concentrados", av["sinais"])
        self.assertEqual(av["proximo_passo"], "redistribuir_verba")
        self.assertIn("Redistribuir a verba entre as campanhas",
                      dados["analise_sugerida"])

    def test_analise_do_pdf_nao_repete_os_numeros_das_tabelas(self):
        dados = _dados([
            {"nome": "Campanha A", "res": 50, "inv": 100.0, "imp": 5000,
             "alc": 4000, "cliques": 300},
            {"nome": "Campanha B", "res": 50, "inv": 110.0, "imp": 5000,
             "alc": 4000, "cliques": 300},
        ])
        analise = dados["analise_sugerida"]
        self.assertNotRegex(analise, r"\d")
        self.assertNotIn("R$", analise)
        # Frequência 1,25 em 15 dias equivale a 2,5 no mês: patamar saudável.
        self.assertIn("frequencia_saudavel", dados["avaliacao"]["sinais"])

    def test_cpa_alto_classifica_em_atencao_e_pede_revisao(self):
        # CTR 1% (abaixo da faixa 2–5%) e CPA de R$ 60,00 com 5 resultados
        dados = _dados([{"nome": "C", "res": 5, "inv": 300.0, "imp": 10000,
                         "alc": 8000, "cliques": 100}])
        _valor, leitura = _linhas_funil(dados)["CTR (taxa de cliques)"]
        self.assertEqual(
            leitura, "Estamos renovando os criativos para elevar a taxa de cliques.")

        av = dados["avaliacao"]
        self.assertEqual(av["classificacao"], "ATENCAO")
        # Amostra pequena não promove nada: o rebaixamento só desce.
        self.assertIn("amostra_pequena", av["sinais"])
        self.assertEqual(av["motivo_principal"], "cpa_atencao")
        self.assertIn("Revisar para quem os anúncios estão sendo mostrados",
                      dados["analise_sugerida"])
        for termo in ("cansado", "não está prendendo", "Baixo"):
            self.assertNotIn(termo, leitura + dados["analise_sugerida"])

    def test_avaliacao_serializavel_acompanha_os_dados(self):
        # É ela, não o texto, que vira payload nas etapas seguintes.
        dados = _dados([{"nome": "C", "res": 344, "inv": 644.98, "imp": 57965,
                         "alc": 16279, "cliques": 900}])
        self.assertEqual(json.loads(json.dumps(dados["avaliacao"])),
                         dados["avaliacao"])
        self.assertEqual(dados["avaliacao"]["perfil"], "varejo_celular")
        self.assertFalse(dados["avaliacao"]["meta_definida"])

    def test_meta_de_cpa_substitui_a_faixa_do_perfil(self):
        registros, mapa = parser_xlsx.ler_registros(io.BytesIO(_planilha(
            [{"nome": "C", "res": 100, "inv": 800.0, "imp": 20000,
              "alc": 12000, "cliques": 500}])))
        # CPA de R$ 8,00: dentro da faixa de varejo_celular (teto 9,00)...
        self.assertEqual(
            parser_xlsx.consolidar(registros, mapa)["avaliacao"]["classificacao"],
            "BOM")
        # ...e acima de uma meta de R$ 5,00 (razão 1,60).
        self.assertEqual(
            parser_xlsx.consolidar(registros, mapa, meta_cpa=5.0)
            ["avaliacao"]["classificacao"], "ATENCAO")

    def test_periodo_curto_endurece_a_leitura_de_frequencia(self):
        # Frequência 1,4: no mês inteiro é público de sobra; na semana, o
        # mesmo número já significa que a audiência viu bastante.
        campanhas = [{"nome": "C", "res": 100, "inv": 200.0, "imp": 14000,
                      "alc": 10000, "cliques": 500}]
        mes = parser_xlsx.ler_export_meta(io.BytesIO(
            _planilha(campanhas, inicio="2026-07-01", fim="2026-07-30")))
        semana = parser_xlsx.ler_export_meta(io.BytesIO(
            _planilha(campanhas, inicio="2026-07-01", fim="2026-07-07")))
        self.assertIn("frequencia_baixa", mes["avaliacao"]["sinais"])
        self.assertIn("frequencia_saturada", semana["avaliacao"]["sinais"])

    def test_ajuste_de_frequencia_nao_tem_degrau_em_catorze_dias(self):
        # Já teve: 14 dias usava os limites cheios e 13 usava 13/30 deles, e
        # um dia a mais no export virava o veredito de frequência do avesso.
        campanhas = [{"nome": "C", "res": 100, "inv": 200.0, "imp": 25000,
                      "alc": 10000, "cliques": 500}]
        sinais = []
        for fim in ("2026-07-13", "2026-07-14", "2026-07-15"):
            dados = parser_xlsx.ler_export_meta(io.BytesIO(
                _planilha(campanhas, inicio="2026-07-01", fim=fim)))
            sinais.append([s for s in dados["avaliacao"]["sinais"]
                           if s.startswith("frequencia_")])
        self.assertEqual(sinais[0], sinais[1])
        self.assertEqual(sinais[1], sinais[2])

    def test_taxa_de_conversao_acima_de_100_e_excelente(self):
        # 250 conversas com 100 cliques: conversas vindas de visualizações
        dados = _dados([{"nome": "C", "res": 250, "inv": 300.0, "imp": 10000,
                         "alc": 8000, "cliques": 100}])
        _valor, leitura = _linhas_funil(dados)["Taxa de Conversão (clique → conversa)"]
        self.assertIn("Excelente eficiência", leitura)

    def test_consolidado_classifica_taxas_recalculadas(self):
        # CPM da unidade A sozinha = R$ 100 (acima); recalculado sobre os
        # totais do grupo = 250/15000*1000 = R$ 16,67 (abaixo da faixa)
        grupo = consolidar_grupo([
            {"nome": "A", "dados": _dados([{"nome": "CA", "res": 10, "inv": 100.0,
                                            "imp": 1000, "alc": 800, "cliques": 50}])},
            {"nome": "B", "dados": _dados([{"nome": "CB", "res": 10, "inv": 100.0,
                                            "imp": 9000, "alc": 7000, "cliques": 200}])},
            {"nome": "C", "dados": _dados([{"nome": "CC", "res": 5, "inv": 50.0,
                                            "imp": 5000, "alc": 4000, "cliques": 100}])},
        ])
        _valor, leitura = _linhas_funil(grupo)["CPM (custo por mil)"]
        self.assertIn("Custo de entrega competitivo", leitura)

        # As três unidades têm o mesmo CPA (R$ 10,00), igual ao do grupo:
        # nenhuma descolada, e o grupo inteiro acima da faixa do perfil.
        av = grupo["avaliacao"]
        self.assertAlmostEqual(av["cpa_grupo"], 10.0, places=2)
        self.assertEqual(av["grupo"]["classificacao"], "ATENCAO")
        self.assertIn("grupo_homogeneo", av["sinais"])
        self.assertEqual(av["proximo_passo"], "elevar_o_patamar_do_grupo")
        self.assertIn("todas no mesmo patamar", grupo["analise_sugerida"])

    def test_consolidado_mede_cada_unidade_contra_o_custo_do_grupo(self):
        # A gasta R$ 400 por 10 resultados (CPA 40); B e C, R$ 50 por 50
        # (CPA 1). CPA do grupo = 500/110 = R$ 4,55 — A fica 8,8x acima.
        grupo = consolidar_grupo([
            {"nome": "Cara", "dados": _dados([{"nome": "CA", "res": 10, "inv": 400.0,
                                               "imp": 9000, "alc": 6000, "cliques": 300}])},
            {"nome": "Barata", "dados": _dados([{"nome": "CB", "res": 50, "inv": 50.0,
                                                 "imp": 5000, "alc": 4000, "cliques": 200}])},
            {"nome": "Média", "dados": _dados([{"nome": "CC", "res": 50, "inv": 50.0,
                                                "imp": 5000, "alc": 4000, "cliques": 200}])},
        ])
        av = grupo["avaliacao"]
        por_nome = {u["nome"]: u for u in av["unidades"]}
        self.assertEqual(por_nome["Cara"]["avaliacao"]["classificacao"], "ATENCAO")
        self.assertEqual(por_nome["Barata"]["avaliacao"]["classificacao"], "OTIMO")
        # A referência de cada unidade é o grupo, não a faixa do perfil.
        self.assertEqual(por_nome["Cara"]["avaliacao"]["referencia"], "grupo")
        self.assertIn("comparada_ao_grupo", por_nome["Cara"]["avaliacao"]["sinais"])

        self.assertIn("unidades_acima_do_grupo", av["sinais"])
        self.assertIn("unidades_abaixo_do_grupo", av["sinais"])
        self.assertIn("dispersao_alta", av["sinais"])
        self.assertEqual(av["proximo_passo"],
                         "levar_o_metodo_das_melhores_as_demais")

        analise = grupo["analise_sugerida"]
        self.assertIn("Cara", analise)      # a praça mais cara é nomeada
        self.assertIn("Barata", analise)    # e a que tem o método a copiar
        self.assertNotIn("R$", analise)     # os números estão na tabela acima

    def test_analise_do_grupo_sai_inteira_com_nomes_longos(self):
        # O bloco 2 do consolidado nomeia duas praças, e o nome vem do
        # operador. Com 20 unidades de nome comprido — todas listadas na
        # legenda desde 12/08/2026 — o relatório passa de uma página; o que
        # não pode é o texto sair cortado ou por cima do rodapé.
        nomes = ["Tim %02d — Maxi Shopping Jundiaí Zona Norte" % i
                 for i in range(20)]
        arquivos = [_arquivo("u%d.xlsx" % i, [
            {"nome": "C", "res": 100 + i * 20, "inv": 100.0 + i * 90,
             "imp": 9000, "alc": 4000, "cliques": 400}]) for i in range(20)]
        self.client.post("/", {"cliente": "TIM Brasil", "arquivos": arquivos})
        analise = self.client.session["relatorio_apex"]["analise_sugerida"]
        # A praça mais cara e a mais barata são nomeadas pelo nome do arquivo
        # até o operador renomear; o texto tem que caber de qualquer forma.
        campos = {"cliente": "TIM Brasil", "periodo": "01/07/2026 a 15/07/2026",
                  "analise": analise}
        campos.update({"unidade_%d" % i: n for i, n in enumerate(nomes)})
        r = self.client.post("/revisao/", campos)
        pdf = _bytes_pdf(r)
        texto = _texto_pdf(pdf)
        self.assertEqual(_paginas_com_rodape_invadido(pdf), [])
        self.assertIn("Objetivo do próximo ciclo", texto)
        self.assertIn(nomes[0], texto)
        self.assertIn(nomes[-1], texto)

    def test_avaliacao_do_grupo_e_serializavel(self):
        grupo = consolidar_grupo([
            {"nome": "A", "dados": _dados([{"nome": "CA", "res": 40, "inv": 80.0,
                                            "imp": 4000, "alc": 3000, "cliques": 200}])},
            {"nome": "B", "dados": _dados([{"nome": "CB", "res": 40, "inv": 90.0,
                                            "imp": 4000, "alc": 3000, "cliques": 200}])},
        ])
        self.assertEqual(json.loads(json.dumps(grupo["avaliacao"])),
                         grupo["avaliacao"])


class ValidacaoUploadTest(TestCase):
    def test_limite_de_vinte_arquivos(self):
        arquivos = [
            _arquivo(f"conta_{i}.xlsx", [{"nome": "C", "res": 1, "inv": 10.0,
                                          "imp": 100, "alc": 80}])
            for i in range(21)
        ]
        r = self.client.post("/", {"cliente": "Grupo", "arquivos": arquivos})
        self.assertEqual(r.status_code, 200)
        self.assertContains(r, "Máximo de 20 arquivos")

    def test_extensao_invalida_aponta_arquivo(self):
        arquivos = [
            _arquivo("ok.xlsx", [{"nome": "C", "res": 1, "inv": 10.0,
                                  "imp": 100, "alc": 80}]),
            SimpleUploadedFile("notas.txt", b"nao e planilha", content_type="text/plain"),
        ]
        r = self.client.post("/", {"cliente": "Grupo", "arquivos": arquivos})
        self.assertEqual(r.status_code, 200)
        self.assertContains(r, "notas.txt")

    def test_arquivo_corrompido_no_meio_do_lote_aponta_arquivo(self):
        ok = {"nome": "C", "res": 1, "inv": 10.0, "imp": 100, "alc": 80}
        arquivos = [
            _arquivo("a.xlsx", [ok]),
            SimpleUploadedFile("b.xlsx", b"bytes invalidos", content_type=XLSX_MIME),
            _arquivo("c.xlsx", [ok]),
        ]
        r = self.client.post("/", {"cliente": "Grupo", "arquivos": arquivos})
        self.assertEqual(r.status_code, 200)
        html = r.content.decode()
        self.assertIn("b.xlsx", html)
        self.assertNotIn("relatorio_apex", self.client.session)


class NomeDoArquivoTest(TestCase):
    """Padrão único nos quatro modos:
    '[empresa]-[funcionalidade]-[início]-[fim].pdf'. O nome sozinho diz de
    quem é o relatório, que relatório é e de que período — na pasta de
    downloads, meses depois, é só isso que sobra."""

    CONTA = {"nome": "C", "res": 40, "inv": 80.0, "imp": 4000, "alc": 2500}

    def _anexos(self, quantos, **kw):
        return [_arquivo(f"u{i}.xlsx", [dict(self.CONTA)], **kw)
                for i in range(quantos)]

    def _nome(self, resposta):
        return resposta["Content-Disposition"].split('filename="')[1].rstrip('"')

    def test_anexo_unico(self):
        self.client.post("/", {"cliente": "TIM BRASIL",
                               "arquivos": self._anexos(1)})
        r = self.client.post("/revisao/", {"cliente": "TIM BRASIL",
                                           "periodo": "01/07/2026 a 31/07/2026",
                                           "analise": ""})
        self.assertEqual(self._nome(r),
                         "TIM-BRASIL-anexounico-1-jul-26-31-jul-26.pdf")

    def test_consolidado(self):
        self.client.post("/", {"cliente": "TIM BRASIL",
                               "arquivos": self._anexos(3)})
        r = self.client.post("/revisao/", {"cliente": "TIM BRASIL",
                                           "periodo": "01/07/2026 a 31/07/2026",
                                           "analise": ""})
        self.assertEqual(self._nome(r),
                         "TIM-BRASIL-consolidado-1-jul-26-31-jul-26.pdf")

    def test_listagem(self):
        self.client.post("/", {"modo": "listagem", "cliente": "TIM BRASIL",
                               "arquivos": self._anexos(3)})
        r = self.client.post("/revisao/", {"cliente": "TIM BRASIL",
                                           "inicio": "2026-07-01",
                                           "fim": "2026-07-31"})
        self.assertEqual(self._nome(r),
                         "TIM-BRASIL-listagem-1-jul-26-31-jul-26.pdf")

    def test_indicador_unico(self):
        self.client.post("/", {"modo": "indicador", "cliente": "TIM BRASIL",
                               "metrica": "investimento_total",
                               "arquivos": self._anexos(3, inicio="2026-07-01",
                                                        fim="2026-07-31")})
        r = self.client.post("/revisao/", {"cliente": "TIM BRASIL",
                                           "metrica": "investimento_total"})
        self.assertEqual(self._nome(r),
                         "TIM-BRASIL-indicadorunico-1-jul-26-31-jul-26.pdf")

    def test_sem_periodo_marca_a_data_de_geracao(self):
        """Uma data só, sem rótulo, seria lida como início de intervalo."""
        self.client.post("/", {"modo": "listagem", "cliente": "TIM BRASIL",
                               "arquivos": self._anexos(2)})
        r = self.client.post("/revisao/", {"cliente": "TIM BRASIL"})
        hoje = date.today()
        self.assertEqual(
            self._nome(r),
            f"TIM-BRASIL-listagem-gerado-{hoje.day}-"
            f"{_MESES_PT[hoje.month - 1]}-{hoje:%y}.pdf")

    def test_acento_e_pontuacao_viram_hifen(self):
        self.client.post("/", {"cliente": "Grupo São José & Cia",
                               "arquivos": self._anexos(1)})
        r = self.client.post("/revisao/", {"cliente": "Grupo São José & Cia",
                                           "periodo": "", "analise": ""})
        self.assertTrue(self._nome(r).startswith("Grupo-Sao-Jose-Cia-anexounico-"),
                        self._nome(r))


class IndicadorDeResultadoTest(TestCase):
    """O export traz o indicador cru da API ("actions:post_engagement") e uma
    conta pode misturar objetivos. Vale para os quatro modos: o rótulo do PDF
    é o do indicador de maior volume, sempre traduzido."""

    def _pdf_unico(self, campanhas):
        self.client.post("/", {"cliente": "ILOC",
                               "arquivos": [_arquivo("a.xlsx", campanhas)]})
        return _texto_pdf(_bytes_pdf(self.client.post("/revisao/", {
            "cliente": "ILOC", "periodo": "", "analise": ""})))

    def test_indicador_dominante_vence_a_primeira_linha(self):
        """Caso real: a campanha de engajamento abre a planilha, mas responde
        por menos de um terço dos resultados."""
        texto = self._pdf_unico([
            {"nome": "Engaja", "res": 120, "inv": 100.0, "imp": 5000, "alc": 4000,
             "indicador": "actions:post_engagement"},
            {"nome": "Mensagens", "res": 305, "inv": 200.0, "imp": 9000, "alc": 7000,
             "indicador": "actions:onsite_conversion."
                          "messaging_conversation_started_7d"},
        ])
        self.assertIn("Conversas Iniciadas", texto)
        self.assertNotIn("Envolvimento com a Publicação", texto)

    def test_nenhum_codigo_de_api_chega_ao_cliente(self):
        for cru, esperado in (
                ("actions:post_engagement", "Envolvimento com a Publicação"),
                ("actions:lead", "Leads"),
                ("actions:link_click", "Cliques no Link"),
                ("actions:landing_page_view", "Visualizações da Página"),
                ("video_thruplay_watched_actions", "Reproduções de Vídeo")):
            texto = self._pdf_unico([{"nome": "C", "res": 10, "inv": 50.0,
                                      "imp": 1000, "alc": 800, "indicador": cru}])
            self.assertIn(esperado, texto)
            self.assertNotIn("actions:", texto)
            self.assertNotIn("_actions", texto)

    def test_linha_sem_resultado_nao_vota_no_indicador(self):
        """Campanha que gastou sem converter entra nos totais, mas não decide
        o rótulo — senão bastaria uma campanha zerada para renomear o PDF."""
        texto = self._pdf_unico([
            {"nome": "Sem resultado", "res": None, "inv": 300.0, "imp": 9000,
             "alc": 7000, "indicador": "actions:lead"},
            {"nome": "Mensagens", "res": 12, "inv": 40.0, "imp": 1000, "alc": 800,
             "indicador": "actions:onsite_conversion."
                          "messaging_conversation_started_7d"},
        ])
        self.assertIn("Conversas Iniciadas", texto)
        self.assertNotIn("Leads", texto)

    def test_indicador_desconhecido_cai_no_cru_e_avisa_no_log(self):
        with self.assertLogs("relatorios.indicadores", level="WARNING") as log:
            texto = self._pdf_unico([{"nome": "C", "res": 10, "inv": 50.0,
                                      "imp": 1000, "alc": 800,
                                      "indicador": "actions:objetivo_novo"}])
        self.assertEqual(len(log.output), 1, "um aviso por conta, não um por seção")
        self.assertIn("actions:objetivo_novo", log.output[0])
        self.assertIn("conta: a", log.output[0])       # de qual anexo veio
        self.assertIn("actions:objetivo_novo", texto)  # gera assim mesmo

    def test_rotulo_do_export_por_extenso_nao_vira_alarme(self):
        """Export em pt-BR já traz linguagem de cliente; não é código de API."""
        with patch("relatorios.indicadores.logger") as log:
            texto = self._pdf_unico([{"nome": "C", "res": 10, "inv": 50.0,
                                      "imp": 1000, "alc": 800,
                                      "indicador": "Compras"}])
        log.warning.assert_not_called()
        self.assertIn("Compras", texto)

    def test_consolidado_usa_o_indicador_dominante_do_grupo(self):
        conversa = ("actions:onsite_conversion."
                    "messaging_conversation_started_7d")
        arquivos = [
            _arquivo("grande.xlsx", [{"nome": "C", "res": 100, "inv": 200.0,
                                      "imp": 9000, "alc": 7000}],
                     indicador=conversa),
            _arquivo("pequena.xlsx", [{"nome": "C", "res": 8, "inv": 30.0,
                                       "imp": 900, "alc": 700}],
                     indicador="actions:lead"),
        ]
        self.client.post("/", {"cliente": "Grupo", "arquivos": arquivos})
        html = self.client.get("/revisao/").content.decode()
        self.assertIn("Conversas Iniciadas", html)
        # O aviso de divergência continua, mas em linguagem de gente
        self.assertIn("não usam o mesmo indicador", html)
        self.assertIn("Leads", html)
        self.assertNotIn("actions:", html)


class FluxoListagemTest(TestCase):
    """Modo 3: tabela com 1 linha por conta, na ordem de envio, sem
    consolidação, sem análise e sem destaque de melhor conta. Como os demais
    modos, passa pela revisão antes de gerar o PDF."""

    CONTAS = [
        ("unidade_centro.xlsx", 40, 80.0, 4000, 2500, 300),
        ("unidade_norte.xlsx", 10, 55.5, 2000, 900, 100),
        ("unidade_sul.xlsx", 25, 100.0, 5000, 3000, 250),
    ]

    def _upload(self, cliente="TIM Brasil", nomes=None):
        arquivos = [
            _arquivo(nome, [{"nome": "C", "res": res, "inv": inv,
                             "imp": imp, "alc": alc, "cliques": cli}])
            for nome, res, inv, imp, alc, cli in self.CONTAS
        ]
        post = {"modo": "listagem", "cliente": cliente, "arquivos": arquivos}
        if nomes:
            post["nome_conta"] = nomes
        return self.client.post("/", post)

    def _post_listagem(self, cliente="TIM Brasil", nomes=None):
        """Fluxo completo: painel → revisão → PDF."""
        self._upload(cliente, nomes)
        return self.client.post("/revisao/", {"cliente": cliente})

    def test_upload_leva_a_revisao_antes_do_pdf(self):
        r = self._upload()
        self.assertRedirects(r, "/revisao/")
        # A revisão mostra a prévia da tabela que vai ao PDF
        html = self.client.get("/revisao/").content.decode()
        self.assertIn("Revisar listagem", html)
        self.assertIn("unidade centro", html)

    def test_gera_pdf_na_revisao(self):
        r = self._post_listagem()
        self.assertEqual(r.status_code, 200)
        self.assertEqual(r["Content-Type"], "application/pdf")
        # O nome do arquivo sai do cliente, como nos outros três modos.
        self.assertIn("TIM-Brasil-listagem", r["Content-Disposition"])

    def test_nomes_editados_na_revisao_vao_para_o_pdf(self):
        self._upload()
        r = self.client.post("/revisao/", {
            "cliente": "TIM Brasil", "unidade_0": "Loja Centro",
            "unidade_1": "Loja Norte", "unidade_2": "Loja Sul"})
        texto = _texto_pdf(_bytes_pdf(r))
        self.assertIn("Loja Centro", texto)
        self.assertNotIn("unidade centro", texto)

    def test_cabecalho_traz_o_titulo_do_modo_e_o_cliente(self):
        """O título é do modo, o nome é do cliente — mesmo cabeçalho dos
        outros três modos. Até 24/08/2026 a listagem tinha título livre e não
        mostrava cliente nenhum."""
        r = self._post_listagem(cliente="Visão Geral — Franquias")
        texto = _texto_pdf(_bytes_pdf(r))
        self.assertIn("Relatório de Listagem", texto)
        # Caixa alta é do CSS (.sub .cliente), igual ao cabeçalho dos outros modos
        self.assertIn("VISÃO GERAL — FRANQUIAS", texto)
        self.assertIn("Visao-Geral-Franquias", r["Content-Disposition"])

    def test_painel_recusa_listagem_sem_cliente(self):
        """Antes o cliente era opcional na listagem — era o único modo sem
        nome nenhum no PDF."""
        r = self._upload(cliente="")
        self.assertEqual(r.status_code, 200)      # não redirecionou
        self.assertIn("Informe o cliente/grupo", r.content.decode())

    def test_linhas_ranqueadas_por_resultados(self):
        texto = _texto_pdf(_bytes_pdf(self._post_listagem()))
        # Enviadas como centro (40) → norte (10) → sul (25); o ranking
        # reordena para centro → sul → norte, independente da ordem de envio.
        pos = [texto.index(nome) for nome in
               ("unidade centro", "unidade sul", "unidade norte")]
        self.assertEqual(pos, sorted(pos))

    def test_empate_em_resultados_desempata_por_investimento(self):
        """Duas gerações da mesma base têm de produzir o mesmo PDF: sem
        desempate explícito, contas de igual volume trocariam de posição."""
        arquivos = [
            _arquivo("magra.xlsx", [{"nome": "C", "res": 30, "inv": 40.0,
                                     "imp": 2000, "alc": 1500}]),
            _arquivo("gorda.xlsx", [{"nome": "C", "res": 30, "inv": 90.0,
                                     "imp": 3000, "alc": 2000}]),
        ]
        self.client.post("/", {"modo": "listagem", "cliente": "TIM Brasil",
                               "arquivos": arquivos})
        texto = _texto_pdf(_bytes_pdf(
            self.client.post("/revisao/", {"cliente": "TIM Brasil"})))
        self.assertLess(texto.index("gorda"), texto.index("magra"))

    def test_valores_por_conta_em_pt_br_sem_consolidar(self):
        texto = _texto_pdf(_bytes_pdf(self._post_listagem()))
        self.assertIn("R$ 80,00", texto)      # investimento da 1ª conta
        self.assertIn("R$ 2,00", texto)       # custo/resultado: 80 / 40
        self.assertIn("R$ 5,55", texto)       # custo/resultado: 55,50 / 10
        self.assertIn("R$ 20,00", texto)      # CPM: 80 / 4000 * 1000
        self.assertIn("2.500", texto)         # alcance pt-BR da 1ª conta
        self.assertIn("Conversas Iniciadas", texto)   # label do resultado
        # Sem consolidação nem análise
        self.assertNotIn("R$ 235,50", texto)  # soma dos investimentos
        self.assertNotIn("Análise", texto)

    def test_ordem_das_colunas(self):
        """O tipo de resultado vem colado ao número dele, e a posição abre a
        linha — ler "Conversas Iniciadas / 40" seguido não exige varrer a
        tabela até a coluna certa."""
        texto = _texto_pdf(_bytes_pdf(self._post_listagem()))
        colunas = ["#", "Conta", "Resultado", "Nº Resultados",
                   "Investimento (R$)", "Custo/Resultado (R$)", "Alcance",
                   "Impressões", "CPM (R$)"]
        pos = [texto.index(c) for c in colunas]
        self.assertEqual(pos, sorted(pos), f"cabeçalho fora de ordem: {texto[:200]}")

    def test_sem_ctr_e_com_cpm(self):
        """O export de Campanhas não traz cliques: a coluna de CTR renderizava
        "—" em toda linha. Trocada por CPM, que sai dos totais já lidos."""
        texto = _texto_pdf(_bytes_pdf(self._post_listagem()))
        self.assertNotIn("CTR", texto)
        self.assertIn("CPM (R$)", texto)
        for cpm in ("R$ 20,00", "R$ 27,75"):   # centro/sul e norte
            self.assertIn(cpm, texto)

    def test_coluna_de_posicao_numera_o_ranking(self):
        texto = _texto_pdf(_bytes_pdf(self._post_listagem()))
        for posicao, nome in enumerate(
                ("unidade centro", "unidade sul", "unidade norte"), start=1):
            self.assertRegex(texto, rf"{posicao} {nome}")

    def test_conta_sem_impressoes_nao_estoura_no_cpm(self):
        arquivos = [_arquivo("sem_dados.xlsx",
                             [{"nome": "C", "res": 0, "inv": 12.0}]),
                    _arquivo("com_dados.xlsx",
                             [{"nome": "C", "res": 4, "inv": 20.0, "imp": 1000}])]
        self.client.post("/", {"modo": "listagem", "cliente": "TIM Brasil",
                               "arquivos": arquivos})
        texto = _texto_pdf(_bytes_pdf(
            self.client.post("/revisao/", {"cliente": "TIM Brasil"})))
        self.assertIn("R$ 20,00", texto)   # CPM da conta com impressões
        self.assertIn("—", texto)          # CPM e custo/resultado da outra

    def _upload_periodos(self):
        """Três anexos com janelas diferentes — o período do relatório é a
        união delas: 01/07 (menor início) a 31/07 (maior fim)."""
        arquivos = [
            _arquivo("centro.xlsx", [{"nome": "C", "res": 40, "inv": 80.0}],
                     inicio="2026-07-05", fim="2026-07-20"),
            _arquivo("norte.xlsx", [{"nome": "C", "res": 10, "inv": 55.5}],
                     inicio="2026-07-01", fim="2026-07-18"),
            _arquivo("sul.xlsx", [{"nome": "C", "res": 25, "inv": 100.0}],
                     inicio="2026-07-10", fim="2026-07-31"),
        ]
        return self.client.post("/", {"modo": "listagem", "cliente": "TIM Brasil",
                                      "arquivos": arquivos})

    def test_periodo_sugerido_a_partir_dos_anexos(self):
        self._upload_periodos()
        html = self.client.get("/revisao/").content.decode()
        # ISO no value: é o formato que o input nativo de data entende
        self.assertIn('value="2026-07-01"', html)
        self.assertIn('value="2026-07-31"', html)

    def test_periodo_editado_vai_para_o_cabecalho_do_pdf(self):
        self._upload_periodos()
        r = self.client.post("/revisao/", {"cliente": "TIM Brasil",
                                           "inicio": "2026-07-01",
                                           "fim": "2026-07-31"})
        self.assertIn("01/07/2026 — 31/07/2026", _texto_pdf(_bytes_pdf(r)))

    def test_periodo_em_branco_omite_o_bloco(self):
        self._upload_periodos()
        texto = _texto_pdf(_bytes_pdf(
            self.client.post("/revisao/", {"cliente": "TIM Brasil"})))
        self.assertNotRegex(texto, r"\d{2}/\d{2}/\d{4} — \d{2}/\d{2}/\d{4}")

    def test_meia_data_e_periodo_invertido_sao_recusados(self):
        for post in ({"inicio": "2026-07-01"},                       # sem fim
                     {"fim": "2026-07-31"},                          # sem início
                     {"inicio": "2026-07-31", "fim": "2026-07-01"}):  # invertido
            self._upload_periodos()
            r = self.client.post("/revisao/", dict(post, cliente="TIM Brasil"))
            self.assertEqual(r.status_code, 200)
            self.assertNotEqual(r["Content-Type"], "application/pdf")

    def test_modo_unico_exige_exatamente_um_arquivo(self):
        ok = {"nome": "C", "res": 1, "inv": 10.0, "imp": 100, "alc": 80}
        r = self.client.post("/", {
            "modo": "unico", "cliente": "ILOC",
            "arquivos": [_arquivo("a.xlsx", [ok]), _arquivo("b.xlsx", [ok])],
        })
        self.assertEqual(r.status_code, 200)
        self.assertContains(r, "exatamente 1 arquivo")

    def test_modo_consolidado_exige_dois_ou_mais(self):
        ok = {"nome": "C", "res": 1, "inv": 10.0, "imp": 100, "alc": 80}
        r = self.client.post("/", {
            "modo": "consolidado", "cliente": "Grupo",
            "arquivos": [_arquivo("a.xlsx", [ok])],
        })
        self.assertEqual(r.status_code, 200)
        self.assertContains(r, "pelo menos 2 arquivos")

    def test_modos_1_e_2_exigem_cliente(self):
        ok = {"nome": "C", "res": 1, "inv": 10.0, "imp": 100, "alc": 80}
        r = self.client.post("/", {
            "modo": "unico", "arquivos": [_arquivo("a.xlsx", [ok])]})
        self.assertEqual(r.status_code, 200)
        self.assertContains(r, "Informe o cliente")


# ----------------------------------------------------------------------
# Modo 4 — Indicador Único
# ----------------------------------------------------------------------
CABECALHO_SEM_CLIQUES = [c for c in CABECALHO if c != "Cliques no link"]


def _arquivo_sem_cliques(nome, campanhas, inicio="2026-07-01", fim="2026-07-15"):
    """Export em que a coluna de cliques não existe — usado para o caso de
    métrica indisponível numa das contas."""
    wb = Workbook()
    ws = wb.active
    ws.append(CABECALHO_SEM_CLIQUES)
    for c in campanhas:
        ws.append([c["nome"], "active", c.get("res"), INDICADOR, c.get("inv"),
                   c.get("imp"), c.get("alc"), inicio, fim])
    buf = io.BytesIO()
    wb.save(buf)
    return SimpleUploadedFile(nome, buf.getvalue(), content_type=XLSX_MIME)


# centro: CPA 2,00 · CTR 7,50%  |  norte: CPA 5,55 · CTR 5,00%
# sul:    CPA 4,00 · CTR 5,00%
# Totais: investimento 235,50 · resultados 75 · impressões 11.000 · cliques 650
CONTAS_INDICADOR = [
    ("unidade_centro.xlsx", 40, 80.0, 4000, 2500, 300),
    ("unidade_norte.xlsx", 10, 55.5, 2000, 900, 100),
    ("unidade_sul.xlsx", 25, 100.0, 5000, 3000, 250),
]


class FluxoIndicadorTest(TestCase):
    """Modo 4: PDF direto comparando UMA métrica entre contas — ordenado pela
    direção de `melhor` no registro, com total agregado conforme a regra da
    métrica (soma × recálculo sobre os brutos)."""

    def _upload(self, metrica, cliente="TIM Brasil", arquivos=None, nomes=None):
        if arquivos is None:
            arquivos = [
                _arquivo(nome, [{"nome": "C", "res": res, "inv": inv,
                                 "imp": imp, "alc": alc, "cliques": cli}])
                for nome, res, inv, imp, alc, cli in CONTAS_INDICADOR
            ]
        post = {"modo": "indicador", "cliente": cliente,
                "metrica": metrica, "arquivos": arquivos}
        if nomes:
            post["nome_conta"] = nomes
        return self.client.post("/", post)

    def _post(self, metrica, cliente="TIM Brasil", arquivos=None, **extra):
        """Fluxo completo: painel → revisão → PDF."""
        self._upload(metrica, cliente, arquivos)
        return self.client.post("/revisao/",
                                {"cliente": cliente, "metrica": metrica, **extra})

    def _texto(self, metrica, **kw):
        return _texto_pdf(_bytes_pdf(self._post(metrica, **kw)))

    def test_upload_leva_a_revisao_com_previa(self):
        r = self._upload("conversas_iniciadas")
        self.assertRedirects(r, "/revisao/")
        html = self.client.get("/revisao/").content.decode()
        self.assertIn("Revisar indicador", html)
        self.assertIn("75", html)          # total já calculado na prévia

    def test_gera_pdf_na_revisao(self):
        r = self._post("conversas_iniciadas")
        self.assertEqual(r.status_code, 200)
        self.assertEqual(r["Content-Type"], "application/pdf")
        self.assertIn('filename="TIM-Brasil-indicadorunico-1-jul-26-15-jul-26.pdf"',
                      r["Content-Disposition"])

    def test_trocar_metrica_na_revisao_nao_exige_reenviar_anexos(self):
        self._upload("conversas_iniciadas")
        r = self.client.post("/revisao/", {"cliente": "TIM", "metrica": "cpa"})
        texto = _texto_pdf(_bytes_pdf(r))
        self.assertIn("R$ 3,14", texto)    # CPA geral recalculado
        self.assertIn("Custo por Resultado", texto)

    def test_cookie_sinaliza_download_quando_ha_token(self):
        # O front manda um token; o PDF volta com esse token num cookie, sinal
        # de que o arquivo saiu — é o que fecha a etapa 02 na tela.
        self._upload("cpa")
        r = self.client.post("/revisao/", {"cliente": "TIM", "metrica": "cpa",
                                           "download_token": "abc123"})
        self.assertEqual(r.status_code, 200)
        self.assertEqual(r.cookies["apex_download"].value, "abc123")

    def test_sem_token_nao_ha_cookie(self):
        r = self._post("cpa")
        self.assertNotIn("apex_download", r.cookies)

    def test_nome_digitado_no_painel_substitui_o_nome_do_arquivo(self):
        # Cada anexo carrega o nome da conta digitado no painel; é ele que vai
        # ao PDF, não o "unidade_centro" derivado do arquivo.
        self._upload("conversas_iniciadas",
                     nomes=["Loja Centro", "Loja Norte", "Loja Sul"])
        r = self.client.post("/revisao/",
                             {"cliente": "TIM", "metrica": "conversas_iniciadas"})
        texto = _texto_pdf(_bytes_pdf(r))
        self.assertIn("Loja Centro", texto)
        self.assertNotIn("unidade centro", texto)

    def test_nome_em_branco_cai_no_nome_do_arquivo(self):
        # Campo vazio → mantém o comportamento antigo (nome derivado do arquivo).
        self._upload("conversas_iniciadas", nomes=["", "Loja Norte", ""])
        r = self.client.post("/revisao/",
                             {"cliente": "TIM", "metrica": "conversas_iniciadas"})
        texto = _texto_pdf(_bytes_pdf(r))
        self.assertIn("unidade centro", texto)   # fallback do arquivo
        self.assertIn("Loja Norte", texto)       # digitado

    def test_nome_editado_na_revisao_vence_o_do_painel(self):
        self._upload("conversas_iniciadas", nomes=["Loja Centro", "N", "S"])
        r = self.client.post("/revisao/", {
            "cliente": "TIM", "metrica": "conversas_iniciadas",
            "unidade_0": "Centro Revisado", "unidade_1": "N", "unidade_2": "S"})
        texto = _texto_pdf(_bytes_pdf(r))
        self.assertIn("Centro Revisado", texto)
        self.assertNotIn("Loja Centro", texto)

    def test_metrica_aditiva_soma_e_mostra_share(self):
        texto = self._texto("conversas_iniciadas")
        self.assertIn("Conversas Iniciadas", texto)
        self.assertIn("75", texto)          # total = 40 + 10 + 25
        self.assertIn("53,3%", texto)       # share de centro
        self.assertIn("13,3%", texto)       # share de norte

    def test_metrica_de_taxa_recalcula_total_e_omite_share(self):
        texto = self._texto("cpa")
        # CPA geral = 235,50 / 75 = 3,14 — e NÃO a média das CPAs (3,85)
        self.assertIn("R$ 3,14", texto)
        self.assertNotIn("R$ 3,85", texto)
        self.assertIn("não é a média dos valores individuais", texto)
        self.assertNotIn("% do total", texto)

    def test_ctr_geral_recalculado_sobre_os_brutos(self):
        texto = self._texto("ctr")
        self.assertIn("5,91%", texto)       # 650 / 11.000
        self.assertNotIn("5,83%", texto)    # média dos CTRs das contas

    def test_ordem_segue_a_direcao_de_melhor(self):
        # conversas → "maior" é melhor: centro (40) · sul (25) · norte (10)
        texto = self._texto("conversas_iniciadas")
        pos = [texto.index(n) for n in ("unidade centro", "unidade sul", "unidade norte")]
        self.assertEqual(pos, sorted(pos))

        # CPA → "menor" é melhor: centro (2,00) · sul (4,00) · norte (5,55)
        texto = self._texto("cpa")
        pos = [texto.index(n) for n in ("unidade centro", "unidade sul", "unidade norte")]
        self.assertEqual(pos, sorted(pos))

    def test_metrica_sem_ranking_mantem_ordem_de_envio(self):
        # investimento_total tem melhor=None → ordem dos anexos, não por valor
        texto = self._texto("investimento_total")
        pos = [texto.index(n) for n in ("unidade centro", "unidade norte", "unidade sul")]
        self.assertEqual(pos, sorted(pos))

    def test_nome_da_unidade_igual_ao_modo_listagem(self):
        # Mesma origem (nome do arquivo, via views._nome_unidade) nos dois modos
        indicador = self._texto("conversas_iniciadas")
        arquivos = [
            _arquivo(nome, [{"nome": "C", "res": res, "inv": inv,
                             "imp": imp, "alc": alc, "cliques": cli}])
            for nome, res, inv, imp, alc, cli in CONTAS_INDICADOR
        ]
        self.client.post("/", {"modo": "listagem", "cliente": "TIM Brasil",
                               "arquivos": arquivos})
        listagem = _texto_pdf(_bytes_pdf(
            self.client.post("/revisao/", {"cliente": "TIM Brasil"})))
        for nome in ("unidade centro", "unidade norte", "unidade sul"):
            self.assertIn(nome, indicador)
            self.assertIn(nome, listagem)

    def test_conta_sem_a_coluna_fica_fora_do_total(self):
        arquivos = [
            _arquivo("centro.xlsx", [{"nome": "C", "res": 40, "inv": 80.0,
                                      "imp": 4000, "alc": 2500, "cliques": 300}]),
            _arquivo_sem_cliques("norte.xlsx", [{"nome": "C", "res": 10,
                                                 "inv": 55.5, "imp": 2000,
                                                 "alc": 900}]),
        ]
        texto = self._texto("ctr", arquivos=arquivos)
        self.assertIn("Dado indisponível no export de: norte", texto)
        # CTR geral = 300 / 4.000 — a conta sem cliques não entra na conta
        self.assertIn("7,50%", texto)
        self.assertNotIn("6,67%", texto)   # 300 / 6.000, se norte entrasse

    def test_periodo_em_hifen(self):
        texto = self._texto("conversas_iniciadas")
        self.assertIn("01-07 a 15-07", texto)
        self.assertNotIn("01/07/2026", texto)

    def test_aviso_de_objetivo_divergente_so_em_metrica_sensivel(self):
        # Cada POST consome os uploads: os anexos são remontados a cada chamada
        def anexos():
            return [
                _arquivo("centro.xlsx", [{"nome": "C", "res": 40, "inv": 80.0,
                                          "imp": 4000, "alc": 2500, "cliques": 300}]),
                _arquivo("norte.xlsx", [{"nome": "C", "res": 10, "inv": 55.5,
                                         "imp": 2000, "alc": 900, "cliques": 100}],
                         indicador="Cliques no link"),
            ]

        self.assertIn("objetivos diferentes", self._texto("cpa", arquivos=anexos()))
        # Métrica de topo não depende do objetivo → sem aviso
        self.assertNotIn("objetivos diferentes",
                         self._texto("impressoes", arquivos=anexos()))

    def test_exige_dois_arquivos_e_metrica(self):
        ok = {"nome": "C", "res": 1, "inv": 10.0, "imp": 100, "alc": 80}
        r = self.client.post("/", {"modo": "indicador", "cliente": "G",
                                   "metrica": "cpa",
                                   "arquivos": [_arquivo("a.xlsx", [ok])]})
        self.assertContains(r, "pelo menos 2 arquivos")

        r = self.client.post("/", {
            "modo": "indicador", "cliente": "G",
            "arquivos": [_arquivo("a.xlsx", [ok]), _arquivo("b.xlsx", [ok])]})
        self.assertContains(r, "Escolha a métrica")

    def test_seletor_agrupado_por_estagio_no_painel(self):
        html = self.client.get("/").content.decode()
        self.assertIn("Topo de Funil", html)
        self.assertIn("Meio de Funil", html)
        self.assertIn("Fundo de Funil", html)
        self.assertIn('value="taxa_conversao"', html)


class RegistroMetricasTest(TestCase):
    """O registro é a fonte única: acrescentar uma métrica não toca view,
    template nem regra de agregação."""

    def test_agregacao_de_taxa_nunca_e_media_das_medias(self):
        # 2 contas de volumes muito diferentes: a média das CPAs (5,50) fica
        # longe do CPA real do grupo (100 / 55 = 1,82)
        nums = [{"investimento": 20.0, "resultados": 50.0},
                {"investimento": 80.0, "resultados": 5.0}]
        self.assertAlmostEqual(metricas.total_geral("cpa", nums), 100 / 55)

    def test_soma_e_direta_para_metrica_aditiva(self):
        nums = [{"investimento": 20.0}, {"investimento": 80.0}]
        self.assertEqual(metricas.total_geral("investimento_total", nums), 100.0)

    def test_denominador_zero_vira_indisponivel(self):
        nums = [{"investimento": 20.0, "resultados": 0.0}]
        self.assertIsNone(metricas.total_geral("cpa", nums))
        self.assertEqual(metricas.formatar("cpa", None), "—")

    def test_metrica_nova_so_com_uma_entrada_no_registro(self):
        nova = dict(metricas.METRICS_REGISTRY)
        nova["custo_por_mil_alcancados"] = {
            "label": "Custo por Mil Alcançados", "unidade": "R$",
            "estagio": "topo", "agregacao": "recalculo",
            "formula": "investimento_total / alcance * 1000",
            "melhor": "menor", "fonte": None,
        }
        with patch.object(metricas, "METRICS_REGISTRY", nova):
            # 1) aparece no seletor, no grupo certo
            topo = dict(metricas.opcoes_agrupadas())["Topo de Funil — Atração"]
            self.assertIn(("custo_por_mil_alcancados", "Custo por Mil Alcançados"),
                          topo)
            # 2) agrega sozinha: 235,50 / 6.400 * 1000 = 36,80
            nums = [{"investimento": 80.0, "alcance": 2500.0},
                    {"investimento": 55.5, "alcance": 900.0},
                    {"investimento": 100.0, "alcance": 3000.0}]
            self.assertAlmostEqual(
                metricas.total_geral("custo_por_mil_alcancados", nums),
                235.5 / 6400 * 1000)
            # 3) gera o PDF sem nenhuma alteração em view/template/agregação
            arquivos = [
                _arquivo(nome, [{"nome": "C", "res": res, "inv": inv,
                                 "imp": imp, "alc": alc, "cliques": cli}])
                for nome, res, inv, imp, alc, cli in CONTAS_INDICADOR
            ]
            self.client.post("/", {
                "modo": "indicador", "cliente": "TIM",
                "metrica": "custo_por_mil_alcancados", "arquivos": arquivos})
            r = self.client.post("/revisao/", {
                "cliente": "TIM", "metrica": "custo_por_mil_alcancados"})
            self.assertEqual(r["Content-Type"], "application/pdf")
            texto = _texto_pdf(_bytes_pdf(r))
            self.assertIn("Custo por Mil Alcançados", texto)
            self.assertIn("R$ 36,80", texto)



class SelecaoDeCampanhasTest(TestCase):
    """Seleção dos grupos de campanha na revisão do consolidado e da listagem.

    Os anexos de um mesmo cliente trazem produtos diferentes (celular, ultra) e
    o operador quer o relatório de um deles. A escolha mora na revisão porque é
    só depois de ler os anexos que os grupos existem; aplicar refaz a leitura
    dos registros guardados na sessão, sem reenviar arquivo.
    """

    CELULAR, ULTRA = "LEADS · CELULAR", "LEADS · ULTRA"

    # centro e norte têm os dois produtos; sul só tem celular.
    # Celular: 65 resultados / R$ 130,00 · Ultra: 15 / R$ 50,00
    CONTAS = [
        ("centro.xlsx", [
            {"nome": "[LEADS][CELULAR][CENTRO][ABO][01JUN26]", "res": 30,
             "inv": 60.0, "imp": 3000, "alc": 2000, "cliques": 200},
            {"nome": "[LEADS][ULTRA][ABO][24JUL26]", "res": 10, "inv": 40.0,
             "imp": 1000, "alc": 800, "cliques": 90},
        ]),
        ("norte.xlsx", [
            {"nome": "[LEADS][CELULAR][NORTE][ABO][01JUN26]", "res": 20,
             "inv": 40.0, "imp": 2000, "alc": 1500, "cliques": 150},
            {"nome": "[LEADS][ULTRA][ABO][28JUL26]", "res": 5, "inv": 10.0,
             "imp": 500, "alc": 400, "cliques": 40},
        ]),
        ("sul.xlsx", [
            {"nome": "[LEADS][CELULAR][SUL][ABO][01JUN26]", "res": 15,
             "inv": 30.0, "imp": 1500, "alc": 1000, "cliques": 120},
        ]),
    ]

    def _upload(self, modo="consolidado", contas=None, **extra):
        arquivos = [_arquivo(nome, campanhas)
                    for nome, campanhas in (contas or self.CONTAS)]
        post = {"modo": modo, "cliente": "TIM Brasil", "arquivos": arquivos}
        post.update(extra)
        return self.client.post("/", post)

    def _aplicar(self, chaves, **extra):
        post = {"cliente": "TIM Brasil", "aplicar_campanhas": "1",
                "campanhas": chaves}
        post.update(extra)
        return self.client.post("/revisao/", post)

    def _sessao(self):
        return self.client.session["relatorio_apex"]

    def _linhas_funil(self, dados):
        return {m: v for etapa in dados["funil"]["etapas"]
                for m, v, _ in etapa["linhas"]}

    # ---- a regra de agrupamento -------------------------------------
    def test_grupo_sai_dos_dois_primeiros_colchetes(self):
        # Região e data variam entre unidades do mesmo produto e ficam de fora
        self.assertEqual(
            parser_xlsx.chave_grupo_campanha("[LEADS][CELULAR-BOLETO][SALTO][ABO][13JUL26]"),
            parser_xlsx.chave_grupo_campanha("[LEADS][CELULAR-BOLETO][ITU][ABO][01SET25]"))
        self.assertNotEqual(
            parser_xlsx.chave_grupo_campanha("[LEADS][CELULAR-BOLETO][ABO][01JUN26]"),
            parser_xlsx.chave_grupo_campanha("[LEADS][ULTRA][ABO][24JUL26]"))
        # O objetivo entra na chave: mesmo produto com outro objetivo é outro grupo
        self.assertNotEqual(parser_xlsx.chave_grupo_campanha("[LEADS][ULTRA][ABO]"),
                            parser_xlsx.chave_grupo_campanha("[VENDAS][ULTRA][ABO]"))

    def test_nome_fora_do_padrao_vira_grupo_dele_mesmo(self):
        self.assertEqual(parser_xlsx.chave_grupo_campanha("Campanha de julho"),
                         "Campanha de julho")
        self.assertEqual(parser_xlsx.chave_grupo_campanha("[SOZINHO]"), "[SOZINHO]")
        self.assertEqual(parser_xlsx.chave_grupo_campanha(""),
                         parser_xlsx.GRUPO_SEM_NOME)

    def test_sem_chaves_nada_e_filtrado(self):
        # Mesma convenção de VEICULACAO_TODAS: é o que mantém o fluxo sem
        # seleção idêntico ao que sempre foi.
        registros = [{"campanha": "[A][B][C]"}, {"campanha": "[A][D][C]"}]
        self.assertEqual(parser_xlsx.filtrar_campanhas(registros, None), registros)
        self.assertEqual(parser_xlsx.filtrar_campanhas(registros, []), registros)

    def test_datas_do_export_vao_serializaveis_para_a_sessao(self):
        # A sessão grava JSON e `date` não é serializável: as colunas de
        # período são normalizadas para texto ISO na leitura.
        self._upload()
        json.dumps(self._sessao())      # levanta TypeError se algo escapar
        registro = self._sessao()["_anexos"][0]["registros"][0]
        self.assertEqual(registro["inicio"], "2026-07-01")

    # ---- consolidado -------------------------------------------------
    def test_a_tela_lista_os_grupos_com_as_campanhas_dentro(self):
        self._upload()
        html = self.client.get("/revisao/").content.decode()
        self.assertIn("Campanhas incluídas", html)
        self.assertIn(self.CELULAR, html)
        self.assertIn(self.ULTRA, html)
        # As campanhas ficam à vista para o agrupamento poder ser conferido
        self.assertIn("[LEADS][CELULAR][CENTRO][ABO][01JUN26]", html)
        self.assertIn("3 anexos", html)

    def test_um_grupo_so_nao_abre_a_selecao(self):
        # Sem escolha a fazer, a caixa marcada sozinha seria ruído
        self._upload(contas=[(nome, [c for c in campanhas if "ULTRA" not in c["nome"]])
                             for nome, campanhas in self.CONTAS])
        html = self.client.get("/revisao/").content.decode()
        self.assertNotIn("Campanhas incluídas", html)
        self.assertNotIn('name="aplicar_campanhas"', html)

    def test_selecionar_um_grupo_refaz_todos_os_numeros(self):
        self._upload()
        self.assertEqual(self._linhas_funil(self._sessao())["Investimento Total"],
                         "R$ 180,00")

        self._aplicar([self.CELULAR])
        linhas = self._linhas_funil(self._sessao())
        self.assertEqual(linhas["Investimento Total"], "R$ 130,00")
        self.assertEqual(linhas["Conversas Iniciadas"], "65")
        # CPA recalculado sobre os brutos do recorte: 130 / 65 = 2,00
        self.assertEqual(linhas["Custo por Conversa (CPA)"], "R$ 2,00")
        self.assertEqual(self._sessao()["_selecao_campanhas"], [self.CELULAR])

    def test_anexo_sem_o_grupo_sai_do_relatorio(self):
        self._upload()
        r = self._aplicar([self.ULTRA])
        dados = self._sessao()
        self.assertEqual([u["nome"] for u in dados["unidades"]], ["centro", "norte"])
        self.assertEqual(self._linhas_funil(dados)["Investimento Total"], "R$ 50,00")
        self.assertIn("1 anexo", r.content.decode())
        # O anexo continua guardado: remarcar o grupo o traz de volta
        self.assertEqual(len(dados["_anexos"]), 3)
        self._aplicar([self.CELULAR, self.ULTRA])
        self.assertEqual(len(self._sessao()["unidades"]), 3)

    def test_selecao_que_esvazia_o_consolidado_e_recusada(self):
        # Só o centro tem ultra: o consolidado ficaria com uma unidade
        contas = [self.CONTAS[0], self.CONTAS[2]]
        self._upload(contas=contas)
        r = self._aplicar([self.ULTRA])
        self.assertIn("menos de 2 unidades", r.content.decode())
        # A seleção anterior continua valendo — nada foi trocado na sessão
        self.assertEqual(self._linhas_funil(self._sessao())["Investimento Total"],
                         "R$ 130,00")

    def test_desmarcar_tudo_e_recusado(self):
        self._upload()
        r = self._aplicar([])
        self.assertIn("Marque pelo menos um grupo", r.content.decode())
        self.assertEqual(len(self._sessao()["unidades"]), 3)

    def test_nome_digitado_antes_de_aplicar_nao_se_perde(self):
        self._upload()
        self._aplicar([self.CELULAR, self.ULTRA], unidade_0="Loja Centro")
        self.assertEqual(self._sessao()["unidades"][0]["nome"], "Loja Centro")
        # E segue valendo depois de um filtro que reordena as unidades
        self._aplicar([self.ULTRA])
        self.assertEqual(self._sessao()["unidades"][0]["nome"], "Loja Centro")

    def test_texto_da_ia_e_descartado_ao_trocar_a_selecao(self):
        # Foi escrito sobre outros números: mantê-lo seria oferecer ao cliente
        # a leitura de um relatório que não é mais este.
        self._upload()
        s = self.client.session
        dados = s["relatorio_apex"]
        dados["analise_ia"] = "TEXTO ESCRITO PELA IA"
        s["relatorio_apex"] = dados
        s.save()

        r = self._aplicar([self.CELULAR])
        self.assertNotIn("analise_ia", self._sessao())
        self.assertNotIn("TEXTO ESCRITO PELA IA", r.content.decode())
        self.assertIn("Relatório refeito com as campanhas selecionadas",
                      r.content.decode())

    def test_cliente_e_periodo_digitados_sobrevivem(self):
        self._upload()
        self._aplicar([self.CELULAR], cliente="TIM Interior",
                      periodo="01/07/2026 a 15/07/2026")
        dados = self._sessao()
        self.assertEqual(dados["cliente"], "TIM Interior")
        self.assertEqual(dados["periodo"], "01/07/2026 a 15/07/2026")

    def test_o_pdf_sai_com_a_selecao_aplicada(self):
        self._upload()
        self._aplicar([self.ULTRA])
        r = self.client.post("/revisao/", {"cliente": "TIM Brasil",
                                           "periodo": "01/07/2026 a 15/07/2026",
                                           "analise": "Texto."})
        texto = _texto_pdf(_bytes_pdf(r))
        self.assertIn("R$ 50,00", texto)
        self.assertIn("Consolidado de 2 unidades", texto)
        self.assertIn("Unidades incluídas no consolidado: centro, norte.", texto)

    # ---- listagem ----------------------------------------------------
    def test_listagem_reordena_com_o_grupo_escolhido(self):
        self._upload(modo="listagem")
        # Com tudo: centro (40) → norte (25) → sul (15)
        self._aplicar([self.ULTRA])
        contas = self._sessao()["contas"]
        self.assertEqual([c["nome"] for c in contas], ["centro", "norte"])
        r = self.client.post("/revisao/", {"cliente": "TIM Brasil"})
        texto = _texto_pdf(_bytes_pdf(r))
        self.assertIn("R$ 40,00", texto)     # centro só com a ultra
        self.assertIn("2 contas", texto)
        self.assertNotIn("R$ 30,00", texto)  # a sul, sem ultra, ficou fora

    def test_listagem_sem_nenhum_anexo_no_grupo_e_recusada(self):
        self._upload(modo="listagem", contas=[self.CONTAS[2]])
        # Um grupo só: a tela nem oferece a seleção, e um POST forjado com
        # grupo inexistente não derruba nada
        r = self._aplicar(["INEXISTENTE"])
        self.assertEqual(r.status_code, 200)
        self.assertEqual(len(self._sessao()["contas"]), 1)

    # ---- indicador único ---------------------------------------------
    # O indicador foi o último modo a ganhar a seleção (24/08/2026), no lugar
    # do filtro por status de veiculação: recortar por produto é o que o
    # cliente reconhece, e o status a tabela já mostra.
    def _upload_indicador(self, contas=None, metrica="conversas_iniciadas"):
        return self._upload(modo="indicador", contas=contas, metrica=metrica)

    def test_indicador_guarda_os_registros_na_sessao(self):
        # Precisa deles para refazer a leitura com outra seleção, como os
        # demais modos — antes o indicador era o único que não guardava.
        self._upload_indicador()
        self.assertIn("_anexos", self._sessao())

    def test_indicador_lista_os_grupos_e_oferece_o_botao(self):
        self._upload_indicador()
        html = self.client.get("/revisao/").content.decode()
        self.assertIn("Campanhas incluídas", html)
        self.assertIn(self.CELULAR, html)
        self.assertIn(self.ULTRA, html)
        self.assertIn('name="aplicar_campanhas"', html)

    def test_indicador_refaz_os_numeros_com_o_grupo_escolhido(self):
        self._upload_indicador(metrica="investimento_total")
        self._aplicar([self.CELULAR], metrica="investimento_total")
        contas = self._sessao()["contas"]
        # centro 60 + norte 40 + sul 30 = 130 (sem os 50 da ultra)
        self.assertEqual(
            sum(c["dados"]["_num"]["investimento"] for c in contas), 130.0)
        self.assertEqual(self._sessao()["_selecao_campanhas"], [self.CELULAR])

    def test_indicador_conta_sem_o_grupo_sai_da_comparacao(self):
        self._upload_indicador(metrica="investimento_total")
        r = self._aplicar([self.ULTRA], metrica="investimento_total")
        self.assertEqual([c["nome"] for c in self._sessao()["contas"]],
                         ["centro", "norte"])
        self.assertIn("1 anexo", r.content.decode())

    def test_indicador_a_metrica_sobrevive_ao_aplicar(self):
        self._upload_indicador(metrica="investimento_total")
        self._aplicar([self.CELULAR], metrica="cpa")
        self.assertEqual(self._sessao()["metrica"], "cpa")

    def test_indicador_o_pdf_sai_com_a_selecao_aplicada(self):
        self._upload_indicador(metrica="investimento_total")
        self._aplicar([self.ULTRA], metrica="investimento_total")
        r = self.client.post("/revisao/", {"cliente": "TIM Brasil",
                                           "metrica": "investimento_total"})
        texto = _texto_pdf(_bytes_pdf(r))
        self.assertIn("R$ 50,00", texto)     # só a ultra: 40 + 10
        # A sul não anunciou ultra: saiu da comparação, e com ela os R$ 30,00
        self.assertNotIn("R$ 30,00", texto)

    # ---- anexo único -------------------------------------------------
    # Uma planilha só também traz produtos diferentes: o recorte por grupo é o
    # mesmo dos outros modos, sobre um anexo em vez de vários.
    def _upload_unico(self, campanhas=None):
        arquivo = _arquivo("centro.xlsx", campanhas or self.CONTAS[0][1])
        return self.client.post("/", {"modo": "unico", "cliente": "TIM Brasil",
                                      "arquivos": [arquivo]})

    def test_unico_lista_os_grupos_do_anexo(self):
        self._upload_unico()
        html = self.client.get("/revisao/").content.decode()
        self.assertIn("Campanhas incluídas", html)
        self.assertIn(self.CELULAR, html)
        self.assertIn(self.ULTRA, html)
        self.assertIn("[LEADS][CELULAR][CENTRO][ABO][01JUN26]", html)
        # "em 1 anexo" seria ruído: não há anexo nenhum com que comparar.
        self.assertNotIn("em 1 anexo", html)

    def test_unico_com_um_grupo_so_nao_abre_a_selecao(self):
        self._upload_unico([c for c in self.CONTAS[0][1] if "ULTRA" not in c["nome"]])
        html = self.client.get("/revisao/").content.decode()
        self.assertNotIn("Campanhas incluídas", html)
        self.assertNotIn('name="aplicar_campanhas"', html)

    def test_unico_selecionar_um_grupo_refaz_os_numeros(self):
        self._upload_unico()
        self.assertEqual(self._linhas_funil(self._sessao())["Investimento Total"],
                         "R$ 100,00")

        self._aplicar([self.CELULAR])
        linhas = self._linhas_funil(self._sessao())
        self.assertEqual(linhas["Investimento Total"], "R$ 60,00")
        self.assertEqual(linhas["Conversas Iniciadas"], "30")
        # CPA recalculado sobre os brutos do recorte: 60 / 30 = 2,00
        self.assertEqual(linhas["Custo por Conversa (CPA)"], "R$ 2,00")
        self.assertEqual(self._sessao()["_selecao_campanhas"], [self.CELULAR])

    def test_unico_a_tabela_por_campanha_perde_a_campanha_filtrada(self):
        # No anexo único a tabela de campanhas vai ao PDF: deixar nela uma
        # campanha que saiu dos totais é a contradição que o cliente enxerga.
        self._upload_unico()
        self._aplicar([self.CELULAR])
        nomes = [l[0] for l in self._sessao()["detalhes_campanha"]["linhas"]]
        self.assertEqual(nomes, ["[LEADS][CELULAR][CENTRO][ABO][01JUN26]"])

    def test_unico_grupo_inexistente_no_post_nao_derruba_nada(self):
        # Grupo que não está entre as opções não passa da validação do form —
        # o POST forjado volta a tela, sem tocar na sessão.
        self._upload_unico()
        r = self._aplicar(["INEXISTENTE"])
        self.assertEqual(r.status_code, 200)
        self.assertEqual(self._linhas_funil(self._sessao())["Investimento Total"],
                         "R$ 100,00")

    def test_unico_desmarcar_tudo_e_recusado(self):
        self._upload_unico()
        r = self._aplicar([])
        self.assertIn("Marque pelo menos um grupo", r.content.decode())
        self.assertEqual(self._linhas_funil(self._sessao())["Investimento Total"],
                         "R$ 100,00")

    def test_unico_cliente_e_periodo_digitados_sobrevivem(self):
        self._upload_unico()
        self._aplicar([self.CELULAR], cliente="TIM Interior",
                      periodo="01/07/2026 a 15/07/2026")
        dados = self._sessao()
        self.assertEqual(dados["cliente"], "TIM Interior")
        self.assertEqual(dados["periodo"], "01/07/2026 a 15/07/2026")

    def test_unico_texto_da_ia_e_descartado_ao_trocar_a_selecao(self):
        self._upload_unico()
        s = self.client.session
        dados = s["relatorio_apex"]
        dados["analise_ia"] = "TEXTO ESCRITO PELA IA"
        s["relatorio_apex"] = dados
        s.save()

        r = self._aplicar([self.CELULAR])
        self.assertNotIn("analise_ia", self._sessao())
        self.assertNotIn("TEXTO ESCRITO PELA IA", r.content.decode())
        self.assertIn("Relatório refeito com as campanhas selecionadas",
                      r.content.decode())

    def test_unico_o_pdf_sai_com_a_selecao_aplicada(self):
        self._upload_unico()
        self._aplicar([self.ULTRA])
        r = self.client.post("/revisao/", {"cliente": "TIM Brasil",
                                           "periodo": "01/07/2026 a 15/07/2026",
                                           "analise": "Texto."})
        texto = _texto_pdf(_bytes_pdf(r))
        self.assertIn("R$ 40,00", texto)
        self.assertNotIn("[LEADS][CELULAR][CENTRO][ABO][01JUN26]", texto)


class AmbienteDeProducaoTest(SimpleTestCase):
    """
    O padrão do `settings` é PRODUÇÃO; quem liga o desenvolvimento é o
    `manage.py`.

    Na VPS as variáveis vêm de /etc/apex-reports/env pelo systemd. O arquivo
    inteiro sumir já derrubava o serviço (`EnvironmentFile` sem `-`), mas uma
    variável apagada à mão passava batido: a aplicação subia com DEBUG ligado
    e com a chave de exemplo, que está publicada no repositório. Agora o
    caminho de produção se recusa a subir sem as variáveis, e o caminho de
    desenvolvimento continua não precisando de nenhuma.

    O comportamento é de tempo de import, então cada caso roda num processo
    limpo, sem nenhuma DJANGO_* herdada de quem chamou a suíte.
    """

    RAIZ = Path(__file__).resolve().parent.parent

    def _rodar(self, argumentos, **ambiente):
        env = {k: v for k, v in os.environ.items() if not k.startswith("DJANGO_")}
        env.update(ambiente, DJANGO_SETTINGS_MODULE="apex_reports.settings")
        return subprocess.run([sys.executable] + argumentos, cwd=str(self.RAIZ),
                              env=env, capture_output=True, text=True)

    def _wsgi(self, **ambiente):
        """Sobe a aplicação como o gunicorn sobe — sem passar pelo manage.py."""
        return self._rodar(["-c", "import apex_reports.wsgi;"
                            "from django.conf import settings;"
                            "print(settings.DEBUG, settings.ALLOWED_HOSTS)"],
                           **ambiente)

    def test_producao_sem_variaveis_recusa_subir(self):
        r = self._wsgi()
        self.assertNotEqual(r.returncode, 0, "subiu sem SECRET_KEY nenhuma")
        self.assertIn("DJANGO_SECRET_KEY", r.stderr)

    def test_producao_com_as_variaveis_sobe_travada(self):
        r = self._wsgi(DJANGO_SECRET_KEY="chave-de-verdade", DJANGO_DEBUG="0",
                       DJANGO_ALLOWED_HOSTS="203.0.113.10,localhost")
        self.assertEqual(r.returncode, 0, r.stderr)
        self.assertEqual(r.stdout.strip(),
                         "False ['203.0.113.10', 'localhost']")

    def test_valor_desconhecido_de_debug_cai_em_desligado(self):
        """Errar o valor tem que errar para o lado seguro — "sim" não liga."""
        for valor in ("", "sim", "0", "false"):
            with self.subTest(valor=valor):
                r = self._wsgi(DJANGO_SECRET_KEY="k", DJANGO_DEBUG=valor)
                self.assertTrue(r.stdout.startswith("False"), r.stdout or r.stderr)

    def test_manage_py_continua_rodando_sem_configurar_nada(self):
        """A promessa do README: `python manage.py runserver` e pronto."""
        r = self._rodar(["manage.py", "check"])
        self.assertEqual(r.returncode, 0, r.stderr)

    def test_deploy_escreve_toda_variavel_que_o_settings_le(self):
        """Guarda contra a deriva entre os dois arquivos: acrescentar uma
        variável ao settings sem escrevê-la no deploy derruba a VPS no próximo
        publish, e só lá.

        Vale também para as OPENAI_*: o deploy reescreve o env inteiro a cada
        publicação, então uma variável que ele não conheça é uma variável que
        o próximo `make deploy` apaga."""
        def variaveis(caminho):
            return set(re.findall(r"(?:DJANGO|OPENAI)_[A-Z_]+",
                                  (self.RAIZ / caminho).read_text(encoding="utf-8")))

        lidas = variaveis("apex_reports/settings.py")
        escritas = variaveis("deploy/deploy.sh")
        self.assertTrue(lidas, "nenhuma variável encontrada no settings")
        self.assertEqual(lidas - escritas, set())


class ComentarioDeTemplateTest(SimpleTestCase):
    """
    `{# … #}` do Django NÃO atravessa linha.

    Um comentário aberto numa linha e fechado noutra não é comentário: é texto,
    e sai impresso — na tela de revisão e, pior, dentro do PDF que vai ao
    cliente. Aconteceu duas vezes em 12/08/2026, nos dois templates. Comentário
    de várias linhas se escreve com um par por linha, ou com
    `{% comment %}`.
    """

    TEMPLATES = Path(__file__).resolve().parent / "templates" / "relatorios"

    def test_nenhum_comentario_de_template_atravessa_linha(self):
        soltos = []
        for arquivo in sorted(self.TEMPLATES.glob("*.html")):
            dentro = False
            for n, linha in enumerate(
                    arquivo.read_text(encoding="utf-8").splitlines(), 1):
                abre, fecha = linha.count("{#"), linha.count("#}")
                if dentro or abre != fecha:
                    soltos.append(f"{arquivo.name}:{n}: {linha.strip()[:60]}")
                dentro = (abre > fecha) if abre != fecha else dentro
        self.assertEqual(soltos, [], "comentário {# #} sem fechar na mesma "
                                     "linha — ele SAI IMPRESSO na página:\n"
                                     + "\n".join(soltos))


class ArquivoEnvTest(SimpleTestCase):
    """
    O `.env` de desenvolvimento — a chave da OpenAI num arquivo, em vez de
    numa variável de terminal que morre ao fechar a janela.

    A regra que não pode quebrar é a precedência: variável que já existe no
    ambiente vence o arquivo. É ela que mantém o `manage.py migrate` do deploy
    rodando com as variáveis de produção, numa máquina que por acaso tenha um
    `.env` esquecido na raiz.
    """

    RAIZ = Path(__file__).resolve().parent.parent

    def _carregar(self, conteudo, ambiente=None):
        from apex_reports.settings import carregar_env
        with tempfile.TemporaryDirectory() as pasta:
            caminho = Path(pasta) / ".env"
            caminho.write_text(conteudo, encoding="utf-8")
            with patch.dict(os.environ, ambiente or {}, clear=False):
                antes = set(os.environ)
                carregar_env(caminho)
                lidas = {k: os.environ[k] for k in set(os.environ) - antes}
                for k in lidas:                       # não vaza para os outros
                    del os.environ[k]
        return lidas

    def test_le_par_por_linha(self):
        self.assertEqual(self._carregar("APEX_TESTE_A=1\nAPEX_TESTE_B=dois\n"),
                         {"APEX_TESTE_A": "1", "APEX_TESTE_B": "dois"})

    def test_ignora_comentario_linha_vazia_e_aspas(self):
        lidas = self._carregar('# comentário\n\nAPEX_TESTE_C="com aspas"\n'
                               "export APEX_TESTE_D='exportado'\n")
        self.assertEqual(lidas, {"APEX_TESTE_C": "com aspas",
                                 "APEX_TESTE_D": "exportado"})

    def test_valor_com_igual_nao_e_cortado(self):
        """Chave de API com `=` no meio não pode chegar truncada — o erro
        apareceria só na primeira chamada, como "chave inválida"."""
        self.assertEqual(self._carregar("APEX_TESTE_E=sk-abc=def==\n"),
                         {"APEX_TESTE_E": "sk-abc=def=="})

    def test_ambiente_vence_o_arquivo(self):
        lidas = self._carregar("APEX_TESTE_F=do-arquivo\n",
                               {"APEX_TESTE_F": "do-ambiente"})
        self.assertEqual(lidas, {})
        self.assertNotIn("APEX_TESTE_F", os.environ)

    def test_arquivo_ausente_nao_e_erro(self):
        from apex_reports.settings import carregar_env
        carregar_env(self.RAIZ / "nao-existe-mesmo.env")   # não levanta

    def test_env_fica_fora_do_repositorio(self):
        """Uma chave no repositório é uma chave pública — mesma regra da
        SECRET_KEY. O `.env.example` é o que fica versionado."""
        ignorados = (self.RAIZ / ".gitignore").read_text(encoding="utf-8").split()
        self.assertIn(".env", ignorados)
        self.assertTrue((self.RAIZ / ".env.example").exists())

    def test_o_exemplo_cita_toda_variavel_que_o_settings_le(self):
        """Mesmo guarda do deploy.sh, do lado do desenvolvimento: variável
        nova sem linha no exemplo é variável que ninguém descobre existir."""
        def variaveis(caminho, padrao=r"(?:DJANGO|OPENAI)_[A-Z_]+"):
            return set(re.findall(padrao,
                                  (self.RAIZ / caminho).read_text(encoding="utf-8")))

        self.assertEqual(variaveis("apex_reports/settings.py")
                         - variaveis(".env.example"), set())


class SuperficieExpostaTest(TestCase):
    """
    A aplicação publica duas rotas, e nada mais.

    O `/admin` do projeto recém-criado sobreviveu até aqui sem ter o que
    administrar — não existe um único modelo. Na prática era um segundo
    formulário de login exposto em HTTP puro: sem domínio não há HTTPS, e quem
    protege o acesso é o basic auth do nginx.
    """

    def test_admin_nao_existe(self):
        self.assertEqual(self.client.get("/admin/").status_code, 404)

    def test_as_rotas_publicadas_sao_so_estas(self):
        from apex_reports.urls import urlpatterns
        self.assertEqual([str(p.pattern) for p in urlpatterns],
                         ["", "revisao/"])

    def test_so_ficam_instalados_os_apps_usados(self):
        """`auth`, `contenttypes` e `messages` vieram do `startproject` e nunca
        foram ligados: nenhum login, nenhum modelo, nenhum template lendo a
        fila de mensagens. Traziam tabela e middleware para não fazer nada."""
        from django.conf import settings
        self.assertEqual(list(settings.INSTALLED_APPS), [
            "django.contrib.sessions",
            "django.contrib.staticfiles",
            "relatorios",
        ])

    def test_o_banco_existe_so_para_a_sessao(self):
        """Um `migrate` num banco novo cria só a tabela de sessão (mais o
        controle de migrações do próprio Django)."""
        from django.db import connection
        self.assertEqual(sorted(connection.introspection.table_names()),
                         ["django_migrations", "django_session"])
