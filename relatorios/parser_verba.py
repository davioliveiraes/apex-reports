# -*- coding: utf-8 -*-
"""
Leitor do export do preset `VERBA` do Gerenciador de Anúncios.

Módulo separado do `parser_xlsx` de propósito, e não por organização: o preset
`VERBA` traz colunas que **colidem** com as do export de desempenho. `Início` e
`Término` aqui são as datas de configuração da campanha; lá, o recorte do
relatório ("Início dos relatórios"). Ler os dois com o mesmo mapa faria uma
campanha que subiu dia 17 parecer ter começado no dia 1º — que é justamente o
erro que este módulo existe para evitar (ver `docs/GUIA_VERBA.md`).

Orçamento não é métrica, é configuração: ele só existe no nível em que foi
definido. Campanha `[CBO]` guarda o valor nela mesma; campanha `[ABO]` guarda
"Usando o orçamento do conjunto de anúncios" e o valor está no outro export.
Por isso a coleta são dois arquivos, e por isso o cruzamento é **por ID**:
nome de campanha é renomeado no meio do mês e o merge perde linhas sem avisar.
"""

import re
from datetime import date, datetime

from openpyxl import load_workbook

from .parser_xlsx import _fmt_data, _mapear_colunas, _norm, _to_float

# Mesmo formato de `parser_xlsx._COLUNAS`: chave → (alternativas, proibidos).
# `_mapear_colunas` casa nome exato primeiro e só depois "contém", então
# "Orçamento" ganha de "Tipo de orçamento" sem depender da ordem das colunas —
# os proibidos são o cinto de segurança para o dia em que o Meta renomear uma
# delas (o guia avisa que isso acontece).
_COLUNAS_VERBA = {
    "campanha_id":    ([["identificacao da campanha"], ["campaign id"]], []),
    "campanha":       ([["nome da campanha"], ["campaign name"]], []),
    "conjunto_id":    ([["identificacao", "conjunto"], ["ad set id"]], []),
    "conjunto":       ([["nome do conjunto"], ["ad set name"]], []),
    "orcamento":      ([["orcamento"], ["budget"]], ["tipo", "type"]),
    "tipo_orcamento": ([["tipo de orcamento"], ["budget type"]], []),
    "lances":         ([["estrategia de lances"], ["bid strategy"]], []),
    # Proibidos afastam as colunas homônimas do export de desempenho, caso o
    # operador tenha exportado com as duas predefinições misturadas.
    # A data em que a CAMPANHA foi configurada para começar. Continua
    # excluindo "relatório"/"reporting" — as duas datas convivem no mesmo
    # arquivo desde 29/08/2026 e confundi-las faria uma campanha que subiu dia
    # 17 parecer ter começado no dia 1º.
    "inicio":         ([["inicio"], ["starts"], ["start time"]],
                       ["relatorio", "reporting"]),
    "termino":        ([["termino"], ["encerramento"], ["ends"], ["end time"]],
                       ["relatorio", "reporting"]),
    "objetivo":       ([["objetivo"], ["objective"]], []),
    "veiculacao":     ([["veiculacao"], ["delivery"]], []),
    "gasto":          ([["valor gasto"], ["valor usado"], ["amount spent"]], []),
    # O recorte do RELATÓRIO — o intervalo escolhido no seletor de datas. É
    # ele que diz a que período o `Valor gasto` se refere, e sem ele o app não
    # tinha como conferir o passo que mais dá erro na coleta.
    "inicio_relatorio":  ([["inicio dos relatorios"], ["reporting starts"]], []),
    "termino_relatorio": ([["termino dos relatorios"],
                           ["encerramento dos relatorios"],
                           ["reporting ends"]], []),
}

# Valores de `Veiculação` que contam como estrutura no ar. Tudo o que não está
# aqui conta como fora — inclusive "em análise" e "programada", que ainda não
# gastam. O valor cru vai para a tabela de conferência da tela: o operador
# precisa poder discordar do que o app decidiu, e para isso precisa ver.
VEICULACOES_ATIVAS = frozenset((
    "active", "ativa", "ativo", "ativas", "ativos",
    "veiculando", "em veiculacao", "delivering", "publicado",
))

# Texto que o Meta escreve na célula de orçamento da campanha quando o valor
# mora um nível abaixo. É o marcador de `[ABO]`, e a única forma de saber que o
# segundo arquivo é obrigatório para esta conta.
_HERDA_DO_CONJUNTO = ("orcamento do conjunto", "ad set budget",
                      "usando o orcamento do conjunto")

_NUMERO = re.compile(r"\d[\d.,]*")


# ----------------------------------------------------------------------
# Leitura de um arquivo
# ----------------------------------------------------------------------
def ler_planilha_verba(arquivo):
    """As linhas de um export do preset `VERBA`.

    Devolvia também o nível deduzido das colunas, quando a coleta eram dois
    arquivos e o operador não dizia qual era qual. Hoje o nível é declarado na
    tela e conferido em `ler_arquivo_verba` — deduzir virou conferir.
    """
    wb = load_workbook(arquivo, data_only=True, read_only=True)
    ws = wb.active
    linhas_brutas = list(ws.iter_rows(values_only=True))
    wb.close()
    if not linhas_brutas:
        raise ValueError("A planilha está vazia.")

    cabecalho, mapa = None, {}
    for i, linha in enumerate(linhas_brutas[:10]):
        m = _mapear_colunas(linha, _COLUNAS_VERBA)
        if "orcamento" in m or "gasto" in m:
            cabecalho, mapa = i, m
            break
    if cabecalho is None:
        raise ValueError(
            "Não foi possível reconhecer as colunas de verba nesta planilha. "
            "Aplique a predefinição VERBA em Colunas → Personalizar colunas "
            "antes de exportar (ver docs/GUIA_VERBA.md)."
        )

    linhas = []
    for bruta in linhas_brutas[cabecalho + 1:]:
        if bruta is None or all(c is None or str(c).strip() == "" for c in bruta):
            continue
        reg = {}
        for chave, idx in mapa.items():
            valor = bruta[idx] if idx < len(bruta) else None
            # A sessão serializa em JSON e não sabe gravar `date`.
            if isinstance(valor, date):
                valor = valor.strftime("%Y-%m-%d")
            reg[chave] = valor
        nome_ref = _norm(reg.get("conjunto") or reg.get("campanha"))
        if nome_ref.startswith(("total", "resultados de")):
            continue
        linhas.append(reg)

    if not linhas:
        raise ValueError("Nenhuma linha de dados encontrada na planilha.")
    return linhas


NIVEL_CAMPANHA = "campanha"
NIVEL_CONJUNTO = "conjunto"


def tem_conjunto(linhas):
    """A planilha traz coluna de conjunto **preenchida**?

    A checagem é pelo valor, não pela presença da coluna: a predefinição VERBA
    pede as duas identificações, então o export de campanha também traz a
    coluna de conjunto — vazia.
    """
    return any(str(r.get("conjunto_id") or r.get("conjunto") or "").strip()
               for r in linhas)


def ler_arquivo_verba(arquivo, nivel):
    """`(linhas, erro)` para UM export, conferido contra o nível declarado.

    Até 29/08/2026 a coleta eram dois arquivos e o nível de cada um era
    deduzido das colunas. Deduzir deixou de fazer sentido quando o orçamento
    saiu da conta: o que o fechamento precisa é do gasto, e o gasto existe nos
    dois níveis. Um arquivo basta — o do nível em que a conta está montada —, e
    quem sabe qual é esse nível é o operador, não o arquivo.

    A declaração não é confiada às cegas: `[CBO]` com export de conjunto (ou o
    contrário) é o engano provável aqui, e ele soma o gasto no nível errado
    sem reclamar de nada.
    """
    try:
        linhas = ler_planilha_verba(arquivo)
    except ValueError as e:
        return None, f'Arquivo "{arquivo.name}": {e}'
    except Exception:
        return None, (
            f'Não foi possível ler "{arquivo.name}". Confira se é o .xlsx '
            "exportado direto do Gerenciador de Anúncios.")

    conjunto = tem_conjunto(linhas)
    if nivel == NIVEL_CONJUNTO and not conjunto:
        return None, (
            'Você marcou <b>ABO</b>, mas este arquivo é o export de '
            '<b>campanha</b> — nenhuma linha traz conjunto de anúncios. Em '
            "ABO o orçamento e o gasto que interessam estão no nível conjunto: "
            "troque para a aba <i>Conjuntos de anúncios</i> e exporte de novo.")
    if nivel == NIVEL_CAMPANHA and conjunto:
        return None, (
            'Você marcou <b>CBO</b>, mas este arquivo é o export de '
            '<b>conjunto de anúncios</b>. Em CBO o orçamento está na campanha: '
            "troque para a aba <i>Campanhas</i> e exporte de novo.")
    return linhas, None



# ----------------------------------------------------------------------
# Orçamento: valor, periodicidade e herança
# ----------------------------------------------------------------------
def partir_orcamento(celula, tipo=None):
    """`(valor, periodicidade, herdado)` da célula de orçamento.

    O export traz valor e periodicidade grudados — `R$ 33,00 Diário`,
    `R$ 1.000,00 Vitalício` —, então o número sai por regex e não por
    `_to_float` direto, que engasgaria no texto ao lado. A coluna *Tipo de
    orçamento*, quando veio, tem a última palavra sobre a periodicidade.

    `herdado=True` é a campanha `[ABO]`: o valor não está aqui, está no export
    de conjunto de anúncios.
    """
    texto = "" if celula is None else str(celula).strip()
    normal = _norm(texto)
    if any(marca in normal for marca in _HERDA_DO_CONJUNTO):
        return None, None, True

    achado = _NUMERO.search(texto)
    valor = _to_float(achado.group(0)) if achado else _to_float(celula)
    if valor is None:
        return None, None, False
    return valor, _periodicidade(tipo, normal), False


def _periodicidade(tipo, texto_normal):
    """"vitalicio" ou "diario" — a fórmula de projeção depende disso."""
    for fonte in (_norm(tipo), texto_normal):
        if "vitalic" in fonte or "lifetime" in fonte:
            return "vitalicio"
        if "diari" in fonte or "daily" in fonte:
            return "diario"
    return "diario"


def ativa(veiculacao):
    """A estrutura está no ar? Só o que está entra no configurado diário."""
    return _norm(veiculacao) in VEICULACOES_ATIVAS


# ----------------------------------------------------------------------
# Cruzamento dos dois níveis
# ----------------------------------------------------------------------
def _data(valor):
    """Texto ISO ou dd/mm/aaaa → `date`; qualquer outra coisa → None."""
    if isinstance(valor, (datetime, date)):
        return valor.date() if isinstance(valor, datetime) else valor
    texto = str(valor or "").strip()[:10]
    for formato in ("%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(texto, formato).date()
        except ValueError:
            continue
    return None


def _diario_equivalente(valor, periodicidade, inicio, termino, dias_do_mes):
    """`(valor diário, precisou do mês)`.

    Vitalício não é diário: dividi-lo pelo período em que ele vale é o que
    permite compará-lo com o contratado. Sem data de término não há período,
    e aí o mês de referência é o palpite menos ruim — sinalizado, para o
    operador saber que aquele número é uma conversão e não uma leitura.
    """
    if valor is None or periodicidade != "vitalicio":
        return valor, False
    i, f = _data(inicio), _data(termino)
    if i and f and f >= i:
        return valor / ((f - i).days + 1), False
    return valor / dias_do_mes, True


def montar_estruturas(linhas, nivel=NIVEL_CAMPANHA, dias_do_mes=30):
    """`(estruturas, avisos)` — uma estrutura por linha do export.

    Sem cruzamento: o arquivo é do nível em que a conta está montada, e cada
    linha dele já é uma estrutura completa — nome, veiculação, início, gasto e
    o orçamento que o Meta tem setado.

    **O orçamento aqui não entra em conta nenhuma.** Desde 29/08/2026 o diário
    do fechamento é o contratado dividido pelos dias do ciclo (R$ 300/semana =
    R$ 43/dia), e não a soma do que está configurado. O valor lido continua
    saindo na tabela de conferência porque é ele que denuncia Meta setado em
    R$ 20 quando o contrato pede R$ 43 — mas denuncia para o operador ler, não
    para o cálculo usar.
    """
    avisos, estruturas, convertidos = [], [], 0
    conjunto = nivel == NIVEL_CONJUNTO

    for linha in linhas:
        valor, periodicidade, herdado = partir_orcamento(
            linha.get("orcamento"), linha.get("tipo_orcamento"))
        orcamento, convertido = _diario_equivalente(
            valor, periodicidade, linha.get("inicio"), linha.get("termino"),
            dias_do_mes)
        convertidos += 1 if convertido else 0

        nome = str(linha.get("conjunto" if conjunto else "campanha") or "").strip()
        estruturas.append({
            "id": str(linha.get("conjunto_id" if conjunto else "campanha_id")
                      or "").strip(),
            "nome": nome or "(sem nome)",
            # No nível conjunto a campanha vira o contexto da linha: sem ela a
            # tabela lista dez conjuntos sem dizer de quem são.
            "campanha": (str(linha.get("campanha") or "").strip()
                         if conjunto else ""),
            "objetivo": str(linha.get("objetivo") or "").strip(),
            "lances": str(linha.get("lances") or "").strip(),
            "veiculacao": str(linha.get("veiculacao") or "").strip() or "—",
            "ativa": ativa(linha.get("veiculacao")),
            "inicio": linha.get("inicio"),
            "termino": linha.get("termino"),
            "inicio_br": _fmt_data(linha.get("inicio")) or "—",
            "tipo": "ABO" if conjunto else "CBO",
            "gasto": _to_float(linha.get("gasto")) or 0.0,
            "orcamento": orcamento,
            "orcamento_bruto": (str(linha.get("orcamento") or "").strip()
                                or ("herdado" if herdado else "—")),
        })

    if convertidos:
        avisos.append(
            f"{convertidos} estrutura(s) com orçamento vitalício sem data de "
            f"término: o valor foi dividido pelos {dias_do_mes} dias do mês "
            "para virar equivalente diário na tabela de conferência.")
    return estruturas, avisos



def periodo_relatado(linhas):
    """`(início, término, erro)` do intervalo que o export declara.

    Sem as colunas, devolve `(None, None, erro)`: não há como seguir. É
    preferível recusar o arquivo a projetar um gasto de período desconhecido.
    """
    inicio, termino = _extremos_do_relatorio(linhas)
    if not (inicio and termino):
        return None, None, (
            "O export não traz as colunas de período do relatório "
            "(`Início dos relatórios` e `Encerramento dos relatórios`). "
            "Acrescente as duas à predefinição VERBA e exporte de novo — ver "
            "docs/GUIA_VERBA.md. Sem elas o app não sabe a que intervalo o "
            "gasto se refere.")
    return inicio, termino, None


def _extremos_do_relatorio(linhas):
    """`(menor início, maior término)` entre as linhas, ou `(None, None)`.

    Menor e maior, e não o valor da primeira linha: planilha editada à mão
    existe, e nesse caso o intervalo real é a união do que houver nas linhas,
    não o que estiver no topo.
    """
    inicios = [d for d in (_data(l.get("inicio_relatorio")) for l in linhas) if d]
    terminos = [d for d in (_data(l.get("termino_relatorio")) for l in linhas) if d]
    if not (inicios and terminos):
        return None, None
    return min(inicios), max(terminos)
