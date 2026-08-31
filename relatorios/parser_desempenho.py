# -*- coding: utf-8 -*-
"""
Leitor do export do preset `DESEMPENHO` do Gerenciador de Anúncios.

Módulo separado do `parser_xlsx` pelo mesmo motivo do `parser_verba`: é outro
recorte de colunas, com outras colisões. O preset `DESEMPENHO` traz quatro
colunas cujos nomes começam iguais — `Resultados`, `Resultados (iniciais)`,
`Indicador de resultados` e `Indicador de resultados (inicial)` — e casá-las
por "contém" na ordem errada faria a análise ler a coluna vazia e concluir que
o mês não teve resultado nenhum. O mapa abaixo blinda as quatro.

**O que este preset NÃO traz:** valor gasto, cliques, CTR, CPC, visualizações
de página, ROAS, retenção de vídeo. Nada aqui pode depender delas. O
`analise_desempenho` respeita isso, e é a razão de ele existir em vez de
reaproveitar o motor da Análise Geral, que lê investimento em toda conta.

O nível do export é o de **conjuntos de anúncios**, e o arquivo pode trazer de
um a muitos. Uma linha só é o caso comum de conta pequena, não o caso especial.
"""

from datetime import date

from openpyxl import load_workbook

from .parser_xlsx import _mapear_colunas, _norm, _to_float

# Mesmo formato de `parser_xlsx._COLUNAS`: chave → (alternativas, proibidos).
# `_mapear_colunas` casa nome exato antes de "contém", e é isso que separa
# `Resultados` de `Resultados (iniciais)` sem depender da ordem das colunas.
# Os proibidos são o cinto de segurança para o dia em que o Meta mudar a
# grafia de uma delas e o passe exato deixar de bater.
_COLUNAS_DESEMPENHO = {
    "inicio":  ([["inicio dos relatorios"], ["reporting starts"]], []),
    "termino": ([["encerramento dos relatorios"], ["termino dos relatorios"],
                 ["reporting ends"]], []),
    # O export pode vir no nível de campanha OU de conjunto — o preset é o
    # mesmo, o que muda é a aba de onde ele saiu. As duas colunas são mapeadas
    # e a análise usa a que existir (ver `analise_desempenho.unidades`). Até
    # 30/08/2026 só o conjunto era lido, e num arquivo de campanhas TODA linha
    # ficava sem nome — a origem dos "Conjunto 1", "Conjunto 2" no texto.
    "campanha":   ([["nome da campanha"], ["campaign name"]], []),
    "conjunto":   ([["nome do conjunto"], ["ad set name"]], []),
    "veiculacao": ([["veiculacao"], ["delivery"]], []),
    # "inicia" cobre "(iniciais)" e "(inicial)" de uma vez.
    "resultados": ([["resultados"], ["results"]],
                   ["custo", "cost", "indicador", "indicator", "inicia",
                    "tipo", "type", "taxa", "rate"]),
    "indicador":  ([["indicador de resultados"], ["result indicator"]],
                   ["inicia"]),
    "custo_resultado": ([["custo por resultados"], ["custo por resultado"],
                         ["cost per result"]], []),
    "alcance":    ([["alcance"], ["reach"]], ["custo", "cost"]),
    "impressoes": ([["impressoes"], ["impressions"]], ["cpm", "custo", "cost"]),
    "frequencia": ([["frequencia"], ["frequency"]], []),
    "cpm":        ([["cpm"]], []),
    "conversas":  ([["conversas por mensagem"],
                    ["messaging conversations started"]], ["custo", "cost"]),
    "custo_conversa": ([["custo por conversa"], ["cost per messaging"]], []),
    "novos_contatos": ([["novos contatos"], ["new messaging contacts"]],
                       ["custo", "cost"]),
    # Opcionais e quase sempre vazias no export real. Ficam mapeadas para não
    # serem confundidas com as principais, não porque a análise as use.
    "resultados_iniciais": ([["resultados (iniciais)"],
                             ["resultados iniciais"]], []),
    "indicador_inicial":   ([["indicador de resultados (inicial)"]], []),
}

# As colunas sem as quais esta análise não existe. `Resultados (iniciais)` e
# `Indicador de resultados (inicial)` ficam **de fora de propósito**: vêm
# vazias no export de referência, e exigi-las recusaria o arquivo certo.
COLUNAS_ESSENCIAIS = ("resultados", "custo_resultado", "alcance", "impressoes",
                      "frequencia", "cpm", "conversas", "custo_conversa",
                      "novos_contatos")

# O nome que o operador procura no personalizador de colunas, para a mensagem
# de erro poder apontar a coluna que falta em vez de dizer "arquivo inválido".
ROTULOS_ESSENCIAIS = {
    "resultados": "Resultados",
    "custo_resultado": "Custo por resultados",
    "alcance": "Alcance",
    "impressoes": "Impressões",
    "frequencia": "Frequência",
    "cpm": "CPM (custo por 1.000 impressões)",
    "conversas": "Conversas por mensagem iniciadas",
    "custo_conversa": "Custo por conversa por mensagem iniciada",
    "novos_contatos": "Novos contatos de mensagem",
}

# Campos que viram número. Fora daqui ficam nome, veiculação, indicador e as
# duas datas — texto, e texto é o que a sessão sabe serializar.
_NUMERICOS = ("resultados", "custo_resultado", "alcance", "impressoes",
              "frequencia", "cpm", "conversas", "custo_conversa",
              "novos_contatos", "resultados_iniciais")

# Veiculações que contam como conjunto no ar. Mesma lista do `parser_verba`,
# repetida aqui e não importada: são dois presets diferentes e um pode ganhar
# um valor que o outro não tem. O valor cru vai para a tela de qualquer jeito.
VEICULACOES_ATIVAS = frozenset((
    "active", "ativa", "ativo", "ativas", "ativos",
    "veiculando", "em veiculacao", "delivering", "publicado",
))


class ErroDePreset(ValueError):
    """Arquivo lido, colunas erradas — com a lista do que faltou.

    Separado de `ValueError` para a view poder mostrar as colunas que faltam
    numa lista, em vez de despejar a frase inteira num parágrafo só.
    """

    def __init__(self, mensagem, faltando=()):
        super().__init__(mensagem)
        self.faltando = list(faltando)


def ler_planilha_desempenho(arquivo):
    """As linhas do export, já normalizadas.

    Levanta `ErroDePreset` quando o cabeçalho não é do preset `DESEMPENHO` ou
    quando falta coluna essencial, e `ValueError` quando o arquivo não tem
    linha de dados nenhuma.
    """
    wb = load_workbook(arquivo, data_only=True, read_only=True)
    ws = wb.active
    brutas = list(ws.iter_rows(values_only=True))
    wb.close()
    if not brutas:
        raise ValueError("A planilha está vazia.")

    # O cabeçalho nem sempre é a primeira linha: exports com filtro aplicado
    # trazem uma linha de contexto antes.
    cabecalho, mapa = None, {}
    for i, linha in enumerate(brutas[:10]):
        m = _mapear_colunas(linha, _COLUNAS_DESEMPENHO)
        if "resultados" in m and "impressoes" in m:
            cabecalho, mapa = i, m
            break
    if cabecalho is None:
        raise ErroDePreset(
            "Não foi possível reconhecer as colunas de desempenho nesta "
            "planilha. Aplique a predefinição DESEMPENHO em Colunas → "
            "Personalizar colunas antes de exportar.")

    faltando = [ROTULOS_ESSENCIAIS[c] for c in COLUNAS_ESSENCIAIS
                if c not in mapa]
    if faltando:
        raise ErroDePreset(
            "O arquivo foi lido, mas não tem todas as colunas da predefinição "
            "DESEMPENHO. Reexporte com a predefinição aplicada em Colunas → "
            "Personalizar colunas.", faltando)

    linhas = []
    for bruta in brutas[cabecalho + 1:]:
        if bruta is None or all(c is None or str(c).strip() == ""
                                for c in bruta):
            continue
        reg = {}
        for chave, idx in mapa.items():
            valor = bruta[idx] if idx < len(bruta) else None
            if isinstance(valor, date):
                # A sessão serializa em JSON e não sabe gravar `date`.
                valor = valor.strftime("%Y-%m-%d")
            if chave in _NUMERICOS:
                valor = _to_float(valor)
            elif valor is not None:
                valor = str(valor).strip() or None
            reg[chave] = valor
        # Linha de total do próprio export — o Meta a escreve no rodapé quando
        # o filtro está ligado, e somá-la dobraria a conta inteira.
        nome_ref = _norm(reg.get("campanha") or reg.get("conjunto"))
        if nome_ref.startswith(("total", "resultados de")):
            continue
        linhas.append(reg)

    if not linhas:
        raise ValueError("Nenhuma linha de dados encontrada na planilha.")
    return linhas


def ler_arquivo_desempenho(arquivo):
    """`(linhas, erro, faltando)` — a leitura com o erro já em português.

    A view não trata exceção: ou recebe as linhas, ou recebe a frase pronta e
    a lista de colunas que faltam para mostrar ao lado dela.
    """
    try:
        return ler_planilha_desempenho(arquivo), None, []
    except ErroDePreset as e:
        return None, f'Arquivo "{arquivo.name}": {e}', e.faltando
    except ValueError as e:
        return None, f'Arquivo "{arquivo.name}": {e}', []
    except Exception:
        # O engano provável não é arquivo corrompido — é o export do preset
        # errado. Os três saem do mesmo Gerenciador, na mesma semana, para o
        # mesmo cliente, e todos abrem sem reclamar.
        return None, (
            f'Não foi possível ler "{arquivo.name}". Confira se é o .xlsx '
            "exportado do Gerenciador de Anúncios com a predefinição "
            "DESEMPENHO — os exports dos presets VERBA e RASTREAMENTO não "
            "servem aqui."), []


def ativa(veiculacao):
    """O conjunto estava no ar no período?

    Tudo que não está em `VEICULACOES_ATIVAS` conta como fora, inclusive
    "em análise" e "programada". O valor cru vai para a tabela de conferência:
    o operador precisa poder discordar do que o app decidiu.
    """
    return _norm(veiculacao) in VEICULACOES_ATIVAS


def periodo_do_relatorio(linhas):
    """`(inicio, termino)` como texto ISO, lidos das próprias linhas.

    As duas datas são iguais em todas as linhas de um mesmo export — é o
    recorte escolhido no seletor do Gerenciador, não uma data por conjunto.
    Lê da primeira linha que as tiver, porque um conjunto criado no meio do
    período não muda o recorte do relatório.
    """
    inicio = termino = None
    for linha in linhas:
        inicio = inicio or linha.get("inicio")
        termino = termino or linha.get("termino")
        if inicio and termino:
            break
    return inicio, termino
